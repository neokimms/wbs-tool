#!/usr/bin/env python3
from __future__ import annotations

from datetime import date
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "레벨",
    "WBS 코드",
    "상위 WBS 코드",
    "작업명",
    "유형",
    "담당",
    "가중치",
    "시작일",
    "종료일",
    "산출물 유형",
    "검수 여부",
    "진척 산식",
    "비고",
]

ROWS = [
    [1, "SI", "", "차세대 업무포털 구축", "프로젝트", "PMO", 100, date(2026, 6, 2), date(2026, 12, 18), "프로젝트 계획서", "Y", "하위 단계 가중치 합산", "Demo E2E root"],
    [2, "SI.1", "SI", "착수", "단계", "PMO", 5, date(2026, 6, 2), date(2026, 6, 12), "착수계획서", "Y", "작업 완료율 x 가중치", ""],
    [3, "SI.1.1", "SI.1", "프로젝트 킥오프", "작업", "PMO", 2, date(2026, 6, 2), date(2026, 6, 5), "회의록", "N", "완료 여부", ""],
    [3, "SI.1.2", "SI.1", "통합 수행계획 수립", "산출물", "PMO", 3, date(2026, 6, 6), date(2026, 6, 12), "수행계획서", "Y", "검토 승인", ""],
    [2, "SI.2", "SI", "분석", "단계", "BA Lead", 15, date(2026, 6, 15), date(2026, 7, 10), "요구사항 정의서", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.2.1", "SI.2", "현행 업무 분석", "작업", "BA Lead", 5, date(2026, 6, 15), date(2026, 6, 24), "현행 분석서", "Y", "검토 승인", ""],
    [3, "SI.2.2", "SI.2", "요구사항 정의", "산출물", "BA Lead", 7, date(2026, 6, 25), date(2026, 7, 4), "요구사항 정의서", "Y", "검토 승인", ""],
    [3, "SI.2.3", "SI.2", "분석 단계 승인", "마일스톤", "PMO", 3, date(2026, 7, 8), date(2026, 7, 10), "승인 체크리스트", "Y", "승인 완료", ""],
    [2, "SI.3", "SI", "설계", "단계", "Architect", 20, date(2026, 7, 13), date(2026, 8, 14), "설계서", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.3.1", "SI.3", "아키텍처 설계", "산출물", "Architect", 8, date(2026, 7, 13), date(2026, 7, 24), "아키텍처 설계서", "Y", "검토 승인", ""],
    [3, "SI.3.2", "SI.3", "화면/인터페이스 설계", "산출물", "UX Lead", 7, date(2026, 7, 27), date(2026, 8, 7), "화면설계서", "Y", "검토 승인", ""],
    [3, "SI.3.3", "SI.3", "상세 설계 승인", "마일스톤", "PMO", 5, date(2026, 8, 10), date(2026, 8, 14), "상세설계 승인서", "Y", "승인 완료", ""],
    [2, "SI.4", "SI", "개발", "단계", "Dev Lead", 25, date(2026, 8, 17), date(2026, 10, 16), "소스/단위테스트 결과", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.4.1", "SI.4", "공통 프레임워크 구성", "작업", "Dev Lead", 7, date(2026, 8, 17), date(2026, 8, 28), "공통 모듈", "N", "작업 완료율 x 가중치", ""],
    [3, "SI.4.2", "SI.4", "업무 기능 개발", "작업", "Dev Team", 13, date(2026, 8, 31), date(2026, 9, 30), "업무 기능", "N", "작업 완료율 x 가중치", ""],
    [3, "SI.4.3", "SI.4", "단위 테스트", "작업", "QA Lead", 5, date(2026, 10, 1), date(2026, 10, 16), "단위테스트 결과서", "Y", "검토 승인", ""],
    [2, "SI.5", "SI", "테스트", "단계", "QA Lead", 20, date(2026, 10, 19), date(2026, 11, 20), "통합테스트 결과", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.5.1", "SI.5", "통합 테스트", "작업", "QA Lead", 8, date(2026, 10, 19), date(2026, 10, 30), "통합테스트 시나리오", "Y", "검토 승인", ""],
    [3, "SI.5.2", "SI.5", "사용자 인수 테스트", "작업", "Business Owner", 8, date(2026, 11, 2), date(2026, 11, 13), "UAT 결과서", "Y", "검토 승인", ""],
    [3, "SI.5.3", "SI.5", "결함 조치 및 회귀 테스트", "작업", "QA Lead", 4, date(2026, 11, 16), date(2026, 11, 20), "결함 조치 목록", "Y", "완료율", ""],
    [2, "SI.6", "SI", "전환", "단계", "Infra Lead", 10, date(2026, 11, 23), date(2026, 12, 4), "전환 계획/결과", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.6.1", "SI.6", "전환 리허설", "작업", "Infra Lead", 4, date(2026, 11, 23), date(2026, 11, 27), "전환 리허설 결과", "Y", "검토 승인", ""],
    [3, "SI.6.2", "SI.6", "운영 전환", "마일스톤", "PMO", 6, date(2026, 11, 30), date(2026, 12, 4), "전환 승인서", "Y", "승인 완료", ""],
    [2, "SI.7", "SI", "안정화", "단계", "Operations", 5, date(2026, 12, 7), date(2026, 12, 18), "안정화 보고서", "Y", "하위 단계 가중치 합산", ""],
    [3, "SI.7.1", "SI.7", "초기 운영 지원", "작업", "Operations", 3, date(2026, 12, 7), date(2026, 12, 14), "운영 지원 내역", "N", "작업 완료율 x 가중치", ""],
    [3, "SI.7.2", "SI.7", "안정화 종료 보고", "산출물", "PMO", 2, date(2026, 12, 15), date(2026, 12, 18), "안정화 종료 보고서", "Y", "검토 승인", ""],
]


def build_workbook(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "WBS"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1D1D1F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8DCE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.append(HEADERS)
    for row in ROWS:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[7].number_format = "yyyy-mm-dd"
        row[8].number_format = "yyyy-mm-dd"

    widths = [8, 18, 18, 34, 14, 18, 10, 14, 14, 22, 12, 24, 28]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    worksheet.auto_filter.ref = worksheet.dimensions

    item_type_validation = DataValidation(
        type="list",
        formula1='"프로젝트,단계,산출물,작업,마일스톤,리스크,이슈,변경요청"',
        allow_blank=False,
    )
    inspection_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    worksheet.add_data_validation(item_type_validation)
    worksheet.add_data_validation(inspection_validation)
    item_type_validation.add("E2:E1000")
    inspection_validation.add("K2:K1000")

    guide = workbook.create_sheet("Guide")
    guide.append(["Purpose", "WBS Platform E2E demo upload workbook"])
    guide.append(["Template key", "si-standard"])
    guide.append(["Flow", "Preview import -> apply to project WBS -> auto approval -> baseline -> sync preflight"])
    guide.append(["Note", "이 파일은 scripts/demo-e2e.sh에서 자동 생성됩니다."])
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 96
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(output_path)


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: generate-demo-wbs-workbook.py <output.xlsx>", file=sys.stderr)
        return 2
    output_path = Path(sys.argv[1]).resolve()
    build_workbook(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
