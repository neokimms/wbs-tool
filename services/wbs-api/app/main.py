from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from email.message import EmailMessage
from email.utils import formatdate
from io import BytesIO
import base64
import asyncio
import json
import os
from pathlib import Path
import re
import secrets
import smtplib
from typing import Any
from uuid import UUID

import asyncpg
import httpx
from email.mime.text import MIMEText
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, PlainTextResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field

try:
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    from apscheduler.triggers.cron import CronTrigger
except ImportError:  # pragma: no cover - allows static checks before container deps are installed
    AsyncIOScheduler = None
    CronTrigger = None


DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://wbs:wbs_dev_password@localhost:5432/wbs_platform",
)
OPENPROJECT_BASE_URL = os.getenv("OPENPROJECT_BASE_URL", "http://localhost:8080")
OPENPROJECT_HOST_HEADER = os.getenv("OPENPROJECT_HOST_HEADER", "")
OPENPROJECT_SYNC_ENABLED = os.getenv("OPENPROJECT_SYNC_ENABLED", "false").lower() in {"1", "true", "yes", "on"}
OPENPROJECT_API_TOKEN = os.getenv("OPENPROJECT_API_TOKEN", "")
OPENPROJECT_AUTH_MODE = os.getenv("OPENPROJECT_AUTH_MODE", "bearer").lower()
OPENPROJECT_DEFAULT_TYPE_ID = os.getenv("OPENPROJECT_DEFAULT_TYPE_ID", "")
OPENPROJECT_TYPE_MAP_JSON = os.getenv("OPENPROJECT_TYPE_MAP_JSON", "{}")
OPENPROJECT_SYNC_PARENT_LINKS = os.getenv("OPENPROJECT_SYNC_PARENT_LINKS", "true").lower() in {"1", "true", "yes", "on"}
PM_ENGINE_ADAPTER = os.getenv("PM_ENGINE_ADAPTER", "openproject").strip().lower()
PORTAL_ORIGIN = os.getenv("PORTAL_ORIGIN", "http://localhost:3010")
ALLOW_FILE_ORIGIN = os.getenv("WBS_ALLOW_FILE_ORIGIN", "true").lower() in {"1", "true", "yes", "on"}
BACKUP_DIR = Path(os.getenv("BACKUP_DIR", "/tmp/wbs-backups"))
REPORT_OUTPUT_DIR = Path(os.getenv("WBS_REPORT_OUTPUT_DIR", "/app/outputs/reports"))
MIGRATION_PATH = Path(__file__).resolve().parent.parent / "migrations" / "001_init.sql"
MIGRATION_PATHS = [
    Path(__file__).resolve().parent.parent / "migrations" / "001_init.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "002_risks_issues.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "003_notifications.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "004_multitenancy.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "005_change_requests.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "006_wbs_tenant_recovery.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "007_user_groups.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "008_agile_wbs.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "009_agile_hybrid_samples.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "010_project_operation_policy.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "011_agile_wbs_sync.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "012_project_members.sql",
    Path(__file__).resolve().parent.parent / "migrations" / "013_announcements.sql",
]
RUN_MIGRATIONS_ON_STARTUP = os.getenv("WBS_RUN_MIGRATIONS_ON_STARTUP", "true").lower() in {"1", "true", "yes", "on"}
SESSION_TTL_HOURS = int(os.getenv("WBS_SESSION_TTL_HOURS", "12"))
LOGIN_FAILURE_LIMIT = int(os.getenv("WBS_LOGIN_FAILURE_LIMIT", "5"))
LOGIN_LOCK_MINUTES = int(os.getenv("WBS_LOGIN_LOCK_MINUTES", "15"))
ENABLE_LOGIN_ALIASES = os.getenv("WBS_ENABLE_LOGIN_ALIASES", "false").lower() in {"1", "true", "yes", "on"}
MULTITENANCY_ENABLED = os.getenv("MULTITENANCY_ENABLED", "false").lower() in {"1", "true", "yes", "on"}
DEFAULT_TENANT_ID = os.getenv("DEFAULT_TENANT_ID", "default")
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
# P2-06: SSO/LDAP
AUTH_BACKEND = os.getenv("AUTH_BACKEND", "local").strip().lower()  # local | ldap
LDAP_SERVER = os.getenv("LDAP_SERVER", "")
LDAP_PORT = int(os.getenv("LDAP_PORT", "389"))
LDAP_USE_SSL = os.getenv("LDAP_USE_SSL", "false").lower() in {"1", "true", "yes", "on"}
LDAP_BIND_DN = os.getenv("LDAP_BIND_DN", "")
LDAP_BIND_PASSWORD = os.getenv("LDAP_BIND_PASSWORD", "")
LDAP_BASE_DN = os.getenv("LDAP_BASE_DN", "")
LDAP_USER_FILTER = os.getenv("LDAP_USER_FILTER", "(mail={email})")
LDAP_ATTR_EMAIL = os.getenv("LDAP_ATTR_EMAIL", "mail")
LDAP_ATTR_NAME = os.getenv("LDAP_ATTR_NAME", "cn")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
NOTIFY_FROM_EMAIL = os.getenv("NOTIFY_FROM_EMAIL", SMTP_USER)
REPORT_SCHEDULER_ENABLED = os.getenv("WBS_REPORT_SCHEDULER_ENABLED", "true").lower() in {"1", "true", "yes", "on"}
REPORT_DEFAULT_TIMEZONE = os.getenv("WBS_REPORT_TIMEZONE", "Asia/Seoul")
REPORT_SMTP_HOST = os.getenv("WBS_SMTP_HOST", SMTP_HOST).strip()
REPORT_SMTP_PORT = int(os.getenv("WBS_SMTP_PORT", str(SMTP_PORT)))
REPORT_SMTP_USERNAME = os.getenv("WBS_SMTP_USERNAME", SMTP_USER).strip()
REPORT_SMTP_PASSWORD = os.getenv("WBS_SMTP_PASSWORD", SMTP_PASSWORD)
REPORT_SMTP_FROM = os.getenv(
    "WBS_SMTP_FROM",
    NOTIFY_FROM_EMAIL or REPORT_SMTP_USERNAME or "wbs-platform@localhost",
).strip()
REPORT_SMTP_USE_TLS = os.getenv("WBS_SMTP_USE_TLS", "true").lower() in {"1", "true", "yes", "on"}
REPORT_SMTP_USE_SSL = os.getenv("WBS_SMTP_USE_SSL", "false").lower() in {"1", "true", "yes", "on"}
AUDIT_RETENTION_DAYS = int(os.getenv("WBS_AUDIT_RETENTION_DAYS", "365"))
MAX_EXCEL_UPLOAD_BYTES = 8 * 1024 * 1024
MAX_WORK_ITEM_ATTACHMENT_BYTES = int(os.getenv("WBS_WORK_ITEM_ATTACHMENT_MAX_BYTES", str(2 * 1024 * 1024)))
EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ALLOWED_USER_ROLES = {"admin", "pmo", "viewer"}
ALLOWED_USER_STATUSES = {"Active", "Suspended"}
MUTATING_ROLES = {"admin", "pmo"}
ADMIN_ROLES = {"admin"}
WORK_ITEM_STATUSES = {"대기", "진행중", "완료", "지연", "보류"}
WORK_ITEM_PRIORITIES = {"높음", "보통", "낮음"}
PROJECT_DELIVERY_MODES = {"waterfall", "agile", "hybrid"}
AGILE_SPRINT_STATUSES = {"Planning", "Active", "Review", "Retrospective", "Closed"}
AGILE_ITEM_TYPES = {"Epic", "Story", "Task", "Spike", "Bug"}
AGILE_ITEM_STATUSES = {"Backlog", "Ready", "In Progress", "Review", "Done"}
AGILE_ITEM_PRIORITIES = {"Must", "Should", "Could", "Wont"}
STORY_POINT_MODES = {"numeric", "fibonacci"}
SPRINT_LENGTH_POLICIES = {"custom", "fixed_1w", "fixed_2w", "fixed_4w"}
DOD_MANAGEMENT_MODES = {"organization", "team"}
SPRINT_LENGTH_DAYS = {"fixed_1w": 7, "fixed_2w": 14, "fixed_4w": 28}
DEFAULT_PROJECT_OPERATION_POLICY = {
    "default_delivery_mode": "waterfall",
    "story_point_mode": "numeric",
    "fibonacci_points": [1, 2, 3, 5, 8, 13],
    "sprint_length_policy": "custom",
    "dod_management": "team",
    "default_dod_items": [],
    "openproject_sprint_version_sync": False,
    "metadata": {},
}
DEFAULT_USER_GROUP_NAME = "기본 그룹"
PROJECT_WORKFLOW_TRANSITIONS = {
    "Draft": {"Review", "Approved", "Closed"},
    "Review": {"Approved", "Rejected", "Closed"},
    "Rejected": {"Draft", "Review", "Closed"},
    "Approved": {"Synced", "Closed"},
    "Synced": {"Closed"},
    "Closed": set(),
}
APPROVAL_ALLOWED_PROJECT_STATUSES = {"Draft", "Rejected", "Review"}
PROJECT_WBS_IMPORT_ALLOWED_STATUSES = {"Draft", "Rejected", "Review"}
ACTUAL_SYNC_REQUIRED_STATUS = "Approved"
DEFAULT_LOGIN_ALIASES = {
    "admin": {"email": "admin@wbs.local", "password": "admin"},
    "pmo": {"email": "pmo@wbs.local", "password": "pmo"},
    "viewer": {"email": "viewer@wbs.local", "password": "viewer"},
}
USER_SELECT = """
id, email, display_name, role, status, tenant_id, group_id, failed_login_count, locked_until,
must_change_password, last_login_at, password_changed_at, created_at
"""
USER_SELECT_U = """
u.id, u.email, u.display_name, u.role, u.status, u.tenant_id, u.group_id, u.failed_login_count,
u.locked_until, u.must_change_password, u.last_login_at, u.password_changed_at,
u.created_at
"""
USER_SELECT_G = """
u.id, u.email, u.display_name, u.role, u.status, u.tenant_id, u.group_id, u.failed_login_count,
u.locked_until, u.must_change_password, u.last_login_at, u.password_changed_at,
u.created_at, g.name AS group_name, g.status AS group_status
"""
SETTING_SELECT = "key, label, category, description, value, is_sensitive, updated_by, created_at, updated_at"
AUDIT_SELECT = """
id, actor_user_id, actor_email, actor_role, event_type, entity_type, entity_id,
summary, metadata, created_at
"""
IMPORT_JOB_RETURNING = """
id, source_file, template_key, template_name, project_type, description, status,
total_rows, accepted_rows, rejected_rows, errors, warnings, preview_rows,
diff_rows, template_version, applied_at, created_at
"""
APPROVAL_SELECT = """
a.id, a.project_id, p.name AS project_name, p.template_key, a.title,
a.request_type, a.status, a.requester, a.reviewer, a.due_date,
a.decision_comment, a.metadata, a.created_at, a.decided_at, a.updated_at
"""
SYNC_RUN_SELECT = """
s.id, s.project_id, p.name AS project_name, p.template_key, s.mode,
s.status, s.actor, s.engine, s.dry_run, s.create_work_packages,
s.validate_payloads, s.total_rows, s.pending_work_packages,
s.synced_work_packages, s.created_work_packages, s.openproject_project_id,
s.metadata, s.error, s.started_at, s.completed_at
"""
BASELINE_SELECT = """
b.id, b.project_id, p.name AS project_name, b.approval_id, b.version,
b.status, b.template_key, b.template_name, b.item_count, b.total_weight,
b.snapshot_rows, b.metadata, b.locked_at, b.created_at
"""
REPORT_SCHEDULE_SELECT = """
id, key, name, report_type, enabled, timezone, day_of_week, hour, minute,
recipients, email_subject, email_body, metadata, last_run_at, next_run_at,
created_at, updated_at
"""
REPORT_RUN_SELECT = """
id, schedule_id, schedule_key, report_type, status, period_start,
period_end, recipient_count, artifact_path, delivery_status,
delivery_detail, error, metadata, triggered_by, started_at, completed_at
"""
AGILE_SPRINT_SELECT = """
id, project_id, tenant_id, name, goal, status, start_date, end_date,
capacity_points, metadata, created_at, updated_at
"""
AGILE_ITEM_SELECT = """
id, project_id, tenant_id, sprint_id, parent_id, wbs_code, item_type,
title, description, story_points, priority, status, assignee, reviewer,
acceptance_criteria, sort_order, metadata, created_at, updated_at
"""
PROJECT_OPERATION_POLICY_SELECT = """
tenant_id, default_delivery_mode, story_point_mode, fibonacci_points,
sprint_length_policy, dod_management, default_dod_items,
openproject_sprint_version_sync, metadata, updated_by, created_at, updated_at
"""

EXCEL_COLUMNS = [
    ("level", "레벨"),
    ("code", "WBS 코드"),
    ("parent_code", "상위 WBS 코드"),
    ("name", "작업명"),
    ("item_type", "유형"),
    ("owner", "담당"),
    ("weight", "가중치"),
    ("start_date", "시작일"),
    ("finish_date", "종료일"),
    ("deliverable_type", "산출물 유형"),
    ("inspection_required", "검수 여부"),
    ("progress_formula", "진척 산식"),
    ("notes", "비고"),
]

HEADER_ALIASES = {
    "레벨": "level",
    "level": "level",
    "wbs코드": "code",
    "wbscode": "code",
    "code": "code",
    "상위wbs코드": "parent_code",
    "상위코드": "parent_code",
    "parentcode": "parent_code",
    "parentwbscode": "parent_code",
    "작업명": "name",
    "작업": "name",
    "name": "name",
    "taskname": "name",
    "subject": "name",
    "유형": "item_type",
    "type": "item_type",
    "itemtype": "item_type",
    "담당": "owner",
    "담당자": "owner",
    "owner": "owner",
    "assignee": "owner",
    "가중치": "weight",
    "weight": "weight",
    "시작일": "start_date",
    "start": "start_date",
    "startdate": "start_date",
    "종료일": "finish_date",
    "finish": "finish_date",
    "finishdate": "finish_date",
    "duedate": "finish_date",
    "산출물유형": "deliverable_type",
    "deliverabletype": "deliverable_type",
    "검수여부": "inspection_required",
    "inspectionrequired": "inspection_required",
    "approvalrequired": "inspection_required",
    "진척산식": "progress_formula",
    "progressformula": "progress_formula",
    "비고": "notes",
    "notes": "notes",
}


async def init_connection(connection: asyncpg.Connection) -> None:
    await connection.set_type_codec(
        "json",
        encoder=json.dumps,
        decoder=json.loads,
        schema="pg_catalog",
    )
    await connection.set_type_codec(
        "jsonb",
        encoder=json.dumps,
        decoder=json.loads,
        schema="pg_catalog",
    )


@asynccontextmanager
async def lifespan(app: FastAPI):
    pool = await asyncpg.create_pool(
        DATABASE_URL,
        min_size=1,
        max_size=8,
        init=init_connection,
    )
    if RUN_MIGRATIONS_ON_STARTUP:
        async with pool.acquire() as connection:
            for mp in MIGRATION_PATHS:
                if mp.exists():
                    await connection.execute(mp.read_text(encoding="utf-8"))
    app.state.pool = pool
    app.state.report_scheduler = None
    if REPORT_SCHEDULER_ENABLED and AsyncIOScheduler and CronTrigger:
        scheduler = AsyncIOScheduler(timezone=REPORT_DEFAULT_TIMEZONE)
        app.state.report_scheduler = scheduler
        scheduler.start()
        await refresh_report_scheduler(app)
    yield
    scheduler = getattr(app.state, "report_scheduler", None)
    if scheduler:
        scheduler.shutdown(wait=False)
    await pool.close()


app = FastAPI(
    title="WBS Platform Extension API",
    version="0.1.0",
    lifespan=lifespan,
)

allowed_origins = [PORTAL_ORIGIN, "http://localhost:3010", "http://127.0.0.1:3010"]
if ALLOW_FILE_ORIGIN:
    allowed_origins.append("null")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ProjectCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=120)
    template_key: str = Field(..., min_length=2, max_length=80)
    owner: str = Field("PMO", min_length=2, max_length=80)
    start_date: date | None = None
    delivery_mode: str = Field("waterfall", pattern="^(waterfall|agile|hybrid)$")
    # 선택 상세 필드 (metadata에 저장)
    end_date: date | None = None
    description: str | None = Field(None, max_length=500)
    client_name: str | None = Field(None, max_length=120)
    budget: str | None = Field(None, max_length=80)
    project_manager: str | None = Field(None, max_length=80)


class WbsImportRow(BaseModel):
    level: int | None = Field(None, ge=1, le=20)
    code: str | None = Field(None, min_length=1, max_length=40)
    name: str = Field(..., min_length=1, max_length=160)
    parent_code: str | None = Field(None, max_length=40)
    item_type: str = Field("작업", min_length=1, max_length=40)
    owner: str | None = Field(None, max_length=80)
    weight: float | None = Field(None, ge=0, le=100)
    start_date: date | None = None
    finish_date: date | None = None
    deliverable_type: str | None = Field(None, max_length=80)
    inspection_required: bool = False
    progress_formula: str | None = Field(None, max_length=200)
    notes: str | None = Field(None, max_length=500)
    metadata: dict[str, Any] = Field(default_factory=dict)


class WbsImportValidation(BaseModel):
    source_file: str = Field("wbs-upload.xlsx", min_length=1, max_length=160)
    rows: list[WbsImportRow]


class ApprovalCreate(BaseModel):
    project_id: UUID
    title: str | None = Field(None, max_length=160)
    request_type: str = Field("WBS Baseline", min_length=1, max_length=80)
    requester: str = Field("PMO", min_length=1, max_length=80)
    reviewer: str | None = Field("PMO Lead", max_length=80)
    due_date: date | None = None
    auto_approve_internal: bool = True
    metadata: dict[str, Any] = Field(default_factory=dict)


class ApprovalDecision(BaseModel):
    reviewer: str = Field("PMO Lead", min_length=1, max_length=80)
    comment: str | None = Field(None, max_length=500)


class ProjectSyncRequest(BaseModel):
    dry_run: bool = True
    create_work_packages: bool = True
    force_project_create: bool = False
    validate_payloads: bool = True
    actor: str = Field("PMO", min_length=1, max_length=80)


class ProjectStatusUpdate(BaseModel):
    status: str = Field(..., min_length=4, max_length=20)
    comment: str | None = Field(None, max_length=500)


class ProjectDeliveryModeUpdate(BaseModel):
    delivery_mode: str = Field(..., pattern="^(waterfall|agile|hybrid)$")
    comment: str | None = Field(None, max_length=500)


class ProjectMemberCreate(BaseModel):
    user_id: UUID
    project_role: str = Field(..., min_length=3, max_length=20)


class ProjectMemberUpdate(BaseModel):
    project_role: str = Field(..., min_length=3, max_length=20)


class WbsItemsBatch(BaseModel):
    rows: list[WbsImportRow]
    source: str = Field("portal-editor", max_length=80)


class WorkPackageDateUpdate(BaseModel):
    start_date: date | None = None
    finish_date: date | None = None


class WorkItemUpdate(BaseModel):
    name: str | None = Field(None, min_length=1, max_length=160)
    owner: str | None = Field(None, max_length=80)
    status: str | None = Field(None, pattern="^(대기|진행중|완료|지연|보류)$")
    progress: int | None = Field(None, ge=0, le=100)
    priority: str | None = Field(None, pattern="^(높음|보통|낮음)$")
    start_date: date | None = None
    finish_date: date | None = None
    reviewer: str | None = Field(None, max_length=80)
    approver: str | None = Field(None, max_length=80)
    team: str | None = Field(None, max_length=80)
    effort: float | None = Field(None, ge=0, le=10000)
    comment: str | None = Field(None, max_length=1000)
    attachment_name: str | None = Field(None, max_length=160)
    attachment_url: str | None = Field(None, max_length=1000)


class AgileSprintCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    goal: str | None = Field("", max_length=1000)
    status: str = Field("Planning", pattern="^(Planning|Active|Review|Retrospective|Closed)$")
    start_date: date | None = None
    end_date: date | None = None
    capacity_points: float | None = Field(0, ge=0, le=10000)
    metadata: dict[str, Any] = Field(default_factory=dict)


class AgileSprintUpdate(BaseModel):
    name: str | None = Field(None, min_length=1, max_length=120)
    goal: str | None = Field(None, max_length=1000)
    status: str | None = Field(None, pattern="^(Planning|Active|Review|Retrospective|Closed)$")
    start_date: date | None = None
    end_date: date | None = None
    capacity_points: float | None = Field(None, ge=0, le=10000)
    metadata: dict[str, Any] | None = None


class AgileItemCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    item_type: str = Field("Story", pattern="^(Epic|Story|Task|Spike|Bug)$")
    description: str | None = Field("", max_length=4000)
    story_points: float | None = Field(0, ge=0, le=10000)
    priority: str = Field("Should", pattern="^(Must|Should|Could|Wont)$")
    status: str = Field("Backlog", pattern="^(Backlog|Ready|In Progress|Review|Done)$")
    sprint_id: UUID | None = None
    parent_id: UUID | None = None
    wbs_code: str | None = Field(None, max_length=50)
    assignee: str | None = Field(None, max_length=120)
    reviewer: str | None = Field(None, max_length=120)
    acceptance_criteria: str | None = Field("", max_length=4000)
    metadata: dict[str, Any] = Field(default_factory=dict)


class AgileItemUpdate(BaseModel):
    title: str | None = Field(None, min_length=1, max_length=200)
    item_type: str | None = Field(None, pattern="^(Epic|Story|Task|Spike|Bug)$")
    description: str | None = Field(None, max_length=4000)
    story_points: float | None = Field(None, ge=0, le=10000)
    priority: str | None = Field(None, pattern="^(Must|Should|Could|Wont)$")
    status: str | None = Field(None, pattern="^(Backlog|Ready|In Progress|Review|Done)$")
    sprint_id: UUID | None = None
    parent_id: UUID | None = None
    wbs_code: str | None = Field(None, max_length=50)
    assignee: str | None = Field(None, max_length=120)
    reviewer: str | None = Field(None, max_length=120)
    acceptance_criteria: str | None = Field(None, max_length=4000)
    metadata: dict[str, Any] | None = None


class LoginRequest(BaseModel):
    email: str = Field(..., min_length=3, max_length=160)
    password: str = Field(..., min_length=1, max_length=200)


# P2-02: 리스크 · 이슈 모델
class RiskCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: str = Field("", max_length=2000)
    severity: str = Field("보통", pattern="^(높음|보통|낮음)$")
    likelihood: str = Field("보통", pattern="^(높음|보통|낮음)$")
    owner: str = Field("PMO", max_length=80)
    mitigation: str = Field("", max_length=2000)
    wbs_code: str | None = Field(None, max_length=50)
    due_date: str | None = None


class RiskUpdate(BaseModel):
    title: str | None = Field(None, min_length=1, max_length=200)
    description: str | None = Field(None, max_length=2000)
    severity: str | None = Field(None, pattern="^(높음|보통|낮음)$")
    likelihood: str | None = Field(None, pattern="^(높음|보통|낮음)$")
    status: str | None = Field(None, pattern="^(Open|Mitigated|Closed)$")
    owner: str | None = Field(None, max_length=80)
    mitigation: str | None = Field(None, max_length=2000)
    wbs_code: str | None = Field(None, max_length=50)
    due_date: str | None = None


class IssueCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: str = Field("", max_length=2000)
    priority: str = Field("보통", pattern="^(높음|보통|낮음)$")
    assignee: str = Field("PMO", max_length=80)
    wbs_code: str | None = Field(None, max_length=50)
    due_date: str | None = None


class IssueUpdate(BaseModel):
    title: str | None = Field(None, min_length=1, max_length=200)
    description: str | None = Field(None, max_length=2000)
    priority: str | None = Field(None, pattern="^(높음|보통|낮음)$")
    status: str | None = Field(None, pattern="^(Open|In Progress|Resolved|Closed)$")
    assignee: str | None = Field(None, max_length=80)
    wbs_code: str | None = Field(None, max_length=50)
    due_date: str | None = None


class PasswordChangeRequest(BaseModel):
    current_password: str = Field(..., min_length=1, max_length=200)
    new_password: str = Field(..., min_length=8, max_length=200)


class UserCreate(BaseModel):
    email: str = Field(..., min_length=3, max_length=160)
    display_name: str = Field(..., min_length=2, max_length=120)
    role: str = Field("viewer", min_length=3, max_length=20)
    password: str = Field(..., min_length=8, max_length=200)
    status: str = Field("Active", min_length=5, max_length=20)
    must_change_password: bool = True
    group_id: UUID | None = None


class UserUpdate(BaseModel):
    display_name: str | None = Field(None, min_length=2, max_length=120)
    role: str | None = Field(None, min_length=3, max_length=20)
    password: str | None = Field(None, min_length=8, max_length=200)
    status: str | None = Field(None, min_length=5, max_length=20)
    must_change_password: bool | None = None
    group_id: UUID | None = None


class UserGroupCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    description: str | None = Field("", max_length=500)


class UserGroupUpdate(BaseModel):
    name: str | None = Field(None, min_length=1, max_length=100)
    description: str | None = Field(None, max_length=500)
    status: str | None = Field(None, pattern="^(Active|Suspended)$")


class SettingUpdate(BaseModel):
    value: dict[str, Any] = Field(default_factory=dict)


class ProjectOperationPolicyUpdate(BaseModel):
    default_delivery_mode: str | None = Field(None, pattern="^(waterfall|agile|hybrid)$")
    story_point_mode: str | None = Field(None, pattern="^(numeric|fibonacci)$")
    fibonacci_points: list[float] | None = Field(None, max_length=20)
    sprint_length_policy: str | None = Field(None, pattern="^(custom|fixed_1w|fixed_2w|fixed_4w)$")
    dod_management: str | None = Field(None, pattern="^(organization|team)$")
    default_dod_items: list[str] | None = Field(None, max_length=20)
    openproject_sprint_version_sync: bool | None = None
    metadata: dict[str, Any] | None = None


class ReportScheduleUpdate(BaseModel):
    name: str | None = Field(None, min_length=2, max_length=120)
    enabled: bool | None = None
    timezone: str | None = Field(None, min_length=2, max_length=80)
    day_of_week: int | None = Field(None, ge=0, le=6)
    hour: int | None = Field(None, ge=0, le=23)
    minute: int | None = Field(None, ge=0, le=59)
    recipients: list[str] | None = Field(None, max_length=10)
    email_subject: str | None = Field(None, min_length=1, max_length=200)
    email_body: str | None = Field(None, min_length=1, max_length=2000)
    metadata: dict[str, Any] | None = None


class ReportRunRequest(BaseModel):
    send_email: bool = True
    triggered_by: str = Field("manual", min_length=1, max_length=80)


def normalize_record(record: asyncpg.Record) -> dict[str, Any]:
    data = {key: value for key, value in dict(record).items() if not key.startswith("_")}
    for key, value in data.items():
        if isinstance(value, date):
            data[key] = value.isoformat()
        if isinstance(value, Decimal):
            data[key] = float(value)
        if key in {
            "errors",
            "warnings",
            "metadata",
            "phases",
            "preview_rows",
            "snapshot_rows",
            "error",
            "recipients",
            "delivery_detail",
        } and isinstance(data[key], str):
            try:
                data[key] = json.loads(data[key])
            except json.JSONDecodeError:
                pass
    return data


def normalize_metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            decoded = json.loads(value)
            return decoded if isinstance(decoded, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def get_pool(request: Request) -> asyncpg.Pool:
    return request.app.state.pool


def auth_token_from_request(request: Request) -> str | None:
    authorization = request.headers.get("authorization", "")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        return None
    return token.strip()


def user_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    user = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    return {
        "id": str(user["id"]),
        "email": user["email"],
        "display_name": user["display_name"],
        "role": user["role"],
        "status": user["status"],
        "tenant_id": user.get("tenant_id") or DEFAULT_TENANT_ID,
        "group_id": str(user["group_id"]) if user.get("group_id") else None,
        "group_name": user.get("group_name"),
        "group_status": user.get("group_status"),
        "must_change_password": bool(user.get("must_change_password", False)),
        "last_login_at": user.get("last_login_at"),
        "password_changed_at": user.get("password_changed_at"),
        "created_at": user.get("created_at"),
    }


def managed_user_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    user = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    return {
        **user_response(user),
        "updated_at": user.get("updated_at"),
        "active_sessions": int(user.get("active_sessions") or 0),
        "failed_login_count": int(user.get("failed_login_count") or 0),
        "locked_until": user.get("locked_until"),
    }


def user_group_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    group = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    return {
        "id": str(group["id"]),
        "tenant_id": group.get("tenant_id") or DEFAULT_TENANT_ID,
        "name": group.get("name"),
        "description": group.get("description") or "",
        "status": group.get("status") or "Active",
        "metadata": normalize_metadata(group.get("metadata")),
        "user_count": int(group.get("user_count") or 0),
        "created_at": group.get("created_at"),
        "updated_at": group.get("updated_at"),
    }


async def ensure_default_user_group(connection: asyncpg.Connection, tenant_id: str) -> dict[str, Any]:
    record = await connection.fetchrow(
        """
        SELECT id, tenant_id, name, description, status, metadata, created_at, updated_at,
               0::integer AS user_count
        FROM wbs_user_groups
        WHERE tenant_id = $1 AND lower(name) = lower($2)
        LIMIT 1
        """,
        tenant_id,
        DEFAULT_USER_GROUP_NAME,
    )
    if not record:
        record = await connection.fetchrow(
            """
            INSERT INTO wbs_user_groups (tenant_id, name, description, metadata)
            VALUES ($1, $2, $3, $4::jsonb)
            RETURNING id, tenant_id, name, description, status, metadata, created_at, updated_at,
                      0::integer AS user_count
            """,
            tenant_id,
            DEFAULT_USER_GROUP_NAME,
            "테넌트 기본 소속 그룹",
            {"system": True},
        )
    return user_group_response(record)


async def resolve_user_group(
    connection: asyncpg.Connection,
    tenant_id: str,
    group_id: UUID | str | None,
    *,
    require_active: bool = True,
) -> dict[str, Any]:
    if not group_id:
        return await ensure_default_user_group(connection, tenant_id)
    parsed_id = safe_uuid(group_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid user group id")
    record = await connection.fetchrow(
        """
        SELECT id, tenant_id, name, description, status, metadata, created_at, updated_at,
               0::integer AS user_count
        FROM wbs_user_groups
        WHERE id = $1 AND tenant_id = $2
        """,
        parsed_id,
        tenant_id,
    )
    if not record:
        raise HTTPException(status_code=400, detail="User group is not registered in this tenant")
    group = user_group_response(record)
    if require_active and group["status"] != "Active":
        raise HTTPException(status_code=400, detail="User group is not active")
    return group


async def fetch_user_by_token(connection: asyncpg.Connection, token: str) -> dict[str, Any] | None:
    record = await connection.fetchrow(
        f"""
        SELECT {USER_SELECT_G}
        FROM wbs_user_sessions s
        JOIN wbs_users u ON u.id = s.user_id
        LEFT JOIN wbs_user_groups g
          ON g.id = u.group_id
         AND g.tenant_id = u.tenant_id
        WHERE s.token = $1
          AND s.expires_at > now()
          AND u.status = 'Active'
        """,
        token,
    )
    return user_response(record) if record else None


def require_roles(request: Request, allowed_roles: set[str]) -> dict[str, Any]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user["role"] not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient role")
    return user


def require_mutating_role(request: Request) -> dict[str, Any]:
    return require_roles(request, MUTATING_ROLES)


def require_admin_role(request: Request) -> dict[str, Any]:
    return require_roles(request, ADMIN_ROLES)


async def get_effective_project_role(
    connection: asyncpg.Connection,
    tenant_id: str,
    project_id: UUID,
    user: dict[str, Any] | None,
) -> str | None:
    """프로젝트별 실질 권한.
    - 전사 admin: 멤버십과 무관하게 테넌트 내 모든 프로젝트에 대해 admin 권한
    - 그 외(pmo/viewer): wbs_project_members에 등록된 경우에만 해당 project_role, 비멤버는 None(접근 불가)
    """
    if not user:
        return None
    if user.get("role") == "admin":
        return "admin"
    user_id = safe_uuid(user.get("id"))
    if not user_id:
        return None
    return await connection.fetchval(
        "SELECT project_role FROM wbs_project_members WHERE tenant_id = $1 AND project_id = $2 AND user_id = $3",
        tenant_id, project_id, user_id,
    )


def ensure_project_mutate_role(project: dict[str, Any]) -> None:
    """fetch_tenant_project()가 채운 프로젝트별 실질 권한이 admin/pmo가 아니면 변경 차단."""
    if project.get("_project_role") not in MUTATING_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient role")


def validate_project_status(status: str) -> str:
    normalized_status = status.strip().title()
    if normalized_status not in PROJECT_WORKFLOW_TRANSITIONS:
        raise HTTPException(status_code=400, detail="Invalid project status")
    return normalized_status


def ensure_project_status_allowed(project: dict[str, Any] | asyncpg.Record, allowed_statuses: set[str], action: str) -> None:
    status = project["status"]
    if status not in allowed_statuses:
        raise HTTPException(
            status_code=409,
            detail=f"{action} is not allowed while project status is {status}",
        )


def ensure_project_transition(current_status: str, next_status: str) -> None:
    current = validate_project_status(current_status)
    target = validate_project_status(next_status)
    if target == current:
        return
    if target not in PROJECT_WORKFLOW_TRANSITIONS[current]:
        raise HTTPException(
            status_code=409,
            detail=f"Project status cannot move from {current} to {target}",
        )


def load_login_aliases_config() -> dict[str, dict[str, str]]:
    raw_config = os.getenv("WBS_LOGIN_ALIASES_JSON", "").strip()
    configured_aliases: Any = DEFAULT_LOGIN_ALIASES
    if raw_config:
        try:
            configured_aliases = json.loads(raw_config)
        except json.JSONDecodeError:
            configured_aliases = DEFAULT_LOGIN_ALIASES

    aliases: dict[str, dict[str, str]] = {}
    if not isinstance(configured_aliases, dict):
        return aliases

    for alias, value in configured_aliases.items():
        alias_key = str(alias).strip().lower()
        if not alias_key:
            continue
        if isinstance(value, str):
            email = value.strip().lower()
            alias_password = ""
        elif isinstance(value, dict):
            email = str(value.get("email", "")).strip().lower()
            alias_password = str(value.get("password", ""))
        else:
            continue
        if email:
            aliases[alias_key] = {"email": email, "password": alias_password}
    return aliases


LOGIN_ALIASES = load_login_aliases_config()


def resolve_login_alias(identifier: str) -> tuple[str, dict[str, str] | None]:
    normalized_identifier = identifier.strip().lower()
    if not ENABLE_LOGIN_ALIASES:
        return normalized_identifier, None
    alias = LOGIN_ALIASES.get(normalized_identifier)
    if not alias:
        return normalized_identifier, None
    return alias["email"], alias


def safe_uuid(value: Any) -> UUID | None:
    if not value:
        return None
    try:
        return value if isinstance(value, UUID) else UUID(str(value))
    except ValueError:
        return None


def normalize_delivery_mode(value: Any) -> str:
    mode = str(value or "waterfall").strip().lower()
    if mode not in PROJECT_DELIVERY_MODES:
        raise HTTPException(status_code=400, detail="Invalid project delivery mode")
    return mode


def default_project_operation_policy(tenant_id: str) -> dict[str, Any]:
    return {
        "tenant_id": tenant_id,
        "default_delivery_mode": DEFAULT_PROJECT_OPERATION_POLICY["default_delivery_mode"],
        "story_point_mode": DEFAULT_PROJECT_OPERATION_POLICY["story_point_mode"],
        "fibonacci_points": list(DEFAULT_PROJECT_OPERATION_POLICY["fibonacci_points"]),
        "sprint_length_policy": DEFAULT_PROJECT_OPERATION_POLICY["sprint_length_policy"],
        "dod_management": DEFAULT_PROJECT_OPERATION_POLICY["dod_management"],
        "default_dod_items": list(DEFAULT_PROJECT_OPERATION_POLICY["default_dod_items"]),
        "openproject_sprint_version_sync": DEFAULT_PROJECT_OPERATION_POLICY["openproject_sprint_version_sync"],
        "metadata": dict(DEFAULT_PROJECT_OPERATION_POLICY["metadata"]),
    }


def normalize_number_list(values: Any, fallback: list[float]) -> list[float]:
    if not isinstance(values, list):
        return list(fallback)
    cleaned: list[float] = []
    for value in values:
        try:
            point = float(value)
        except (TypeError, ValueError):
            continue
        if point <= 0 or point > 10000:
            continue
        if not any(abs(existing - point) < 0.001 for existing in cleaned):
            cleaned.append(point)
    return sorted(cleaned) or list(fallback)


def normalize_dod_items(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    cleaned: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in cleaned:
            cleaned.append(text[:200])
        if len(cleaned) >= 20:
            break
    return cleaned


def normalize_project_operation_policy_value(policy: dict[str, Any], tenant_id: str) -> dict[str, Any]:
    base = default_project_operation_policy(tenant_id)
    merged = {**base, **normalize_metadata(policy)}
    merged["default_delivery_mode"] = normalize_delivery_mode(merged.get("default_delivery_mode"))

    story_point_mode = str(merged.get("story_point_mode") or base["story_point_mode"]).strip().lower()
    merged["story_point_mode"] = story_point_mode if story_point_mode in STORY_POINT_MODES else base["story_point_mode"]
    merged["fibonacci_points"] = normalize_number_list(
        merged.get("fibonacci_points"),
        base["fibonacci_points"],
    )

    sprint_length_policy = str(merged.get("sprint_length_policy") or base["sprint_length_policy"]).strip().lower()
    merged["sprint_length_policy"] = (
        sprint_length_policy if sprint_length_policy in SPRINT_LENGTH_POLICIES else base["sprint_length_policy"]
    )

    dod_management = str(merged.get("dod_management") or base["dod_management"]).strip().lower()
    merged["dod_management"] = dod_management if dod_management in DOD_MANAGEMENT_MODES else base["dod_management"]
    merged["default_dod_items"] = normalize_dod_items(merged.get("default_dod_items"))
    merged["openproject_sprint_version_sync"] = bool(merged.get("openproject_sprint_version_sync"))
    merged["metadata"] = normalize_metadata(merged.get("metadata"))
    merged["tenant_id"] = tenant_id
    return merged


def project_operation_policy_response(record: asyncpg.Record | dict[str, Any] | None, tenant_id: str) -> dict[str, Any]:
    if record:
        policy = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    else:
        policy = default_project_operation_policy(tenant_id)
    normalized = normalize_project_operation_policy_value(policy, tenant_id)
    if normalized.get("updated_by"):
        normalized["updated_by"] = str(normalized["updated_by"])
    return normalized


async def fetch_project_operation_policy(connection: asyncpg.Connection, tenant_id: str) -> dict[str, Any]:
    record = await connection.fetchrow(
        f"""
        SELECT {PROJECT_OPERATION_POLICY_SELECT}
        FROM wbs_project_operation_policies
        WHERE tenant_id = $1
        """,
        tenant_id,
    )
    return project_operation_policy_response(record, tenant_id)


def sprint_end_date_from_policy(start_date: date, policy: dict[str, Any], payload_end_date: date | None = None) -> date:
    sprint_policy = policy.get("sprint_length_policy") or "custom"
    if sprint_policy in SPRINT_LENGTH_DAYS:
        return start_date + timedelta(days=SPRINT_LENGTH_DAYS[sprint_policy] - 1)
    return payload_end_date or start_date + timedelta(days=13)


def validate_story_points_for_policy(value: float | int | Decimal | None, policy: dict[str, Any]) -> float:
    points = float(value or 0)
    if policy.get("story_point_mode") != "fibonacci" or points == 0:
        return points
    allowed = normalize_number_list(policy.get("fibonacci_points"), DEFAULT_PROJECT_OPERATION_POLICY["fibonacci_points"])
    if not any(abs(points - allowed_point) < 0.001 for allowed_point in allowed):
        allowed_label = ", ".join(str(int(p)) if float(p).is_integer() else str(p) for p in allowed)
        raise HTTPException(status_code=400, detail=f"Story Point must follow Fibonacci policy: {allowed_label}")
    return points


def metadata_with_policy_dod(metadata: dict[str, Any], policy: dict[str, Any]) -> dict[str, Any]:
    merged = normalize_metadata(metadata)
    merged.setdefault("dod_management", policy.get("dod_management") or "team")
    if not merged.get("dod_items") and policy.get("default_dod_items"):
        merged["dod_items"] = [{"text": item, "done": False} for item in policy["default_dod_items"]]
    if policy.get("dod_management") == "organization":
        merged["dod_source"] = "organization_policy"
    elif policy.get("default_dod_items"):
        merged["dod_source"] = "team_policy_default"
    return merged


def agile_sprint_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    sprint = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    sprint["id"] = str(sprint["id"])
    sprint["project_id"] = str(sprint["project_id"])
    sprint["metadata"] = normalize_metadata(sprint.get("metadata"))
    return sprint


def agile_item_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    item = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    item["id"] = str(item["id"])
    item["project_id"] = str(item["project_id"])
    item["sprint_id"] = str(item["sprint_id"]) if item.get("sprint_id") else None
    item["parent_id"] = str(item["parent_id"]) if item.get("parent_id") else None
    item["metadata"] = normalize_metadata(item.get("metadata"))
    return item


def validate_user_role(role: str) -> str:
    normalized_role = role.strip().lower()
    if normalized_role not in ALLOWED_USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid user role")
    return normalized_role


def validate_user_status(status: str) -> str:
    normalized_status = status.strip().title()
    if normalized_status not in ALLOWED_USER_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid user status")
    return normalized_status


def setting_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    setting = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    if setting.get("is_sensitive"):
        setting["value"] = {"masked": True}
    return setting


def audit_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    event = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    if event.get("actor_user_id"):
        event["actor_user_id"] = str(event["actor_user_id"])
    return event


async def insert_audit_event(
    connection: asyncpg.Connection,
    *,
    request: Request | None = None,
    event_type: str,
    summary: str,
    entity_type: str | None = None,
    entity_id: Any = None,
    metadata: dict[str, Any] | None = None,
    actor: dict[str, Any] | None = None,
    actor_email: str | None = None,
    actor_role: str | None = None,
) -> None:
    request_user = getattr(request.state, "user", None) if request else None
    resolved_actor = actor or request_user or {}
    email = actor_email or resolved_actor.get("email")
    role = actor_role or resolved_actor.get("role")
    await connection.execute(
        """
        INSERT INTO wbs_audit_events
          (actor_user_id, actor_email, actor_role, event_type, entity_type,
           entity_id, summary, metadata, tenant_id)
        VALUES
          ($1, $2, $3, $4, $5, $6, $7, $8::jsonb, $9)
        """,
        safe_uuid(resolved_actor.get("id")),
        email,
        role,
        event_type,
        entity_type,
        str(entity_id) if entity_id is not None else None,
        summary,
        metadata or {},
        get_tenant_id(request) if request else DEFAULT_TENANT_ID,
    )


def _try_ldap_login(email: str, password: str) -> bool:
    """동기 LDAP 바인드 시도. 스레드풀에서 실행됨."""
    if not LDAP_SERVER:
        return False
    try:
        import ldap3  # type: ignore
        server = ldap3.Server(LDAP_SERVER, port=LDAP_PORT, use_ssl=LDAP_USE_SSL, connect_timeout=5)
        # 서비스 계정으로 유저 DN 조회
        bind_conn = ldap3.Connection(server, user=LDAP_BIND_DN, password=LDAP_BIND_PASSWORD, auto_bind=True)
        search_filter = LDAP_USER_FILTER.replace("{email}", ldap3.utils.conv.escape_filter_chars(email))
        bind_conn.search(LDAP_BASE_DN, search_filter, attributes=[LDAP_ATTR_EMAIL, LDAP_ATTR_NAME])
        if not bind_conn.entries:
            return False
        user_dn = bind_conn.entries[0].entry_dn
        bind_conn.unbind()
        # 유저 DN으로 실제 바인드 (패스워드 검증)
        user_conn = ldap3.Connection(server, user=user_dn, password=password)
        result = user_conn.bind()
        user_conn.unbind()
        return result
    except Exception:
        return False


def ldap_diagnostic_step(key: str, label: str, status: str, message: str, detail: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "key": key,
        "label": label,
        "status": status,
        "message": message,
        "detail": detail or {},
    }


def _run_ldap_diagnostics(email: str | None = None, password: str | None = None) -> dict[str, Any]:
    steps: list[dict[str, Any]] = []
    config = {
        "auth_backend": AUTH_BACKEND,
        "ldap_server": LDAP_SERVER,
        "ldap_port": LDAP_PORT,
        "ldap_use_ssl": LDAP_USE_SSL,
        "ldap_bind_dn_configured": bool(LDAP_BIND_DN),
        "ldap_bind_password_configured": bool(LDAP_BIND_PASSWORD),
        "ldap_base_dn_configured": bool(LDAP_BASE_DN),
        "ldap_user_filter": LDAP_USER_FILTER,
        "ldap_attr_email": LDAP_ATTR_EMAIL,
        "ldap_attr_name": LDAP_ATTR_NAME,
    }

    missing = [
        label for label, value in {
            "LDAP_SERVER": LDAP_SERVER,
            "LDAP_BASE_DN": LDAP_BASE_DN,
            "LDAP_USER_FILTER": LDAP_USER_FILTER,
        }.items() if not value
    ]
    if missing:
        steps.append(ldap_diagnostic_step("config", "환경 변수", "fail", f"필수 설정 누락: {', '.join(missing)}", config))
        return {"success": False, "steps": steps, "config": config}
    steps.append(ldap_diagnostic_step("config", "환경 변수", "pass", "LDAP 필수 설정 확인 완료", config))

    try:
        import ldap3  # type: ignore
    except Exception as exc:  # noqa: BLE001 - surface optional dependency status
        steps.append(ldap_diagnostic_step("dependency", "ldap3 모듈", "fail", "ldap3 패키지를 불러올 수 없습니다.", {"error": str(exc)}))
        return {"success": False, "steps": steps, "config": config}

    bind_conn = None
    try:
        server = ldap3.Server(LDAP_SERVER, port=LDAP_PORT, use_ssl=LDAP_USE_SSL, connect_timeout=5)
        bind_conn = ldap3.Connection(
            server,
            user=LDAP_BIND_DN or None,
            password=LDAP_BIND_PASSWORD or None,
            receive_timeout=5,
        )
        if not bind_conn.bind():
            steps.append(ldap_diagnostic_step(
                "service_bind",
                "서비스 계정 바인드",
                "fail",
                "LDAP 서버 연결 또는 서비스 계정 바인드 실패",
                {"result": bind_conn.result},
            ))
            return {"success": False, "steps": steps, "config": config}
        steps.append(ldap_diagnostic_step(
            "service_bind",
            "서비스 계정 바인드",
            "pass",
            "LDAP 서버 연결 및 서비스 계정 바인드 성공",
            {"server": LDAP_SERVER, "port": LDAP_PORT, "ssl": LDAP_USE_SSL},
        ))

        if not email:
            steps.append(ldap_diagnostic_step("user_search", "사용자 검색", "warn", "테스트 이메일이 없어 사용자 검색을 건너뜀"))
            return {"success": True, "steps": steps, "config": config}

        search_filter = LDAP_USER_FILTER.replace("{email}", ldap3.utils.conv.escape_filter_chars(email))
        found = bind_conn.search(LDAP_BASE_DN, search_filter, attributes=[LDAP_ATTR_EMAIL, LDAP_ATTR_NAME])
        if not found or not bind_conn.entries:
            steps.append(ldap_diagnostic_step(
                "user_search",
                "사용자 검색",
                "fail",
                "LDAP 사용자 검색 결과가 없습니다.",
                {"base_dn": LDAP_BASE_DN, "filter": search_filter, "result": bind_conn.result},
            ))
            return {"success": False, "steps": steps, "config": config}

        entry = bind_conn.entries[0]
        user_dn = entry.entry_dn
        steps.append(ldap_diagnostic_step(
            "user_search",
            "사용자 검색",
            "pass",
            "LDAP 사용자 DN 검색 성공",
            {"user_dn": user_dn, "matched_entries": len(bind_conn.entries)},
        ))
        bind_conn.unbind()
        bind_conn = None

        if not password:
            steps.append(ldap_diagnostic_step("user_bind", "사용자 바인드", "warn", "테스트 비밀번호가 없어 사용자 바인드를 건너뜀"))
            return {"success": True, "steps": steps, "config": config}

        user_conn = ldap3.Connection(server, user=user_dn, password=password, receive_timeout=5)
        user_ok = user_conn.bind()
        result = user_conn.result
        user_conn.unbind()
        if not user_ok:
            steps.append(ldap_diagnostic_step(
                "user_bind",
                "사용자 바인드",
                "fail",
                "사용자 자격증명 검증 실패",
                {"result": result},
            ))
            return {"success": False, "steps": steps, "config": config}
        steps.append(ldap_diagnostic_step("user_bind", "사용자 바인드", "pass", "사용자 자격증명 검증 성공"))
        return {"success": True, "steps": steps, "config": config}
    except Exception as exc:  # noqa: BLE001 - diagnostics should report all failures
        steps.append(ldap_diagnostic_step(
            "runtime",
            "실행 오류",
            "fail",
            "LDAP 진단 중 오류가 발생했습니다.",
            {"error": str(exc), "type": exc.__class__.__name__},
        ))
        return {"success": False, "steps": steps, "config": config}
    finally:
        try:
            if bind_conn:
                bind_conn.unbind()
        except Exception:
            pass


async def send_notification(
    connection: asyncpg.Connection,
    *,
    user_id: UUID,
    event_type: str,
    title: str,
    body: str = "",
    entity_type: str | None = None,
    entity_id: str | None = None,
    metadata: dict[str, Any] | None = None,
    email_to: str | None = None,
) -> None:
    """인앱 알림 DB 저장 + SMTP 이메일 발송 (실패해도 메인 플로우 차단 안 함)."""
    try:
        await connection.execute(
            """
            INSERT INTO wbs_notifications
              (user_id, event_type, title, body, entity_type, entity_id, metadata)
            VALUES ($1,$2,$3,$4,$5,$6,$7::jsonb)
            """,
            user_id, event_type, title, body, entity_type, entity_id, metadata or {},
        )
    except Exception:
        pass

    if SMTP_HOST and email_to:
        def _send() -> None:
            try:
                msg = MIMEText(body or title, "plain", "utf-8")
                msg["Subject"] = f"[AX WBS] {title}"
                msg["From"] = NOTIFY_FROM_EMAIL or SMTP_USER
                msg["To"] = email_to
                with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=5) as s:
                    s.ehlo()
                    if s.has_extn("STARTTLS"):
                        s.starttls()
                    if SMTP_USER and SMTP_PASSWORD:
                        s.login(SMTP_USER, SMTP_PASSWORD)
                    s.sendmail(msg["From"], [email_to], msg.as_string())
            except Exception:
                pass
        asyncio.get_event_loop().run_in_executor(None, _send)


def selected_tenant_id_from_request(request: Request) -> str:
    """X-Tenant-ID 헤더로 테넌트 식별. MULTITENANCY_ENABLED=false면 항상 default."""
    if not MULTITENANCY_ENABLED:
        return DEFAULT_TENANT_ID
    return request.headers.get("X-Tenant-ID", DEFAULT_TENANT_ID).strip() or DEFAULT_TENANT_ID


def get_tenant_id(request: Request) -> str:
    return getattr(request.state, "tenant_id", None) or selected_tenant_id_from_request(request)


async def validate_request_tenant(
    connection: asyncpg.Connection,
    request: Request,
    *,
    enforce_active: bool = True,
) -> dict[str, Any]:
    tenant_id = selected_tenant_id_from_request(request)
    request.state.tenant_id = tenant_id
    if not MULTITENANCY_ENABLED:
        request.state.tenant = {"id": tenant_id, "name": "Default Tenant", "status": "Active"}
        return request.state.tenant

    record = await connection.fetchrow(
        "SELECT id, name, status, metadata, created_at, updated_at FROM wbs_tenants WHERE id = $1",
        tenant_id,
    )
    if not record:
        raise HTTPException(status_code=403, detail=f"Tenant is not registered: {tenant_id}")

    tenant = normalize_record(record)
    request.state.tenant = tenant
    if enforce_active and tenant.get("status") != "Active":
        raise HTTPException(status_code=403, detail=f"Tenant is not active: {tenant_id}")
    return tenant


def report_schedule_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    schedule = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    schedule["recipients"] = normalize_email_recipients(schedule.get("recipients"))
    schedule["smtp_configured"] = bool(REPORT_SMTP_HOST)
    schedule["scheduler_enabled"] = REPORT_SCHEDULER_ENABLED and AsyncIOScheduler is not None
    return schedule


def report_run_response(record: asyncpg.Record | dict[str, Any]) -> dict[str, Any]:
    run = normalize_record(record) if isinstance(record, asyncpg.Record) else dict(record)
    if run.get("schedule_id"):
        run["schedule_id"] = str(run["schedule_id"])
    return run


def normalize_email_recipients(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        try:
            decoded = json.loads(value)
            value = decoded
        except json.JSONDecodeError:
            value = [value]
    if not isinstance(value, list):
        return []

    recipients: list[str] = []
    seen: set[str] = set()
    for item in value:
        email = str(item or "").strip()
        if not email or "@" not in email:
            continue
        normalized = email.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        recipients.append(email)
    return recipients


def report_job_id(schedule_key: str) -> str:
    return f"report-schedule:{schedule_key}"


def scheduler_runtime_state(app: FastAPI) -> dict[str, Any]:
    scheduler = getattr(app.state, "report_scheduler", None)
    jobs = scheduler.get_jobs() if scheduler else []
    return {
        "enabled": REPORT_SCHEDULER_ENABLED,
        "apscheduler_available": AsyncIOScheduler is not None and CronTrigger is not None,
        "running": bool(scheduler and getattr(scheduler, "running", False)),
        "jobs": len(jobs),
        "timezone": REPORT_DEFAULT_TIMEZONE,
    }


async def fetch_report_schedule(connection: asyncpg.Connection, schedule_key: str) -> dict[str, Any] | None:
    record = await connection.fetchrow(
        f"""
        SELECT {REPORT_SCHEDULE_SELECT}
        FROM wbs_report_schedules
        WHERE key = $1
        """,
        schedule_key,
    )
    return normalize_record(record) if record else None


async def refresh_report_scheduler(app: FastAPI) -> None:
    scheduler = getattr(app.state, "report_scheduler", None)
    if not scheduler or not CronTrigger:
        return

    for job in scheduler.get_jobs():
        if job.id.startswith(("weekly-report:", "report-schedule:")):
            scheduler.remove_job(job.id)

    async with app.state.pool.acquire() as connection:
        await connection.execute(
            """
            UPDATE wbs_report_schedules
            SET next_run_at = NULL,
                updated_at = now()
            WHERE enabled = false
              AND next_run_at IS NOT NULL
            """
        )
        records = await connection.fetch(
            f"""
            SELECT {REPORT_SCHEDULE_SELECT}
            FROM wbs_report_schedules
            WHERE enabled = true
            ORDER BY key
            """
        )

    for record in records:
        schedule = normalize_record(record)
        trigger_kwargs = {
            "hour": int(schedule["hour"]),
            "minute": int(schedule["minute"]),
            "timezone": schedule["timezone"] or REPORT_DEFAULT_TIMEZONE,
        }
        if schedule.get("report_type") == "weekly_project_status":
            trigger_kwargs["day_of_week"] = str(schedule["day_of_week"])
        trigger = CronTrigger(**trigger_kwargs)
        scheduler.add_job(
            scheduled_report_job,
            trigger=trigger,
            id=report_job_id(str(schedule["key"])),
            args=[app, str(schedule["key"])],
            replace_existing=True,
            coalesce=True,
            max_instances=1,
        )
        job = scheduler.get_job(report_job_id(str(schedule["key"])))
        next_run_at = job.next_run_time if job else None
        async with app.state.pool.acquire() as connection:
            await connection.execute(
                """
                UPDATE wbs_report_schedules
                SET next_run_at = $2,
                    updated_at = now()
                WHERE key = $1
                """,
                schedule["key"],
                next_run_at,
            )


async def scheduled_report_job(app: FastAPI, schedule_key: str) -> None:
    async with app.state.pool.acquire() as connection:
        schedule = await fetch_report_schedule(connection, schedule_key)
    if not schedule or not schedule.get("enabled"):
        return
    await execute_report_schedule(app, schedule, send_email=True, triggered_by="scheduler")
    await refresh_report_scheduler(app)


def worksheet_timestamp(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def append_report_table(worksheet, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    if worksheet.max_row > 1 or worksheet.cell(row=1, column=1).value:
        worksheet.append([])
    worksheet.append([title])
    title_row = worksheet.max_row
    worksheet.cell(title_row, 1).font = Font(bold=True, size=13)
    worksheet.append(headers)
    header_row = worksheet.max_row
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    thin_border = Border(bottom=Side(style="thin", color="D7DFEA"))
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        worksheet.append([worksheet_timestamp(value) for value in row])


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            length = max(length, len(str(cell.value or "")))
        worksheet.column_dimensions[column].width = min(max(length + 2, 12), 42)


async def build_weekly_report_workbook(
    pool: asyncpg.Pool,
    period_start: datetime,
    period_end: datetime,
) -> tuple[bytes, dict[str, Any]]:
    async with pool.acquire() as connection:
        project_count = await connection.fetchval("SELECT count(*) FROM wbs_projects")
        template_count = await connection.fetchval("SELECT count(*) FROM wbs_templates")
        pending_approvals = await connection.fetchval(
            "SELECT count(*) FROM wbs_approval_requests WHERE status = 'Pending'"
        )
        preview_imports = await connection.fetchval("SELECT count(*) FROM wbs_import_jobs WHERE status = 'Preview'")
        locked_baselines = await connection.fetchval("SELECT count(*) FROM wbs_project_baselines WHERE status = 'Locked'")
        status_rows = await connection.fetch(
            """
            SELECT status, count(*) AS count
            FROM wbs_projects
            GROUP BY status
            ORDER BY status
            """
        )
        projects = await connection.fetch(
            """
            SELECT name, owner, status, template_key, start_date,
                   openproject_project_id, created_at, updated_at
            FROM wbs_projects
            ORDER BY updated_at DESC
            LIMIT 100
            """
        )
        approvals = await connection.fetch(
            f"""
            SELECT {APPROVAL_SELECT}
            FROM wbs_approval_requests a
            JOIN wbs_projects p ON p.id = a.project_id
            WHERE a.created_at >= $1 OR a.updated_at >= $1
            ORDER BY a.updated_at DESC
            LIMIT 100
            """,
            period_start,
        )
        imports = await connection.fetch(
            """
            SELECT source_file, template_key, template_name, status, total_rows,
                   accepted_rows, rejected_rows, applied_at, created_at
            FROM wbs_import_jobs
            WHERE created_at >= $1 OR applied_at >= $1
            ORDER BY created_at DESC
            LIMIT 100
            """,
            period_start,
        )
        sync_runs = await connection.fetch(
            f"""
            SELECT {SYNC_RUN_SELECT}
            FROM wbs_sync_runs s
            JOIN wbs_projects p ON p.id = s.project_id
            WHERE s.started_at >= $1
            ORDER BY s.started_at DESC
            LIMIT 100
            """,
            period_start,
        )
        wbs_progress = await connection.fetch(
            """
            SELECT
                p.name AS project_name,
                p.owner,
                p.status AS project_status,
                count(i.code)::integer AS total_items,
                count(i.code) FILTER (
                    WHERE COALESCE(i.metadata->>'status', '') = '완료'
                       OR CASE
                            WHEN COALESCE(i.metadata->>'progress', '') ~ '^[0-9]+(\\.[0-9]+)?$'
                            THEN (i.metadata->>'progress')::numeric
                            ELSE 0
                          END >= 100
                )::integer AS done_items,
                count(i.code) FILTER (
                    WHERE i.finish_date < CURRENT_DATE
                      AND COALESCE(i.metadata->>'status', '') <> '완료'
                      AND CASE
                            WHEN COALESCE(i.metadata->>'progress', '') ~ '^[0-9]+(\\.[0-9]+)?$'
                            THEN (i.metadata->>'progress')::numeric
                            ELSE 0
                          END < 100
                )::integer AS overdue_items,
                round(avg(
                    CASE
                        WHEN COALESCE(i.metadata->>'progress', '') ~ '^[0-9]+(\\.[0-9]+)?$'
                        THEN (i.metadata->>'progress')::numeric
                        ELSE 0
                    END
                ), 1) AS avg_progress
            FROM wbs_projects p
            LEFT JOIN wbs_project_wbs_items i
              ON i.project_id = p.id AND COALESCE(i.item_type, '작업') <> '프로젝트'
            GROUP BY p.id
            ORDER BY p.updated_at DESC
            LIMIT 100
            """
        )

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    append_report_table(
        summary,
        "WBS 주간 프로젝트 보고서",
        ["항목", "값"],
        [
            ["보고 기간", f"{period_start.date().isoformat()} ~ {period_end.date().isoformat()}"],
            ["생성 시각", period_end],
            ["프로젝트", project_count],
            ["템플릿", template_count],
            ["승인 대기", pending_approvals],
            ["Excel 미반영 Preview", preview_imports],
            ["Locked Baseline", locked_baselines],
        ],
    )
    append_report_table(
        summary,
        "프로젝트 상태 분포",
        ["상태", "건수"],
        [[row["status"], row["count"]] for row in status_rows],
    )
    autosize_worksheet(summary)

    project_sheet = wb.create_sheet("Projects")
    append_report_table(
        project_sheet,
        "프로젝트 현황",
        ["프로젝트", "담당", "상태", "템플릿", "시작일", "외부 프로젝트 ID", "생성", "수정"],
        [
            [
                row["name"],
                row["owner"],
                row["status"],
                row["template_key"],
                row["start_date"],
                row["openproject_project_id"],
                row["created_at"],
                row["updated_at"],
            ]
            for row in projects
        ],
    )
    autosize_worksheet(project_sheet)

    progress_sheet = wb.create_sheet("WBS Progress")
    append_report_table(
        progress_sheet,
        "내부 WBS 진행 현황",
        ["프로젝트", "담당", "상태", "작업 항목", "완료", "지연", "평균 진척률"],
        [
            [
                row["project_name"],
                row["owner"],
                row["project_status"],
                row["total_items"],
                row["done_items"],
                row["overdue_items"],
                f"{row['avg_progress'] or 0}%",
            ]
            for row in wbs_progress
        ],
    )
    autosize_worksheet(progress_sheet)

    approval_sheet = wb.create_sheet("Approvals")
    append_report_table(
        approval_sheet,
        "최근 승인 이력",
        ["프로젝트", "제목", "상태", "요청자", "검토자", "생성", "결정"],
        [
            [
                row["project_name"],
                row["title"],
                row["status"],
                row["requester"],
                row["reviewer"],
                row["created_at"],
                row["decided_at"],
            ]
            for row in approvals
        ],
    )
    autosize_worksheet(approval_sheet)

    import_sheet = wb.create_sheet("Excel Imports")
    append_report_table(
        import_sheet,
        "최근 Excel 반영",
        ["파일", "템플릿", "상태", "전체", "정상", "오류", "생성", "반영"],
        [
            [
                row["source_file"],
                row["template_name"] or row["template_key"],
                row["status"],
                row["total_rows"],
                row["accepted_rows"],
                row["rejected_rows"],
                row["created_at"],
                row["applied_at"],
            ]
            for row in imports
        ],
    )
    autosize_worksheet(import_sheet)

    sync_sheet = wb.create_sheet("Baseline Runs")
    append_report_table(
        sync_sheet,
        "최근 기준선 반영 실행",
        ["프로젝트", "모드", "상태", "전체", "생성", "외부 프로젝트 ID", "시작", "완료"],
        [
            [
                row["project_name"],
                row["mode"],
                row["status"],
                row["total_rows"],
                row["created_work_packages"],
                row["openproject_project_id"],
                row["started_at"],
                row["completed_at"],
            ]
            for row in sync_runs
        ],
    )
    autosize_worksheet(sync_sheet)

    buffer = BytesIO()
    wb.save(buffer)
    metadata = {
        "project_count": project_count,
        "pending_approvals": pending_approvals,
        "preview_imports": preview_imports,
        "sync_runs": len(sync_runs),
        "wbs_progress_projects": len(wbs_progress),
        "approvals": len(approvals),
        "imports": len(imports),
        "data_source": "internal_wbs",
    }
    return buffer.getvalue(), metadata


def save_report_artifact(filename: str, workbook_bytes: bytes) -> str:
    REPORT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORT_OUTPUT_DIR / filename
    path.write_bytes(workbook_bytes)
    return str(path)


def send_weekly_report_email(schedule: dict[str, Any], workbook_bytes: bytes, filename: str) -> dict[str, Any]:
    recipients = normalize_email_recipients(schedule.get("recipients"))
    if not recipients:
        return {"status": "skipped", "reason": "no_recipients"}
    if not REPORT_SMTP_HOST:
        return {"status": "skipped", "reason": "smtp_not_configured", "recipients": recipients}

    message = EmailMessage()
    message["Subject"] = schedule.get("email_subject") or "[WBS] 주간 프로젝트 보고서"
    message["From"] = REPORT_SMTP_FROM
    message["To"] = ", ".join(recipients)
    message["Date"] = formatdate(localtime=True)
    message.set_content(schedule.get("email_body") or "주간 WBS 프로젝트 현황 보고서를 첨부합니다.")
    message.add_attachment(
        workbook_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )

    smtp_cls = smtplib.SMTP_SSL if REPORT_SMTP_USE_SSL else smtplib.SMTP
    with smtp_cls(REPORT_SMTP_HOST, REPORT_SMTP_PORT, timeout=30) as smtp:
        if REPORT_SMTP_USE_TLS and not REPORT_SMTP_USE_SSL:
            smtp.starttls()
        if REPORT_SMTP_USERNAME:
            smtp.login(REPORT_SMTP_USERNAME, REPORT_SMTP_PASSWORD)
        smtp.send_message(message)

    return {"status": "sent", "recipients": recipients, "smtp_host": REPORT_SMTP_HOST}


async def execute_weekly_report_schedule(
    app: FastAPI,
    schedule: dict[str, Any],
    *,
    send_email: bool,
    triggered_by: str,
    request: Request | None = None,
) -> dict[str, Any]:
    period_end = datetime.now(timezone.utc)
    period_start = period_end - timedelta(days=7)
    started_at = datetime.now(timezone.utc)
    filename = f"wbs-weekly-report-{period_end.strftime('%Y%m%d-%H%M%S')}.xlsx"
    workbook_bytes: bytes | None = None
    artifact_path: str | None = None
    delivery = {"status": "skipped", "reason": "send_email_false"}
    error_payload: dict[str, Any] | None = None
    status = "Generated"
    metadata: dict[str, Any] = {}

    try:
        workbook_bytes, metadata = await build_weekly_report_workbook(app.state.pool, period_start, period_end)
        artifact_path = save_report_artifact(filename, workbook_bytes)
        if send_email:
            delivery = send_weekly_report_email(schedule, workbook_bytes, filename)
        status = "Sent" if delivery.get("status") == "sent" else "Generated"
    except Exception as exc:  # noqa: BLE001 - record scheduler failures instead of crashing the job loop
        status = "Failed"
        error_payload = {
            "message": str(exc),
            "type": exc.__class__.__name__,
        }

    async with app.state.pool.acquire() as connection:
        async with connection.transaction():
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_report_runs
                  (schedule_id, schedule_key, report_type, status, period_start,
                   period_end, recipient_count, artifact_path, delivery_status,
                   delivery_detail, error, metadata, triggered_by, started_at,
                   completed_at)
                VALUES
                  ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb, $11::jsonb,
                   $12::jsonb, $13, $14, now())
                RETURNING {REPORT_RUN_SELECT}
                """,
                safe_uuid(schedule.get("id")),
                schedule["key"],
                schedule.get("report_type") or "weekly_project_status",
                status,
                period_start,
                period_end,
                len(normalize_email_recipients(schedule.get("recipients"))),
                artifact_path,
                delivery.get("status", "skipped"),
                delivery,
                error_payload,
                metadata,
                triggered_by,
                started_at,
            )
            await connection.execute(
                """
                UPDATE wbs_report_schedules
                SET last_run_at = now(),
                    updated_at = now()
                WHERE key = $1
                """,
                schedule["key"],
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="report.weekly_sent" if status == "Sent" else "report.weekly_generated" if status == "Generated" else "report.weekly_failed",
                entity_type="report_schedule",
                entity_id=schedule["key"],
                summary=f"Weekly report {status.lower()}: {schedule['key']}",
                metadata={
                    "delivery_status": delivery.get("status"),
                    "artifact_path": artifact_path,
                    "triggered_by": triggered_by,
                    "error": error_payload,
                },
                actor={"email": "scheduler@wbs.local", "role": "system"} if not request else None,
            )

    return report_run_response(record)


def send_schedule_text_email(schedule: dict[str, Any], subject: str, body: str) -> dict[str, Any]:
    recipients = normalize_email_recipients(schedule.get("recipients"))
    if not recipients:
        return {"status": "skipped", "reason": "no_recipients"}
    if not REPORT_SMTP_HOST:
        return {"status": "skipped", "reason": "smtp_not_configured"}

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = REPORT_SMTP_FROM
    message["To"] = ", ".join(recipients)
    message["Date"] = formatdate(localtime=True)
    message.set_content(body)

    smtp_cls = smtplib.SMTP_SSL if REPORT_SMTP_USE_SSL else smtplib.SMTP
    with smtp_cls(REPORT_SMTP_HOST, REPORT_SMTP_PORT, timeout=30) as smtp:
        if REPORT_SMTP_USE_TLS and not REPORT_SMTP_USE_SSL:
            smtp.starttls()
        if REPORT_SMTP_USERNAME:
            smtp.login(REPORT_SMTP_USERNAME, REPORT_SMTP_PASSWORD)
        smtp.send_message(message)

    return {"status": "sent", "recipients": recipients, "smtp_host": REPORT_SMTP_HOST}


async def build_notification_digest(pool: asyncpg.Pool, schedule: dict[str, Any]) -> tuple[str, str, dict[str, Any]]:
    report_type = schedule.get("report_type")
    subject = schedule.get("email_subject") or f"[WBS] {schedule.get('name') or schedule.get('key')}"
    intro = schedule.get("email_body") or ""
    if report_type not in {"risk_escalation", "approval_reminder"}:
        return subject, intro.strip(), {
            "item_count": 0,
            "digest_type": str(report_type or "unknown"),
            "unsupported": True,
        }

    async with pool.acquire() as connection:
        if report_type == "risk_escalation":
            table_exists = await connection.fetchval("SELECT to_regclass('public.wbs_risks') IS NOT NULL")
            rows = await connection.fetch(
                """
                SELECT r.title, r.severity, r.status, r.owner, r.due_date, p.name AS project_name
                FROM wbs_risks r
                JOIN wbs_projects p ON p.id = r.project_id
                WHERE r.status != 'Closed'
                  AND r.severity = '높음'
                ORDER BY r.due_date NULLS LAST, r.created_at DESC
                LIMIT 25
                """
            ) if table_exists else []
            lines = [intro, "", "미처리 고위험 리스크"]
            if rows:
                lines.extend(
                    f"- [{row['project_name']}] {row['title']} · 담당 {row['owner']} · 상태 {row['status']} · 기한 {row['due_date'] or '-'}"
                    for row in rows
                )
            else:
                lines.append("- 대상 항목 없음")
            return subject, "\n".join(line for line in lines if line is not None).strip(), {
                "item_count": len(rows),
                "digest_type": "risk_escalation",
            }

        table_exists = await connection.fetchval("SELECT to_regclass('public.wbs_approval_requests') IS NOT NULL")
        rows = await connection.fetch(
            """
            SELECT a.title, a.request_type, a.requester, a.reviewer, a.due_date, p.name AS project_name
            FROM wbs_approval_requests a
            JOIN wbs_projects p ON p.id = a.project_id
            WHERE a.status = 'Pending'
            ORDER BY a.due_date NULLS LAST, a.created_at DESC
            LIMIT 25
            """
        ) if table_exists else []
        lines = [intro, "", "승인 대기 항목"]
        if rows:
            lines.extend(
                f"- [{row['project_name']}] {row['title']} · 요청 {row['requester']} · 승인자 {row['reviewer'] or '-'} · 기한 {row['due_date'] or '-'}"
                for row in rows
            )
        else:
            lines.append("- 대상 항목 없음")
        return subject, "\n".join(line for line in lines if line is not None).strip(), {
            "item_count": len(rows),
            "digest_type": "approval_reminder",
        }


async def execute_notification_report_schedule(
    app: FastAPI,
    schedule: dict[str, Any],
    *,
    send_email: bool,
    triggered_by: str,
    request: Request | None = None,
) -> dict[str, Any]:
    period_end = datetime.now(timezone.utc)
    period_start = period_end - timedelta(days=1)
    started_at = datetime.now(timezone.utc)
    delivery = {"status": "skipped", "reason": "send_email_false"}
    error_payload: dict[str, Any] | None = None
    status = "Generated"
    metadata: dict[str, Any] = {}

    try:
        subject, body, metadata = await build_notification_digest(app.state.pool, schedule)
        if send_email:
            if int(metadata.get("item_count") or 0) > 0:
                delivery = send_schedule_text_email(schedule, subject, body)
            else:
                delivery = {"status": "skipped", "reason": "no_items"}
        status = "Sent" if delivery.get("status") == "sent" else "Generated"
    except Exception as exc:  # noqa: BLE001 - record scheduler failures instead of crashing the job loop
        status = "Failed"
        error_payload = {
            "message": str(exc),
            "type": exc.__class__.__name__,
        }

    async with app.state.pool.acquire() as connection:
        async with connection.transaction():
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_report_runs
                  (schedule_id, schedule_key, report_type, status, period_start,
                   period_end, recipient_count, artifact_path, delivery_status,
                   delivery_detail, error, metadata, triggered_by, started_at,
                   completed_at)
                VALUES
                  ($1, $2, $3, $4, $5, $6, $7, NULL, $8, $9::jsonb, $10::jsonb,
                   $11::jsonb, $12, $13, now())
                RETURNING {REPORT_RUN_SELECT}
                """,
                safe_uuid(schedule.get("id")),
                schedule["key"],
                schedule.get("report_type") or "notification_digest",
                status,
                period_start,
                period_end,
                len(normalize_email_recipients(schedule.get("recipients"))),
                delivery.get("status", "skipped"),
                delivery,
                error_payload,
                metadata,
                triggered_by,
                started_at,
            )
            await connection.execute(
                """
                UPDATE wbs_report_schedules
                SET last_run_at = now(),
                    updated_at = now()
                WHERE key = $1
                """,
                schedule["key"],
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type=f"report.{schedule.get('report_type')}.{status.lower()}",
                entity_type="report_schedule",
                entity_id=schedule["key"],
                summary=f"{schedule.get('name') or schedule['key']} {status.lower()}",
                metadata={
                    "delivery_status": delivery.get("status"),
                    "triggered_by": triggered_by,
                    "error": error_payload,
                    **metadata,
                },
                actor={"email": "scheduler@wbs.local", "role": "system"} if not request else None,
            )

    return report_run_response(record)


async def execute_report_schedule(
    app: FastAPI,
    schedule: dict[str, Any],
    *,
    send_email: bool,
    triggered_by: str,
    request: Request | None = None,
) -> dict[str, Any]:
    if schedule.get("report_type") == "weekly_project_status":
        return await execute_weekly_report_schedule(
            app,
            schedule,
            send_email=send_email,
            triggered_by=triggered_by,
            request=request,
        )
    return await execute_notification_report_schedule(
        app,
        schedule,
        send_email=send_email,
        triggered_by=triggered_by,
        request=request,
    )


def apply_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


@app.middleware("http")
async def authenticate_api_requests(request: Request, call_next):
    path = request.url.path
    public_paths = {"/api/auth/login"}
    if request.method == "OPTIONS" or not path.startswith("/api/"):
        return apply_security_headers(await call_next(request))

    is_tenant_management = path == "/api/tenants" or path.startswith("/api/tenants/")
    is_auth_context = path in {"/api/auth/me", "/api/auth/logout"}
    async with get_pool(request).acquire() as connection:
        try:
            await validate_request_tenant(
                connection,
                request,
                enforce_active=not (is_tenant_management or is_auth_context),
            )
        except HTTPException as exc:
            return apply_security_headers(JSONResponse(status_code=exc.status_code, content={"detail": exc.detail}))

        if path in public_paths:
            return apply_security_headers(await call_next(request))

        token = auth_token_from_request(request)
        if not token:
            return apply_security_headers(JSONResponse(status_code=401, content={"detail": "Authentication required"}))

        user = await fetch_user_by_token(connection, token)

    if not user:
        return apply_security_headers(JSONResponse(status_code=401, content={"detail": "Invalid or expired session"}))
    if (
        MULTITENANCY_ENABLED
        and not (is_tenant_management or is_auth_context)
        and user.get("role") != "admin"
        and user.get("tenant_id") != get_tenant_id(request)
    ):
        return apply_security_headers(JSONResponse(status_code=403, content={"detail": "User is not assigned to this tenant"}))

    request.state.user = user
    return apply_security_headers(await call_next(request))


def parse_json_object(value: str, *, default: dict[str, Any] | None = None) -> dict[str, Any]:
    try:
        decoded = json.loads(value or "{}")
        return decoded if isinstance(decoded, dict) else (default or {})
    except json.JSONDecodeError:
        return default or {}


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_./-]+", "", str(value or "").strip().lower())


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_code(value: Any) -> str | None:
    text = normalize_text(value)
    return re.sub(r"\s+", "", text) if text else None


def normalize_template_key(value: str) -> str:
    key = re.sub(r"[^a-z0-9_-]+", "-", value.strip().lower())
    key = re.sub(r"-{2,}", "-", key).strip("-")
    if not key:
        raise HTTPException(status_code=400, detail="Template key is required")
    return key[:80]


def normalize_openproject_identifier(value: str, fallback: str) -> str:
    identifier = re.sub(r"[^a-z0-9-]+", "-", value.strip().lower())
    identifier = re.sub(r"-{2,}", "-", identifier).strip("-")
    return (identifier or fallback)[:100]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Invalid number: {value}") from exc


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "t", "yes", "y", "예", "검수", "필수", "o"}


def as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Invalid date: {value}")

# 하위 호환 alias
def parse_date(value: Any) -> date | None:
    try:
        return as_date(value)
    except ValueError:
        return None


def infer_parent_from_code(code: str) -> str | None:
    if "." not in code:
        return None
    return code.rsplit(".", 1)[0]


def template_code_prefix(template_key: str) -> str:
    known_prefixes = {
        "si-standard": "SI",
        "migration-data": "MIG",
        "maintenance": "OPS",
        "agile-standard": "AGL",
        "hybrid-standard": "HYB",
    }
    if template_key in known_prefixes:
        return known_prefixes[template_key]
    words = re.findall(r"[A-Z0-9]+", template_key.upper())
    if not words:
        return "WBS"
    return "".join(word[0] for word in words)[:6]


def template_delivery_mode(template: dict[str, Any]) -> str:
    key = str(template.get("key") or "").strip().lower()
    project_type = str(template.get("project_type") or "").strip().lower()
    if project_type == "agile" or key.startswith("agile"):
        return "agile"
    if project_type == "hybrid" or key.startswith("hybrid"):
        return "hybrid"
    return "waterfall"


def parent_depth(code: str, parent_map: dict[str, str | None]) -> int:
    depth = 1
    seen = {code}
    parent = parent_map.get(code)
    while parent and parent not in seen:
        seen.add(parent)
        depth += 1
        parent = parent_map.get(parent)
    return depth


def code_depth(code: str) -> int:
    return len(code.split(".")) if code else 1


def generated_child_code(parent_code: str, item_type: str, counters: dict[str, dict[str, int]], used_codes: set[str]) -> str:
    bucket = counters.setdefault(parent_code, {"normal": 0, "milestone": 0})
    if item_type == "마일스톤":
        while True:
            bucket["milestone"] += 1
            candidate = f"{parent_code}.M{bucket['milestone']}"
            if candidate not in used_codes:
                return candidate

    while True:
        bucket["normal"] += 1
        candidate = f"{parent_code}.{bucket['normal']}"
        if candidate not in used_codes:
            return candidate


def assign_missing_wbs_codes(rows: list[dict[str, Any]], root_code: str) -> list[dict[str, Any]]:
    assigned_rows = [dict(row) for row in rows]
    used_codes = {row["code"] for row in assigned_rows if row.get("code")}
    level_stack: dict[int, str] = {}
    counters: dict[str, dict[str, int]] = {}
    root_assigned = root_code in used_codes

    for row in assigned_rows:
        level = row.get("level")
        code = normalize_code(row.get("code"))
        parent_code = normalize_code(row.get("parent_code"))
        item_type = row.get("item_type") or "작업"

        if code:
            if not parent_code and level and level > 1:
                parent_code = level_stack.get(level - 1)
            if not parent_code:
                parent_code = infer_parent_from_code(code)
            row["code"] = code
            row["parent_code"] = parent_code
            resolved_level = level or code_depth(code)
            level_stack[resolved_level] = code
            for stale_level in [key for key in level_stack if key > resolved_level]:
                del level_stack[stale_level]
            used_codes.add(code)
            continue

        if not parent_code and level and level > 1:
            parent_code = level_stack.get(level - 1)

        if not parent_code and not root_assigned:
            code = root_code
            root_assigned = True
        elif parent_code:
            code = generated_child_code(parent_code, item_type, counters, used_codes)
        else:
            code = generated_child_code(root_code, item_type, counters, used_codes)
            parent_code = root_code if root_assigned else None

        row["code"] = code
        row["parent_code"] = parent_code
        row["code_generated"] = True
        used_codes.add(code)
        resolved_level = level or (parent_depth(code, {item["code"]: item.get("parent_code") for item in assigned_rows if item.get("code")}))
        level_stack[resolved_level] = code
        for stale_level in [key for key in level_stack if key > resolved_level]:
            del level_stack[stale_level]

    return assigned_rows


def auto_code_warnings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "row": row.get("row_number"),
            "field": "code",
            "message": f"WBS code was generated as {row['code']}",
        }
        for row in rows
        if row.get("code_generated")
    ]


def root_code_from_rows(rows: list[dict[str, Any]]) -> str | None:
    roots = [row for row in rows if row.get("code") and not row.get("parent_code")]
    if len(roots) == 1:
        return roots[0]["code"]
    return None


def renumber_wbs_rows(rows: list[dict[str, Any]], root_code: str | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not rows:
        return [], []

    ordered_rows = sorted(rows, key=lambda row: (row.get("sort_order") or 0, row.get("code") or ""))
    children_by_parent: dict[str | None, list[dict[str, Any]]] = {}
    for row in ordered_rows:
        children_by_parent.setdefault(row.get("parent_code"), []).append(dict(row))

    roots = children_by_parent.get(None, [])
    if not roots:
        roots = [ordered_rows[0]]
        children_by_parent.setdefault(None, roots)

    resolved_root_code = root_code or roots[0].get("code") or "WBS"
    changes: list[dict[str, Any]] = []
    renumbered_rows: list[dict[str, Any]] = []

    def walk(row: dict[str, Any], new_code: str, new_parent_code: str | None) -> None:
        old_code = row.get("code")
        if old_code != new_code or row.get("parent_code") != new_parent_code:
            changes.append(
                {
                    "name": row.get("name"),
                    "old_code": old_code,
                    "new_code": new_code,
                    "old_parent_code": row.get("parent_code"),
                    "new_parent_code": new_parent_code,
                }
            )

        next_row = dict(row)
        next_row["code"] = new_code
        next_row["parent_code"] = new_parent_code
        renumbered_rows.append(next_row)

        counters = {"normal": 0, "milestone": 0}
        for child in children_by_parent.get(old_code, []):
            if child.get("item_type") == "마일스톤":
                counters["milestone"] += 1
                child_code = f"{new_code}.M{counters['milestone']}"
            else:
                counters["normal"] += 1
                child_code = f"{new_code}.{counters['normal']}"
            walk(child, child_code, new_code)

    for index, root in enumerate(roots, start=1):
        new_root_code = resolved_root_code if index == 1 else f"{resolved_root_code}.{index}"
        walk(root, new_root_code, None)

    for index, row in enumerate(renumbered_rows, start=1):
        row["sort_order"] = index

    return renumbered_rows, changes


def parse_wbs_workbook(contents: bytes) -> list[dict[str, Any]]:
    if len(contents) > MAX_EXCEL_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Excel file is too large")

    try:
        workbook = load_workbook(BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel workbook: {exc}") from exc

    worksheet = workbook[workbook.sheetnames[0]]
    header_row = None
    header_map: dict[str, int] = {}

    for row_number in range(1, min(worksheet.max_row, 10) + 1):
        candidate: dict[str, int] = {}
        for column_number in range(1, worksheet.max_column + 1):
            header = HEADER_ALIASES.get(normalize_header(worksheet.cell(row_number, column_number).value))
            if header:
                candidate[header] = column_number
        if "name" in candidate:
            header_row = row_number
            header_map = candidate
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="작업명 헤더를 찾을 수 없습니다")

    rows: list[dict[str, Any]] = []
    level_stack: dict[int, str] = {}

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        raw = {
            field: worksheet.cell(row_number, column_number).value
            for field, column_number in header_map.items()
        }
        if not any(normalize_text(raw.get(field)) for field in ("code", "parent_code", "name")):
            continue

        try:
            level_value = as_float(raw.get("level"))
            level = int(level_value) if level_value else None
            code = normalize_code(raw.get("code"))
            parent_code = normalize_code(raw.get("parent_code"))
            if code and not parent_code and level and level > 1:
                parent_code = level_stack.get(level - 1)
            if code and not parent_code:
                parent_code = infer_parent_from_code(code)
            if code and level:
                level_stack[level] = code
                for stale_level in [key for key in level_stack if key > level]:
                    del level_stack[stale_level]

            rows.append(
                {
                    "row_number": row_number,
                    "level": level,
                    "code": code,
                    "name": normalize_text(raw.get("name")),
                    "parent_code": parent_code,
                    "item_type": normalize_text(raw.get("item_type")) or "작업",
                    "owner": normalize_text(raw.get("owner")),
                    "weight": as_float(raw.get("weight")),
                    "start_date": as_date(raw.get("start_date")),
                    "finish_date": as_date(raw.get("finish_date")),
                    "deliverable_type": normalize_text(raw.get("deliverable_type")),
                    "inspection_required": as_bool(raw.get("inspection_required")),
                    "progress_formula": normalize_text(raw.get("progress_formula")),
                    "notes": normalize_text(raw.get("notes")),
                }
            )
        except ValueError as exc:
            rows.append(
                {
                    "row_number": row_number,
                    "code": normalize_code(raw.get("code")),
                    "name": normalize_text(raw.get("name")),
                    "parent_code": normalize_code(raw.get("parent_code")),
                    "item_type": normalize_text(raw.get("item_type")) or "작업",
                    "parse_error": str(exc),
                }
            )

    return rows


def validate_wbs_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    codes: dict[str, int] = {}

    if not rows:
        return ([{"row": None, "field": "file", "message": "No WBS rows found"}], warnings)

    for index, row in enumerate(rows, start=1):
        row_number = row.get("row_number", index)
        code = row.get("code")
        name = row.get("name")

        if row.get("parse_error"):
            errors.append({"row": row_number, "field": "format", "message": row["parse_error"]})
        if not code:
            errors.append({"row": row_number, "field": "code", "message": "WBS code is required"})
        elif code in codes:
            errors.append({"row": row_number, "field": "code", "message": "Duplicate WBS code"})
        else:
            codes[code] = row_number
        if not name:
            errors.append({"row": row_number, "field": "name", "message": "Task name is required"})

        weight = row.get("weight")
        if weight is not None and not 0 <= weight <= 100:
            errors.append({"row": row_number, "field": "weight", "message": "Weight must be between 0 and 100"})

        if row.get("finish_date") and row.get("start_date") and row["finish_date"] < row["start_date"]:
            errors.append({"row": row_number, "field": "finish_date", "message": "Finish date is earlier than start date"})

    parent_map = {row["code"]: row.get("parent_code") for row in rows if row.get("code")}
    weight_map = {row["code"]: row.get("weight") for row in rows if row.get("code")}
    child_weight: dict[str, float] = {}

    for row in rows:
        code = row.get("code")
        parent_code = row.get("parent_code")
        if not code:
            continue
        if parent_code and parent_code not in codes:
            errors.append(
                {
                    "row": row.get("row_number"),
                    "field": "parent_code",
                    "message": "Parent code does not exist in import file",
                }
            )
        parent_key = parent_code or "__root__"
        child_weight[parent_key] = child_weight.get(parent_key, 0) + float(row.get("weight") or 0)

    for code in parent_map:
        seen = {code}
        parent_code = parent_map.get(code)
        while parent_code:
            if parent_code in seen:
                errors.append({"row": codes.get(code), "field": "parent_code", "message": "Circular hierarchy detected"})
                break
            seen.add(parent_code)
            parent_code = parent_map.get(parent_code)

    for parent_code, total_weight in child_weight.items():
        if not total_weight:
            continue
        expected = 100 if parent_code == "__root__" else weight_map.get(parent_code)
        if expected is not None and abs(total_weight - float(expected)) > 0.01:
            warnings.append(
                {
                    "parent_code": None if parent_code == "__root__" else parent_code,
                    "message": f"Sibling weights add up to {total_weight:.2f}, expected {float(expected):.2f}",
                }
            )

    return errors, warnings


def serialize_wbs_row(row: dict[str, Any]) -> dict[str, Any]:
    serialized = dict(row)
    for key in ("start_date", "finish_date"):
        if isinstance(serialized.get(key), date):
            serialized[key] = serialized[key].isoformat()
    return serialized


def comparable_wbs_row(row: dict[str, Any]) -> dict[str, Any]:
    metadata = normalize_metadata(row.get("metadata"))
    return {
        "parent_code": row.get("parent_code"),
        "name": row.get("name"),
        "item_type": row.get("item_type") or "작업",
        "owner": row.get("owner"),
        "weight": float(row["weight"]) if row.get("weight") is not None else None,
        "start_date": row.get("start_date").isoformat() if isinstance(row.get("start_date"), date) else row.get("start_date"),
        "finish_date": row.get("finish_date").isoformat() if isinstance(row.get("finish_date"), date) else row.get("finish_date"),
        "deliverable_type": row.get("deliverable_type", metadata.get("deliverable_type")),
        "inspection_required": row.get("inspection_required", metadata.get("inspection_required", False)),
        "progress_formula": row.get("progress_formula", metadata.get("progress_formula")),
        "notes": row.get("notes", metadata.get("notes")),
    }


def build_wbs_diff_rows(existing_rows: list[dict[str, Any]], next_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_by_code = {row["code"]: row for row in existing_rows if row.get("code")}
    next_by_code = {row["code"]: row for row in next_rows if row.get("code")}
    diff_rows: list[dict[str, Any]] = []

    for code, row in next_by_code.items():
        existing = existing_by_code.get(code)
        if not existing:
            diff_rows.append({"change": "added", "code": code, "name": row.get("name")})
            continue

        before = comparable_wbs_row(existing)
        after = comparable_wbs_row(row)
        changed_fields = [
            {"field": field, "before": before.get(field), "after": after.get(field)}
            for field in sorted(after)
            if before.get(field) != after.get(field)
        ]
        if changed_fields:
            diff_rows.append(
                {
                    "change": "changed",
                    "code": code,
                    "name": row.get("name"),
                    "fields": changed_fields,
                }
            )

    for code, row in existing_by_code.items():
        if code not in next_by_code:
            diff_rows.append({"change": "removed", "code": code, "name": row.get("name")})

    return sorted(diff_rows, key=lambda item: (item["code"], item["change"]))


def restore_wbs_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    restored_rows: list[dict[str, Any]] = []
    for row in rows:
        restored = dict(row)
        for key in ("start_date", "finish_date"):
            value = restored.get(key)
            if isinstance(value, str) and value:
                restored[key] = date.fromisoformat(value)
        restored_rows.append(restored)
    return restored_rows


def template_phases(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    roots = [row for row in rows if not row.get("parent_code")]
    root_code = roots[0]["code"] if len(roots) == 1 else None
    phase_rows = [row for row in rows if row.get("parent_code") == root_code] if root_code else roots
    return [
        {
            "code": row["code"],
            "name": row["name"],
            "weight": row.get("weight"),
        }
        for row in phase_rows
        if row.get("code") and row.get("name")
    ]


def style_workbook_header(worksheet: Any, header_fill: PatternFill, header_font: Font, border: Border) -> None:
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_workbook_body(worksheet: Any, border: Border) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_column_widths(worksheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width


def add_agile_sample_sheets(
    workbook: Workbook,
    template: dict[str, Any],
    rows: list[dict[str, Any]],
    header_fill: PatternFill,
    header_font: Font,
    border: Border,
) -> None:
    delivery_mode = template_delivery_mode(template)
    if delivery_mode not in {"agile", "hybrid"}:
        return

    row_by_code = {row.get("code"): row for row in rows if row.get("code")}
    backlog = workbook.create_sheet("Agile Backlog")
    backlog.freeze_panes = "A2"
    backlog.append(
        [
            "유형",
            "제목",
            "설명",
            "Story Point",
            "우선순위",
            "상태",
            "담당자",
            "검토자",
            "Sprint",
            "연결 WBS 코드",
            "Acceptance Criteria",
            "Definition of Done",
        ]
    )
    style_workbook_header(backlog, header_fill, header_font, border)

    agile_row_count = 0
    for row in rows:
        metadata = normalize_metadata(row.get("metadata"))
        item_type = str(metadata.get("agile_type") or row.get("item_type") or "").strip()
        is_agile_item = item_type in AGILE_ITEM_TYPES or metadata.get("story_points") is not None
        if not is_agile_item:
            continue
        backlog.append(
            [
                item_type or "Task",
                row.get("name"),
                row.get("description") or metadata.get("notes"),
                metadata.get("story_points"),
                metadata.get("priority") or "Must",
                metadata.get("status") or "Backlog",
                row.get("owner"),
                metadata.get("reviewer"),
                metadata.get("sprint"),
                metadata.get("wbs_code") or (row.get("parent_code") if delivery_mode == "hybrid" else row.get("code")),
                metadata.get("acceptance_criteria"),
                metadata.get("definition_of_done"),
            ]
        )
        agile_row_count += 1

    style_workbook_body(backlog, border)
    set_column_widths(backlog, [14, 34, 32, 14, 14, 16, 16, 16, 18, 18, 44, 44])
    backlog.auto_filter.ref = backlog.dimensions
    type_validation = DataValidation(type="list", formula1='"Epic,Story,Task,Spike,Bug"', allow_blank=True)
    priority_validation = DataValidation(type="list", formula1='"Must,Should,Could,Wont"', allow_blank=True)
    status_validation = DataValidation(type="list", formula1='"Backlog,Ready,In Progress,Review,Done"', allow_blank=True)
    backlog.add_data_validation(type_validation)
    backlog.add_data_validation(priority_validation)
    backlog.add_data_validation(status_validation)
    type_validation.add("A2:A1000")
    priority_validation.add("E2:E1000")
    status_validation.add("F2:F1000")

    sprint_plan = workbook.create_sheet("Sprint Plan")
    sprint_plan.freeze_panes = "A2"
    sprint_plan.append(["Sprint", "목표", "시작일", "종료일", "계획 SP", "상태", "포함 항목"])
    style_workbook_header(sprint_plan, header_fill, header_font, border)

    sprints: dict[str, dict[str, Any]] = {}
    for row in rows:
        metadata = normalize_metadata(row.get("metadata"))
        sprint = metadata.get("sprint")
        if not sprint:
            continue
        entry = sprints.setdefault(
            str(sprint),
            {
                "goal": "",
                "start_date": None,
                "finish_date": None,
                "planned_points": 0,
                "planned_points_explicit": False,
                "status": "Planning",
                "items": [],
            },
        )
        if metadata.get("sprint_goal"):
            entry["goal"] = metadata.get("sprint_goal")
        if metadata.get("planned_points") is not None:
            entry["planned_points"] = metadata.get("planned_points")
            entry["planned_points_explicit"] = True
            entry["status"] = metadata.get("status") or entry["status"]
        if row.get("start_date") and not entry["start_date"]:
            entry["start_date"] = row.get("start_date")
        if row.get("finish_date") and not entry["finish_date"]:
            entry["finish_date"] = row.get("finish_date")
        if metadata.get("story_points") is not None:
            entry["items"].append(f"{row.get('name')} ({metadata.get('story_points')}SP)")
            if not entry.get("planned_points_explicit"):
                entry["planned_points"] = (entry.get("planned_points") or 0) + int(metadata.get("story_points") or 0)

    for sprint, entry in sprints.items():
        sprint_plan.append(
            [
                sprint,
                entry.get("goal"),
                entry.get("start_date"),
                entry.get("finish_date"),
                entry.get("planned_points"),
                entry.get("status"),
                ", ".join(entry.get("items") or []),
            ]
        )

    style_workbook_body(sprint_plan, border)
    set_column_widths(sprint_plan, [18, 46, 14, 14, 12, 16, 80])
    for row in sprint_plan.iter_rows(min_row=2):
        row[2].number_format = "yyyy-mm-dd"
        row[3].number_format = "yyyy-mm-dd"
    sprint_plan.auto_filter.ref = sprint_plan.dimensions

    if delivery_mode != "hybrid":
        return

    mapping = workbook.create_sheet("Hybrid Mapping")
    mapping.freeze_panes = "A2"
    mapping.append(["상위 WBS 코드", "상위 WBS 항목", "Agile 항목", "Sprint", "Story Point", "상태", "진행 기준"])
    style_workbook_header(mapping, header_fill, header_font, border)
    for row in rows:
        metadata = normalize_metadata(row.get("metadata"))
        if metadata.get("agile_type") not in AGILE_ITEM_TYPES and metadata.get("story_points") is None:
            continue
        wbs_code = metadata.get("wbs_code") or row.get("parent_code")
        parent = row_by_code.get(wbs_code)
        mapping.append(
            [
                wbs_code,
                parent.get("name") if parent else "",
                row.get("name"),
                metadata.get("sprint"),
                metadata.get("story_points"),
                metadata.get("status"),
                metadata.get("acceptance_criteria") or metadata.get("definition_of_done"),
            ]
        )
    style_workbook_body(mapping, border)
    set_column_widths(mapping, [18, 28, 34, 18, 14, 16, 56])
    mapping.auto_filter.ref = mapping.dimensions


def build_template_workbook(template: dict[str, Any], rows: list[dict[str, Any]]) -> BytesIO:
    parent_map = {row["code"]: row.get("parent_code") for row in rows if row.get("code")}
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "WBS"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1D1D1F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8DCE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.append([label for _, label in EXCEL_COLUMNS])
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        metadata = normalize_metadata(row.get("metadata"))
        start_date = row.get("start_date")
        finish_date = row.get("finish_date")
        worksheet.append(
            [
                parent_depth(row["code"], parent_map),
                row["code"],
                row.get("parent_code"),
                row["name"],
                row.get("item_type") or "작업",
                row.get("owner"),
                float(row["weight"]) if row.get("weight") is not None else None,
                start_date,
                finish_date,
                metadata.get("deliverable_type"),
                "Y" if metadata.get("inspection_required") else "N",
                metadata.get("progress_formula")
                or ("하위 단계 가중치 합산" if not row.get("parent_code") else "작업 완료율 x 가중치"),
                metadata.get("notes"),
            ]
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[7].number_format = "yyyy-mm-dd"
        row[8].number_format = "yyyy-mm-dd"

    widths = [8, 18, 18, 30, 14, 16, 10, 14, 14, 18, 12, 24, 28]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    worksheet.auto_filter.ref = worksheet.dimensions

    item_type_validation = DataValidation(
        type="list",
        formula1='"프로젝트,단계,산출물,작업,마일스톤,리스크,이슈,변경요청"',
        allow_blank=False,
    )
    inspection_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    worksheet.add_data_validation(item_type_validation)
    worksheet.add_data_validation(inspection_validation)
    item_type_validation.add("E2:E1000")
    inspection_validation.add("K2:K1000")

    guide = workbook.create_sheet("Guide")
    guide.append(["Template", template["name"]])
    guide.append(["Key", template["key"]])
    guide.append(["Project Type", template["project_type"]])
    guide.append(["Delivery Mode", template_delivery_mode(template)])
    guide.append(["Rule", "작업명은 필수입니다. WBS 코드는 비워두면 레벨과 행 순서 기준으로 자동 생성합니다."])
    if template_delivery_mode(template) in {"agile", "hybrid"}:
        guide.append(["Agile Sheets", "Agile Backlog와 Sprint Plan 시트에서 Epic/Story/Task, Story Point, Sprint 정보를 관리합니다."])
    if template_delivery_mode(template) == "hybrid":
        guide.append(["Hybrid Mapping", "상위 WBS 코드와 Agile 항목을 연결해 기준선 WBS와 Sprint 실행 현황을 함께 추적합니다."])
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 96
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    add_agile_sample_sheets(workbook, template, rows, header_fill, header_font, border)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_import_errors_workbook(job: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    issues = workbook.active
    issues.title = "Issues"
    issues.append(["Type", "Row", "Field", "Code", "Message"])
    for issue_type, entries in (("Error", job.get("errors") or []), ("Warning", job.get("warnings") or [])):
        for issue in entries:
            issues.append(
                [
                    issue_type,
                    issue.get("row"),
                    issue.get("field") or issue.get("parent_code"),
                    issue.get("code") or issue.get("parent_code"),
                    issue.get("message"),
                ]
            )

    diff = workbook.create_sheet("Diff")
    diff.append(["Change", "Code", "Name", "Field", "Before", "After"])
    for item in job.get("diff_rows") or []:
        fields = item.get("fields") or [{}]
        for field in fields:
            diff.append(
                [
                    item.get("change"),
                    item.get("code"),
                    item.get("name"),
                    field.get("field"),
                    field.get("before"),
                    field.get("after"),
                ]
            )

    rows = workbook.create_sheet("Rows")
    rows.append([label for _, label in EXCEL_COLUMNS])
    for row in job.get("preview_rows") or []:
        rows.append([row.get(key) for key, _ in EXCEL_COLUMNS])

    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor="1D1D1F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 22

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def fetch_template(connection: asyncpg.Connection, template_key: str) -> dict[str, Any] | None:
    record = await connection.fetchrow(
        """
        SELECT key, name, project_type, description, phases
        FROM wbs_templates
        WHERE key = $1
        """,
        template_key,
    )
    return normalize_record(record) if record else None


async def fetch_project(
    connection: asyncpg.Connection,
    project_id: UUID,
    tenant_id: str | None = None,
    *,
    for_update: bool = False,
) -> dict[str, Any] | None:
    params: list[Any] = [project_id]
    tenant_filter = ""
    if tenant_id is not None:
        params.append(tenant_id)
        tenant_filter = f" AND tenant_id = ${len(params)}"
    lock_clause = " FOR UPDATE" if for_update else ""
    record = await connection.fetchrow(
        f"""
        SELECT id, name, template_key, owner, status, start_date, delivery_mode,
               openproject_project_id, metadata, tenant_id, created_at, updated_at
        FROM wbs_projects
        WHERE id = $1{tenant_filter}{lock_clause}
        """,
        *params,
    )
    return normalize_record(record) if record else None


async def fetch_tenant_project(
    connection: asyncpg.Connection,
    request: Request,
    project_id: UUID,
    *,
    for_update: bool = False,
) -> dict[str, Any]:
    project = await fetch_project(
        connection,
        project_id,
        get_tenant_id(request),
        for_update=for_update,
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    project["_project_role"] = await get_effective_project_role(
        connection, get_tenant_id(request), project_id, getattr(request.state, "user", None)
    )
    if project["_project_role"] is None:
        # 비멤버에게는 프로젝트 존재 자체를 노출하지 않음
        raise HTTPException(status_code=404, detail="Project not found")
    return project


async def fetch_template_items(connection: asyncpg.Connection, template_key: str) -> list[dict[str, Any]]:
    records = await connection.fetch(
        """
        SELECT code, parent_code, name, item_type, owner, weight,
               start_date, finish_date, sort_order, metadata, tenant_id
        FROM wbs_template_items
        WHERE template_key = $1
        ORDER BY sort_order, code
        """,
        template_key,
    )
    return [normalize_record(record) for record in records]


async def fetch_project_wbs_items(connection: asyncpg.Connection, project_id: UUID) -> list[dict[str, Any]]:
    records = await connection.fetch(
        """
        SELECT code, parent_code, name, item_type, owner, weight,
               start_date, finish_date, sort_order, metadata, tenant_id
        FROM wbs_project_wbs_items
        WHERE project_id = $1
        ORDER BY sort_order, code
        """,
        project_id,
    )
    return [normalize_record(record) for record in records]


def work_item_status(row: dict[str, Any]) -> str:
    metadata = normalize_metadata(row.get("metadata"))
    raw = str(metadata.get("status") or row.get("status") or "").strip()
    if raw == "진행 중":
        raw = "진행중"
    if raw == "미시작":
        raw = "대기"
    if raw in WORK_ITEM_STATUSES:
        return raw
    finish = parse_date(row.get("finish_date"))
    if finish and finish < date.today():
        return "지연"
    return "대기"


def work_item_progress(row: dict[str, Any]) -> int:
    metadata = normalize_metadata(row.get("metadata"))
    try:
        value = float(metadata.get("progress", row.get("progress", 0)) or 0)
    except (TypeError, ValueError):
        value = 0
    return max(0, min(100, round(value)))


def work_item_identity_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def work_item_people_values(row: dict[str, Any], metadata: dict[str, Any] | None = None) -> list[str]:
    metadata = metadata or normalize_metadata(row.get("metadata"))
    values: list[str] = []
    for value in (row.get("owner"), metadata.get("reviewer"), metadata.get("approver"), metadata.get("team")):
        if not value:
            continue
        values.extend(part.strip() for part in re.split(r"[,;/|·]+", str(value)) if part.strip())
    return values


def work_item_user_tokens(user: dict[str, Any]) -> list[str]:
    email = str(user.get("email") or "")
    values = [user.get("display_name"), email, email.partition("@")[0]]
    return [work_item_identity_key(value) for value in values if work_item_identity_key(value)]


def work_item_matches_user(row: dict[str, Any], user: dict[str, Any]) -> bool:
    metadata = normalize_metadata(row.get("metadata"))
    haystack = " ".join(work_item_people_values(row, metadata)).lower()
    keys = [work_item_identity_key(value) for value in work_item_people_values(row, metadata)]
    for token in work_item_user_tokens(user):
        if token in keys or (token and token in work_item_identity_key(haystack)):
            return True
    return False


def serialize_work_item_row(row: dict[str, Any] | asyncpg.Record, project: dict[str, Any] | None = None) -> dict[str, Any]:
    item = normalize_record(row) if isinstance(row, asyncpg.Record) else dict(row)
    item["metadata"] = normalize_metadata(item.get("metadata"))
    item["status"] = work_item_status(item)
    item["progress"] = work_item_progress(item)
    item["project_id"] = str(item.get("project_id") or project.get("id") if project else item.get("project_id") or "")
    if project:
        item["project_name"] = project.get("name")
        item["project_status"] = project.get("status")
        item["template_key"] = project.get("template_key")
    return item


async def fetch_project_work_item(
    connection: asyncpg.Connection,
    project_id: UUID,
    item_code: str,
    *,
    for_update: bool = False,
) -> dict[str, Any] | None:
    lock_clause = " FOR UPDATE" if for_update else ""
    record = await connection.fetchrow(
        f"""
        SELECT project_id, code, parent_code, name, item_type, owner, weight,
               start_date, finish_date, sort_order, metadata, tenant_id
        FROM wbs_project_wbs_items
        WHERE project_id = $1 AND code = $2
        {lock_clause}
        """,
        project_id,
        item_code,
    )
    return normalize_record(record) if record else None


def append_work_item_log(metadata: dict[str, Any], key: str, entry: dict[str, Any], limit: int = 100) -> None:
    values = metadata.get(key)
    values = values if isinstance(values, list) else []
    values.append(entry)
    metadata[key] = values[-limit:]


def work_item_change_summary(before: dict[str, Any], after: dict[str, Any]) -> str:
    labels = {
        "name": "작업명",
        "owner": "담당자",
        "status": "상태",
        "progress": "진척률",
        "priority": "우선순위",
        "start_date": "시작일",
        "finish_date": "종료일",
        "reviewer": "검토자",
        "approver": "승인자",
        "team": "팀",
        "effort": "공수",
    }
    changes = []
    for key, label in labels.items():
        if str(before.get(key) or "") != str(after.get(key) or ""):
            changes.append(f"{label}: {before.get(key) or '-'} -> {after.get(key) or '-'}")
    return ", ".join(changes)


async def resolve_work_item_notification_users(
    connection: asyncpg.Connection,
    tenant_id: str,
    people: list[str],
) -> list[dict[str, Any]]:
    keys = [work_item_identity_key(value) for value in people if work_item_identity_key(value)]
    if not keys:
        return []
    users = await connection.fetch(
        """
        SELECT id, email, display_name, role, status
        FROM wbs_users
        WHERE tenant_id = $1 AND status = 'Active'
        """,
        tenant_id,
    )
    matched: dict[str, dict[str, Any]] = {}
    for user_record in users:
        user = normalize_record(user_record)
        user_keys = work_item_user_tokens(user)
        if any(key in user_keys or any(key and key in user_key for user_key in user_keys) for key in keys):
            matched[str(user["id"])] = user
    return list(matched.values())


async def notify_work_item_users(
    connection: asyncpg.Connection,
    *,
    tenant_id: str,
    project: dict[str, Any],
    row: dict[str, Any],
    metadata: dict[str, Any],
    event_type: str,
    title: str,
    body: str,
    only_people: list[str] | None = None,
) -> int:
    people = only_people or work_item_people_values(row, metadata)
    users = await resolve_work_item_notification_users(connection, tenant_id, people)
    sent = 0
    for user in users:
        user_id = safe_uuid(user.get("id"))
        if not user_id:
            continue
        await send_notification(
            connection,
            user_id=user_id,
            event_type=event_type,
            title=title,
            body=body,
            entity_type="work_item",
            entity_id=f"{project['id']}:{row['code']}",
            metadata={
                "project_id": str(project["id"]),
                "project_name": project.get("name"),
                "code": row.get("code"),
                "data_source": "internal_wbs",
            },
            email_to=user.get("email") if event_type in {"work_item.delayed", "work_item.approval"} else None,
        )
        sent += 1
    return sent


async def replace_project_wbs_items(
    connection: asyncpg.Connection,
    *,
    project_id: UUID,
    tenant_id: str,
    rows: list[dict[str, Any]],
    source_import_job_id: UUID | None = None,
) -> None:
    await connection.execute("DELETE FROM wbs_project_wbs_items WHERE project_id = $1", project_id)
    for index, row in enumerate(rows, start=1):
        existing_metadata = normalize_metadata(row.get("metadata"))
        metadata = {
            **existing_metadata,
            "deliverable_type": row.get("deliverable_type", existing_metadata.get("deliverable_type")),
            "inspection_required": row.get("inspection_required", existing_metadata.get("inspection_required", False)),
            "progress_formula": row.get("progress_formula")
            or existing_metadata.get("progress_formula")
            or ("하위 단계 가중치 합산" if not row.get("parent_code") else "작업 완료율 x 가중치"),
            "notes": row.get("notes", existing_metadata.get("notes")),
        }
        await connection.execute(
            """
            INSERT INTO wbs_project_wbs_items
              (project_id, tenant_id, code, parent_code, name, item_type, owner, weight,
               start_date, finish_date, sort_order, metadata, source_import_job_id)
            VALUES
              ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12::jsonb, $13)
            """,
            project_id,
            tenant_id,
            row["code"],
            row.get("parent_code"),
            row["name"],
            row.get("item_type") or "작업",
            row.get("owner"),
            row.get("weight"),
            row.get("start_date"),
            row.get("finish_date"),
            index,
            metadata,
            source_import_job_id,
        )


async def replace_template_items(
    connection: asyncpg.Connection,
    *,
    template_key: str,
    template_name: str,
    project_type: str,
    description: str,
    rows: list[dict[str, Any]],
) -> None:
    phases = template_phases(rows)
    await connection.execute(
        """
        INSERT INTO wbs_templates (key, name, project_type, description, phases)
        VALUES ($1, $2, $3, $4, $5::jsonb)
        ON CONFLICT (key) DO UPDATE
        SET name = EXCLUDED.name,
            project_type = EXCLUDED.project_type,
            description = EXCLUDED.description,
            phases = EXCLUDED.phases
        """,
        template_key,
        template_name,
        project_type,
        description,
        phases,
    )
    await connection.execute("DELETE FROM wbs_template_items WHERE template_key = $1", template_key)

    for index, row in enumerate(rows, start=1):
        existing_metadata = normalize_metadata(row.get("metadata"))
        metadata = {
            "deliverable_type": row.get("deliverable_type", existing_metadata.get("deliverable_type")),
            "inspection_required": row.get("inspection_required", existing_metadata.get("inspection_required", False)),
            "progress_formula": row.get("progress_formula")
            or existing_metadata.get("progress_formula")
            or ("하위 단계 가중치 합산" if not row.get("parent_code") else "작업 완료율 x 가중치"),
            "notes": row.get("notes", existing_metadata.get("notes")),
        }
        await connection.execute(
            """
            INSERT INTO wbs_template_items
              (template_key, code, parent_code, name, item_type, owner, weight,
               start_date, finish_date, sort_order, metadata)
            VALUES
              ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb)
            """,
            template_key,
            row["code"],
            row.get("parent_code"),
            row["name"],
            row.get("item_type") or "작업",
            row.get("owner"),
            row.get("weight"),
            row.get("start_date"),
            row.get("finish_date"),
            index,
            metadata,
        )

    version = await connection.fetchval(
        "SELECT COALESCE(max(version), 0) + 1 FROM wbs_template_versions WHERE template_key = $1",
        template_key,
    )
    await connection.execute(
        """
        INSERT INTO wbs_template_versions
          (template_key, version, template_name, project_type, description,
           item_count, snapshot_rows, metadata)
        VALUES
          ($1, $2, $3, $4, $5, $6, $7::jsonb, $8::jsonb)
        """,
        template_key,
        version,
        template_name,
        project_type,
        description,
        len(rows),
        [serialize_wbs_row(row) for row in rows],
        {"source": "replace_template_items"},
    )


async def prepare_template_import(
    connection: asyncpg.Connection,
    *,
    template_key: str,
    parsed_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    existing_rows = await fetch_template_items(connection, template_key)
    root_code = root_code_from_rows(existing_rows) or template_code_prefix(template_key)
    rows = assign_missing_wbs_codes(parsed_rows, root_code)
    errors, warnings = validate_wbs_rows(rows)
    warnings = [*warnings, *auto_code_warnings(rows)]
    serialized_rows = [serialize_wbs_row(row) for row in rows]
    diff_rows = build_wbs_diff_rows(existing_rows, rows)
    return rows, errors, warnings, serialized_rows, diff_rows


async def insert_import_job(
    connection: asyncpg.Connection,
    *,
    source_file: str,
    status: str,
    total_rows: int,
    accepted_rows: int,
    rejected_rows: int,
    errors: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    preview_rows: list[dict[str, Any]],
    diff_rows: list[dict[str, Any]] | None = None,
    template_version: int | None = None,
    template_key: str | None = None,
    template_name: str | None = None,
    project_type: str | None = None,
    description: str | None = None,
    applied: bool = False,
) -> asyncpg.Record:
    return await connection.fetchrow(
        f"""
        INSERT INTO wbs_import_jobs
          (source_file, template_key, template_name, project_type, description,
           status, template_version, total_rows, accepted_rows, rejected_rows,
           errors, warnings, preview_rows, diff_rows, applied_at)
        VALUES
          ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb, $12::jsonb,
           $13::jsonb, $14::jsonb, CASE WHEN $15::boolean THEN now() ELSE NULL END)
        RETURNING {IMPORT_JOB_RETURNING}
        """,
        source_file,
        template_key,
        template_name,
        project_type,
        description,
        status,
        template_version,
        total_rows,
        accepted_rows,
        rejected_rows,
        errors,
        warnings,
        preview_rows,
        diff_rows or [],
        applied,
    )


def sync_error_payload(error: Exception | HTTPException | str) -> dict[str, Any]:
    if isinstance(error, HTTPException):
        return {
            "status_code": error.status_code,
            "detail": error.detail,
        }
    if isinstance(error, str):
        return {"message": error}
    return {
        "type": error.__class__.__name__,
        "message": str(error),
    }


async def insert_sync_run(
    connection: asyncpg.Connection,
    *,
    project_id: UUID,
    mode: str,
    status: str,
    actor: str,
    dry_run: bool,
    create_work_packages: bool,
    validate_payloads: bool,
    total_rows: int = 0,
    pending_work_packages: int = 0,
    synced_work_packages: int = 0,
    created_work_packages: int = 0,
    openproject_project_id: str | None = None,
    metadata: dict[str, Any] | None = None,
    error: dict[str, Any] | None = None,
) -> dict[str, Any]:
    record = await connection.fetchrow(
        f"""
        WITH inserted AS (
          INSERT INTO wbs_sync_runs
            (project_id, mode, status, actor, engine, dry_run,
             create_work_packages, validate_payloads, total_rows,
             pending_work_packages, synced_work_packages, created_work_packages,
             openproject_project_id, metadata, error, completed_at)
          VALUES
            ($1, $2, $3, $4, 'openproject', $5, $6, $7, $8, $9, $10,
             $11, $12, $13::jsonb, $14::jsonb, now())
          RETURNING *
        )
        SELECT {SYNC_RUN_SELECT}
        FROM inserted s
        JOIN wbs_projects p ON p.id = s.project_id
        """,
        project_id,
        mode,
        status,
        actor,
        dry_run,
        create_work_packages,
        validate_payloads,
        total_rows,
        pending_work_packages,
        synced_work_packages,
        created_work_packages,
        openproject_project_id,
        metadata or {},
        error,
    )
    return normalize_record(record)


def baseline_snapshot_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    snapshot_rows: list[dict[str, Any]] = []
    for row in rows:
        snapshot_rows.append(
            serialize_wbs_row(
                {
                    "code": row.get("code"),
                    "parent_code": row.get("parent_code"),
                    "name": row.get("name"),
                    "item_type": row.get("item_type") or "작업",
                    "owner": row.get("owner"),
                    "weight": row.get("weight"),
                    "start_date": row.get("start_date"),
                    "finish_date": row.get("finish_date"),
                    "sort_order": row.get("sort_order"),
                    "metadata": normalize_metadata(row.get("metadata")),
                }
            )
        )
    return snapshot_rows


def baseline_total_weight(rows: list[dict[str, Any]]) -> float:
    root_weight = next(
        (row.get("weight") for row in rows if not row.get("parent_code") and row.get("weight") is not None),
        None,
    )
    if root_weight is not None:
        return float(root_weight)

    top_level_weights = [
        float(row["weight"])
        for row in rows
        if not row.get("parent_code") and row.get("weight") is not None
    ]
    if top_level_weights:
        return sum(top_level_weights)

    return sum(float(row["weight"]) for row in rows if row.get("weight") is not None)


def baseline_summary(baseline: dict[str, Any] | None) -> dict[str, Any]:
    if not baseline:
        return {
            "locked": False,
            "status": "Unlocked",
        }

    return {
        "locked": baseline.get("status") == "Locked",
        "id": baseline.get("id"),
        "approval_id": baseline.get("approval_id"),
        "version": baseline.get("version"),
        "status": baseline.get("status"),
        "template_key": baseline.get("template_key"),
        "template_name": baseline.get("template_name"),
        "item_count": baseline.get("item_count"),
        "total_weight": baseline.get("total_weight"),
        "locked_at": baseline.get("locked_at"),
        "metadata": normalize_metadata(baseline.get("metadata")),
    }


async def fetch_latest_project_baseline(
    connection: asyncpg.Connection,
    project_id: UUID,
) -> dict[str, Any] | None:
    record = await connection.fetchrow(
        f"""
        SELECT {BASELINE_SELECT}
        FROM wbs_project_baselines b
        JOIN wbs_projects p ON p.id = b.project_id
        WHERE b.project_id = $1
        ORDER BY b.version DESC
        LIMIT 1
        """,
        project_id,
    )
    return normalize_record(record) if record else None


async def create_project_baseline(
    connection: asyncpg.Connection,
    *,
    project_id: UUID,
    approval_id: UUID,
    template_key: str,
    actor: str,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    existing = await connection.fetchrow(
        f"""
        SELECT {BASELINE_SELECT}
        FROM wbs_project_baselines b
        JOIN wbs_projects p ON p.id = b.project_id
        WHERE b.approval_id = $1
        """,
        approval_id,
    )
    if existing:
        return normalize_record(existing)

    template = await fetch_template(connection, template_key)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found for baseline")
    project_rows = await fetch_project_wbs_items(connection, project_id)
    baseline_source = "project_wbs" if project_rows else "template"
    rows = project_rows or template_rows_or_phases(template, await fetch_template_items(connection, template_key))
    snapshot_rows = baseline_snapshot_rows(rows)
    version = await connection.fetchval(
        "SELECT COALESCE(max(version), 0) + 1 FROM wbs_project_baselines WHERE project_id = $1",
        project_id,
    )
    record = await connection.fetchrow(
        f"""
        WITH inserted AS (
          INSERT INTO wbs_project_baselines
            (project_id, approval_id, version, status, template_key, template_name,
             item_count, total_weight, snapshot_rows, metadata)
          VALUES
            ($1, $2, $3, 'Locked', $4, $5, $6, $7, $8::jsonb, $9::jsonb)
          RETURNING *
        )
        SELECT {BASELINE_SELECT}
        FROM inserted b
        JOIN wbs_projects p ON p.id = b.project_id
        """,
        project_id,
        approval_id,
        version,
        template["key"],
        template["name"],
        len(snapshot_rows),
        baseline_total_weight(rows),
        snapshot_rows,
        {
            "locked_by": actor,
            "source": "approval",
            "baseline_source": baseline_source,
            **(metadata or {}),
        },
    )
    return normalize_record(record)


async def record_project_sync_run(
    request: Request,
    *,
    project_id: UUID,
    plan: dict[str, Any],
    payload: ProjectSyncRequest,
    mode: str,
    status: str,
    created_work_packages: int = 0,
    openproject_project_id: str | None = None,
    metadata: dict[str, Any] | None = None,
    error: dict[str, Any] | None = None,
) -> dict[str, Any]:
    summary = normalize_metadata(plan.get("summary"))
    async with get_pool(request).acquire() as connection:
        sync_run = await insert_sync_run(
            connection,
            project_id=project_id,
            mode=mode,
            status=status,
            actor=payload.actor,
            dry_run=payload.dry_run,
            create_work_packages=payload.create_work_packages,
            validate_payloads=payload.validate_payloads,
            total_rows=int(summary.get("total_rows") or 0),
            pending_work_packages=int(summary.get("pending_work_packages") or 0),
            synced_work_packages=int(summary.get("synced_work_packages") or 0),
            created_work_packages=created_work_packages,
            openproject_project_id=openproject_project_id or normalize_metadata(plan.get("openproject")).get("project_id"),
            metadata=metadata,
            error=error,
        )
        await insert_audit_event(
            connection,
            request=request,
            event_type="pm_engine.sync_recorded",
            entity_type="project",
            entity_id=project_id,
            summary=f"PM engine sync {status}",
            metadata={
                "mode": mode,
                "dry_run": payload.dry_run,
                "engine": sync_run.get("engine"),
                "sync_run_id": str(sync_run.get("id")) if sync_run.get("id") else None,
            },
        )
        return sync_run


async def fetch_import_job_for_update(connection: asyncpg.Connection, job_id: UUID) -> dict[str, Any] | None:
    record = await connection.fetchrow(
        f"""
        SELECT {IMPORT_JOB_RETURNING}
        FROM wbs_import_jobs
        WHERE id = $1
        FOR UPDATE
        """,
        job_id,
    )
    return normalize_record(record) if record else None


def import_job_response(record: asyncpg.Record, *, include_rows: bool = False, row_limit: int = 50) -> dict[str, Any]:
    job = normalize_record(record)
    preview_rows = job.get("preview_rows") or []
    diff_rows = job.get("diff_rows") or []
    job["preview_count"] = len(preview_rows)
    job["diff_count"] = len(diff_rows)
    if include_rows:
        job["rows"] = preview_rows[:row_limit]
        job["diff_rows"] = diff_rows[:row_limit]
    job.pop("preview_rows", None)
    return job


async def fetch_approval(
    connection: asyncpg.Connection,
    approval_id: UUID,
    tenant_id: str | None = None,
) -> dict[str, Any] | None:
    params: list[Any] = [approval_id]
    tenant_filter = ""
    if tenant_id is not None:
        params.append(tenant_id)
        tenant_filter = f" AND a.tenant_id = ${len(params)}"
    record = await connection.fetchrow(
        f"""
        SELECT {APPROVAL_SELECT}
        FROM wbs_approval_requests a
        JOIN wbs_projects p ON p.id = a.project_id
        WHERE a.id = $1{tenant_filter}
        """,
        *params,
    )
    return normalize_record(record) if record else None


async def fetch_approval_for_update(
    connection: asyncpg.Connection,
    approval_id: UUID,
    tenant_id: str | None = None,
) -> dict[str, Any] | None:
    params: list[Any] = [approval_id]
    tenant_filter = ""
    if tenant_id is not None:
        params.append(tenant_id)
        tenant_filter = f" AND a.tenant_id = ${len(params)}"
    record = await connection.fetchrow(
        f"""
        SELECT {APPROVAL_SELECT}
        FROM wbs_approval_requests a
        JOIN wbs_projects p ON p.id = a.project_id
        WHERE a.id = $1{tenant_filter}
        FOR UPDATE OF a
        """,
        *params,
    )
    return normalize_record(record) if record else None


class OpenProjectClient:
    def __init__(
        self,
        base_url: str,
        api_token: str,
        auth_mode: str = "bearer",
        host_header: str = OPENPROJECT_HOST_HEADER,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.api_token = api_token
        self.auth_mode = auth_mode
        self.host_header = host_header

    def request_options(self) -> dict[str, Any]:
        headers = {"Accept": "application/hal+json", "Content-Type": "application/json"}
        if self.host_header:
            headers["Host"] = self.host_header
        options: dict[str, Any] = {"headers": headers}
        if self.api_token:
            if self.auth_mode == "basic":
                options["auth"] = ("apikey", self.api_token)
            else:
                headers["Authorization"] = f"Bearer {self.api_token}"
        return options

    async def request(self, method: str, path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        try:
            async with httpx.AsyncClient(timeout=20.0, follow_redirects=False) as client:
                response = await client.request(
                    method,
                    f"{self.base_url}{path}",
                    json=payload,
                    **self.request_options(),
                )
        except httpx.RequestError as exc:
            raise HTTPException(
                status_code=502,
                detail={
                    "message": "OpenProject API is unreachable",
                    "base_url": self.base_url,
                    "error": str(exc),
                },
            ) from exc

        if response.status_code >= 400:
            raise HTTPException(
                status_code=502,
                detail={
                    "message": "OpenProject API request failed",
                    "status_code": response.status_code,
                    "response": self.response_detail(response),
                },
            )
        if not response.content:
            return {}
        try:
            return response.json()
        except ValueError as exc:
            raise HTTPException(status_code=502, detail="OpenProject API returned invalid JSON") from exc

    @staticmethod
    def response_detail(response: httpx.Response) -> Any:
        try:
            return response.json()
        except ValueError:
            return response.text[:1000]

    async def probe(
        self,
        *,
        name: str,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        accepted_error_statuses: set[int] | None = None,
    ) -> dict[str, Any]:
        result: dict[str, Any] = {
            "name": name,
            "method": method,
            "path": path,
            "base_url": self.base_url,
        }
        try:
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=False) as client:
                response = await client.request(
                    method,
                    f"{self.base_url}{path}",
                    json=payload,
                    **self.request_options(),
                )
        except httpx.RequestError as exc:
            return {
                **result,
                "ok": False,
                "status": "fail",
                "message": "OpenProject endpoint is unreachable",
                "error": str(exc),
            }

        result["status_code"] = response.status_code
        accepted_error_statuses = accepted_error_statuses or set()
        if 200 <= response.status_code < 400 or response.status_code in accepted_error_statuses:
            result.update({"ok": True, "status": "pass"})
            if response.content:
                detail = self.response_detail(response)
                if isinstance(detail, dict):
                    result["resource_type"] = detail.get("_type")
                    if detail.get("id"):
                        result["resource_id"] = detail["id"]
                    if detail.get("name"):
                        result["resource_name"] = detail["name"]
                    if response.status_code >= 400:
                        result["message"] = detail.get("message", "Endpoint is reachable")
                        result["response"] = detail
            return result

        return {
            **result,
            "ok": False,
            "status": "fail",
            "message": "OpenProject endpoint returned an error",
            "response": self.response_detail(response),
        }

    async def create_project(self, project: dict[str, Any], identifier: str) -> dict[str, Any]:
        return await self.request(
            "POST",
            "/api/v3/projects",
            {
                "_type": "Project",
                "name": project["name"],
                "identifier": identifier,
                "active": True,
            },
        )

    def build_version_payload(
        self,
        *,
        openproject_project_id: str,
        sprint: dict[str, Any],
    ) -> dict[str, Any]:
        status = "closed" if str(sprint.get("status") or "").lower() == "closed" else "open"
        payload: dict[str, Any] = {
            "_type": "Version",
            "name": str(sprint.get("name") or "Sprint")[:255],
            "description": {
                "format": "plain",
                "raw": str(sprint.get("goal") or ""),
            },
            "status": status,
            "sharing": "none",
            "_links": {
                "definingProject": {"href": f"/api/v3/projects/{openproject_project_id}"},
            },
        }
        if sprint.get("start_date"):
            payload["startDate"] = sprint["start_date"]
        if sprint.get("end_date"):
            payload["endDate"] = sprint["end_date"]
        return payload

    async def create_version_from_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        return await self.request("POST", "/api/v3/versions", payload)

    def build_work_package_payload(
        self,
        *,
        openproject_project_id: str,
        row: dict[str, Any],
        parent_href: str | None = None,
    ) -> dict[str, Any]:
        type_map = parse_json_object(OPENPROJECT_TYPE_MAP_JSON)
        type_id = type_map.get(row.get("item_type") or "") or OPENPROJECT_DEFAULT_TYPE_ID
        links: dict[str, Any] = {
            "project": {"href": f"/api/v3/projects/{openproject_project_id}"},
        }
        if type_id:
            links["type"] = {"href": f"/api/v3/types/{type_id}"}
        if parent_href and OPENPROJECT_SYNC_PARENT_LINKS:
            links["parent"] = {"href": parent_href}

        metadata = normalize_metadata(row.get("metadata"))
        description_lines = [
            f"WBS code: {row['code']}",
            f"Type: {row.get('item_type') or '작업'}",
        ]
        if row.get("owner"):
            description_lines.append(f"Owner: {row['owner']}")
        if row.get("weight") is not None:
            description_lines.append(f"Weight: {row['weight']}")
        if metadata.get("deliverable_type"):
            description_lines.append(f"Deliverable: {metadata['deliverable_type']}")

        payload: dict[str, Any] = {
            "_type": "WorkPackage",
            "subject": f"{row['code']} {row['name']}"[:255],
            "description": {
                "format": "markdown",
                "raw": "\n".join(description_lines),
            },
            "_links": links,
        }
        if row.get("start_date"):
            payload["startDate"] = row["start_date"]
        if row.get("finish_date"):
            payload["dueDate"] = row["finish_date"]

        return payload

    async def validate_work_package_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        response = await self.request("POST", "/api/v3/work_packages/form", payload)
        embedded = normalize_metadata(response.get("_embedded"))
        validation_errors = normalize_metadata(embedded.get("validationErrors"))
        if validation_errors:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "OpenProject rejected the work package payload during form validation",
                    "validation_errors": validation_errors,
                },
            )
        return response

    async def create_work_package_from_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        return await self.request("POST", "/api/v3/work_packages?notify=false", payload)

    async def update_work_package_dates(
        self,
        *,
        work_package_id: str,
        start_date: date | None,
        finish_date: date | None,
    ) -> dict[str, Any]:
        current = await self.request("GET", f"/api/v3/work_packages/{work_package_id}")
        lock_version = current.get("lockVersion")
        if lock_version is None:
            raise HTTPException(
                status_code=502,
                detail="OpenProject work package lockVersion is missing",
            )
        return await self.request(
            "PATCH",
            f"/api/v3/work_packages/{work_package_id}",
            {
                "lockVersion": lock_version,
                "startDate": start_date.isoformat() if start_date else None,
                "dueDate": finish_date.isoformat() if finish_date else None,
            },
        )

    async def create_work_package(
        self,
        *,
        openproject_project_id: str,
        row: dict[str, Any],
        parent_href: str | None = None,
    ) -> dict[str, Any]:
        payload = self.build_work_package_payload(
            openproject_project_id=openproject_project_id,
            row=row,
            parent_href=parent_href,
        )
        return await self.create_work_package_from_payload(payload)


def openproject_engine_status() -> dict[str, Any]:
    return {
        "adapter": "openproject",
        "base_url": OPENPROJECT_BASE_URL,
        "host_header_configured": bool(OPENPROJECT_HOST_HEADER),
        "enabled": OPENPROJECT_SYNC_ENABLED,
        "token_configured": bool(OPENPROJECT_API_TOKEN),
        "auth_mode": OPENPROJECT_AUTH_MODE,
        "default_type_configured": bool(OPENPROJECT_DEFAULT_TYPE_ID),
        "type_map_configured": bool(parse_json_object(OPENPROJECT_TYPE_MAP_JSON)),
        "parent_links": OPENPROJECT_SYNC_PARENT_LINKS,
    }


def pm_engine_status(setting_value: dict[str, Any] | None = None) -> dict[str, Any]:
    setting_value = setting_value or {}
    runtime = openproject_engine_status()
    adapter = PM_ENGINE_ADAPTER or setting_value.get("adapter") or runtime["adapter"]
    portal_enabled = bool(setting_value.get("portal_enabled", setting_value.get("enabled", True)))
    if adapter in {"disabled", "none", "internal"}:
        portal_enabled = False
    display_name = (
        "내부 WBS Only"
        if not portal_enabled
        else "Mock PM Engine"
        if adapter == "mock"
        else setting_value.get("display_name") or "OpenProject"
    )
    runtime_enabled = portal_enabled and (runtime["enabled"] or adapter == "mock")
    actual_sync_ready = portal_enabled and (adapter == "mock" or (runtime["enabled"] and bool(OPENPROJECT_API_TOKEN)))
    return {
        **runtime,
        "adapter": adapter,
        "display_name": display_name,
        "provider": runtime["adapter"],
        "mode": setting_value.get("mode") or ("internal-wbs" if not portal_enabled else "ce-api-adapter"),
        "dependency_boundary": setting_value.get("dependency_boundary") or "pm-engine-api",
        "actual_sync_control": setting_value.get("actual_sync_control") or "OPENPROJECT_SYNC_ENABLED",
        "portal_enabled": portal_enabled,
        "enabled": runtime_enabled,
        "capabilities": {
            "preflight": True,
            "dry_run": True,
            "actual_sync": actual_sync_ready,
            "work_package_payload_validation": True,
            "hierarchy_parent_links": portal_enabled and runtime["parent_links"],
            "mock_adapter": portal_enabled and adapter == "mock",
        },
        "runtime": runtime,
    }


def openproject_preflight_check(name: str, status: str, message: str, **extra: Any) -> dict[str, Any]:
    return {
        "name": name,
        "status": status,
        "ok": status == "pass",
        "message": message,
        **extra,
    }


async def run_openproject_preflight() -> dict[str, Any]:
    if PM_ENGINE_ADAPTER == "mock":
        return {
            "engine": pm_engine_status({"adapter": "mock", "display_name": "Mock PM Engine"}),
            "state": "ready",
            "ready_for_actual_sync": True,
            "checks": [
                openproject_preflight_check(
                    "pm_engine_adapter",
                    "pass",
                    "Mock PM engine adapter is ready for local product validation",
                ),
                openproject_preflight_check(
                    "external_api",
                    "skip",
                    "OpenProject API calls are skipped by the mock adapter",
                ),
            ],
        }

    client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
    checks: list[dict[str, Any]] = []

    api_root_check = await client.probe(
        name="api_root",
        method="GET",
        path="/api/v3",
        accepted_error_statuses={401, 403},
    )
    checks.append(api_root_check)
    checks.append(
        openproject_preflight_check(
            "sync_enabled",
            "pass" if OPENPROJECT_SYNC_ENABLED else "warn",
            "Actual OpenProject sync is enabled"
            if OPENPROJECT_SYNC_ENABLED
            else "Actual OpenProject sync is disabled; dry-run and planning endpoints remain available",
        )
    )
    checks.append(
        openproject_preflight_check(
            "api_token",
            "pass" if OPENPROJECT_API_TOKEN else "warn",
            "OPENPROJECT_API_TOKEN is configured"
            if OPENPROJECT_API_TOKEN
            else "OPENPROJECT_API_TOKEN is not configured; actual sync will be blocked",
        )
    )

    if OPENPROJECT_API_TOKEN:
        checks.append(await client.probe(name="authenticated_user", method="GET", path="/api/v3/users/me"))
    else:
        checks.append(
            openproject_preflight_check(
                "authenticated_user",
                "skip",
                "Skipped because OPENPROJECT_API_TOKEN is not configured",
            )
        )

    failed = any(check.get("status") == "fail" for check in checks)
    api_root_failed = any(check.get("name") == "api_root" and check.get("status") == "fail" for check in checks)
    auth_failed = any(
        check.get("name") == "authenticated_user" and check.get("status") == "fail" for check in checks
    )
    ready = (
        not failed
        and OPENPROJECT_SYNC_ENABLED
        and bool(OPENPROJECT_API_TOKEN)
        and any(check.get("name") == "authenticated_user" and check.get("ok") for check in checks)
    )
    if api_root_failed:
        state = "offline"
    elif auth_failed:
        state = "auth_failed"
    elif failed:
        state = "blocked"
    elif ready:
        state = "ready"
    else:
        state = "dry_run_only"

    return {
        "engine": pm_engine_status(),
        "state": state,
        "ready_for_actual_sync": ready,
        "checks": checks,
    }


def build_openproject_sync_plan(
    project: dict[str, Any],
    template: dict[str, Any],
    rows: list[dict[str, Any]],
    baseline: dict[str, Any] | None = None,
    policy: dict[str, Any] | None = None,
    sprints: list[dict[str, Any] | asyncpg.Record] | None = None,
) -> dict[str, Any]:
    project_uuid = str(project["id"])
    identifier = normalize_openproject_identifier(
        f"{project['name']}-{project_uuid[:8]}",
        f"wbs-{project_uuid[:8]}",
    )
    metadata = normalize_metadata(project.get("metadata"))
    engine_metadata = normalize_metadata(metadata.get("pm_engine"))
    synced_work_packages = normalize_metadata(engine_metadata.get("work_packages"))
    operation_policy = normalize_project_operation_policy_value(policy or {}, project.get("tenant_id") or DEFAULT_TENANT_ID)
    synced_versions = normalize_metadata(engine_metadata.get("versions"))

    planned_rows = []
    for row in rows:
        planned_rows.append(
            {
                "code": row["code"],
                "parent_code": row.get("parent_code"),
                "name": row["name"],
                "item_type": row.get("item_type") or "작업",
                "owner": row.get("owner"),
                "weight": row.get("weight"),
                "subject": f"{row['code']} {row['name']}"[:255],
                "start_date": row.get("start_date"),
                "finish_date": row.get("finish_date"),
                "metadata": normalize_metadata(row.get("metadata")),
                "already_synced": row["code"] in synced_work_packages,
            }
        )

    planned_versions = []
    if operation_policy.get("openproject_sprint_version_sync"):
        for sprint in sprints or []:
            sprint_data = agile_sprint_response(sprint) if isinstance(sprint, asyncpg.Record) else dict(sprint)
            planned_versions.append(
                {
                    "sprint_id": str(sprint_data.get("id")),
                    "name": sprint_data.get("name"),
                    "goal": sprint_data.get("goal"),
                    "start_date": sprint_data.get("start_date"),
                    "end_date": sprint_data.get("end_date"),
                    "status": sprint_data.get("status"),
                    "already_synced": sprint_data.get("name") in synced_versions,
                }
            )

    return {
        "engine": pm_engine_status(),
        "project": project,
        "template": template,
        "wbs_source": project.get("wbs_source", "template"),
        "operation_policy": {
            "story_point_mode": operation_policy["story_point_mode"],
            "sprint_length_policy": operation_policy["sprint_length_policy"],
            "dod_management": operation_policy["dod_management"],
            "openproject_sprint_version_sync": operation_policy["openproject_sprint_version_sync"],
        },
        "openproject": {
            "project_id": project.get("openproject_project_id") or engine_metadata.get("project_id"),
            "project_identifier": engine_metadata.get("project_identifier") or identifier,
            "project_already_synced": bool(project.get("openproject_project_id") or engine_metadata.get("project_id")),
        },
        "baseline": baseline_summary(baseline),
        "rows": planned_rows,
        "sprint_versions": planned_versions,
        "summary": {
            "total_rows": len(planned_rows),
            "pending_work_packages": len([row for row in planned_rows if not row["already_synced"]]),
            "synced_work_packages": len([row for row in planned_rows if row["already_synced"]]),
            "pending_versions": len([row for row in planned_versions if not row["already_synced"]]),
            "synced_versions": len([row for row in planned_versions if row["already_synced"]]),
        },
    }


def build_openproject_payload_sample(
    project: dict[str, Any],
    rows: list[dict[str, Any]],
) -> dict[str, Any] | None:
    metadata = normalize_metadata(project.get("metadata"))
    engine_metadata = normalize_metadata(metadata.get("pm_engine"))
    work_packages = normalize_metadata(engine_metadata.get("work_packages"))
    openproject_project_id = project.get("openproject_project_id") or engine_metadata.get("project_id")
    sample_project_id = openproject_project_id or "OPENPROJECT_PROJECT_ID"
    href_by_code = {
        code: item.get("href")
        for code, item in work_packages.items()
        if isinstance(item, dict) and item.get("href")
    }
    sample_row = next((row for row in rows if row.get("code") not in work_packages), rows[0] if rows else None)
    if not sample_row:
        return None

    client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
    payload = client.build_work_package_payload(
        openproject_project_id=str(sample_project_id),
        row=sample_row,
        parent_href=href_by_code.get(sample_row.get("parent_code")),
    )
    return {
        "row_code": sample_row["code"],
        "project_id_source": "stored" if openproject_project_id else "placeholder",
        "form_endpoint": "/api/v3/work_packages/form",
        "create_endpoint": "/api/v3/work_packages?notify=false",
        "payload": payload,
    }


def operation_check(
    key: str,
    label: str,
    status: str,
    message: str,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "key": key,
        "label": label,
        "status": status,
        "message": message,
        "metadata": metadata or {},
    }


def backup_health_check() -> dict[str, Any]:
    backup_files = sorted(BACKUP_DIR.glob("*.dump"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not backup_files:
        return operation_check(
            "backup_rehearsal",
            "Backup rehearsal",
            "warn",
            f"No PostgreSQL backup found in {BACKUP_DIR}",
            {"backup_dir": str(BACKUP_DIR), "latest_backup": None},
        )

    latest = backup_files[0]
    latest_at = datetime.fromtimestamp(latest.stat().st_mtime, tz=timezone.utc)
    age_hours = round((datetime.now(timezone.utc) - latest_at).total_seconds() / 3600, 1)
    status = "pass" if age_hours <= 168 else "warn"
    message = f"Latest backup {latest.name}, {age_hours}h old"
    return operation_check(
        "backup_rehearsal",
        "Backup rehearsal",
        status,
        message,
        {
            "backup_dir": str(BACKUP_DIR),
            "latest_backup": latest.name,
            "latest_backup_at": latest_at.isoformat(),
            "age_hours": age_hours,
        },
    )


def template_rows_or_phases(template: dict[str, Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if rows:
        return rows
    return [
        {
            "code": phase["code"],
            "parent_code": None,
            "name": phase["name"],
            "item_type": "단계",
            "owner": "PMO",
            "weight": phase.get("weight"),
            "start_date": None,
            "finish_date": None,
            "metadata": {},
        }
        for phase in template.get("phases", [])
    ]


async def fetch_project_sync_context(
    request: Request,
    project_id: UUID,
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, project_id)
        template = await fetch_template(connection, project["template_key"])
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        project_rows = await fetch_project_wbs_items(connection, project_id)
        rows = project_rows or await fetch_template_items(connection, project["template_key"])
        project["wbs_source"] = "project" if project_rows else "template"

    return project, template, template_rows_or_phases(template, rows)


def prometheus_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("\n", "\\n").replace('"', '\\"')


def prometheus_metric(name: str, value: int | float, labels: dict[str, str] | None = None) -> str:
    if labels:
        label_text = ",".join(f'{key}="{prometheus_escape(str(label_value))}"' for key, label_value in labels.items())
        return f"{name}{{{label_text}}} {value}"
    return f"{name} {value}"


@app.get("/health")
async def health(request: Request) -> dict[str, str]:
    async with get_pool(request).acquire() as connection:
        await connection.fetchval("SELECT 1")
    return {
        "status": "ok",
        "database": "postgresql",
        "openproject_base_url": OPENPROJECT_BASE_URL,
    }


@app.post("/api/auth/login")
async def login(payload: LoginRequest, request: Request) -> dict[str, Any]:
    login_identifier = payload.email.strip().lower()
    normalized_email, login_alias = resolve_login_alias(login_identifier)
    alias_used = login_alias is not None
    expires_at = datetime.now(timezone.utc) + timedelta(hours=SESSION_TTL_HOURS)
    token = secrets.token_urlsafe(32)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            await connection.execute("DELETE FROM wbs_user_sessions WHERE expires_at <= now()")
            user_record = await connection.fetchrow(
                f"""
                SELECT {USER_SELECT_G}, u.password_hash
                FROM wbs_users u
                LEFT JOIN wbs_user_groups g
                  ON g.id = u.group_id
                 AND g.tenant_id = u.tenant_id
                WHERE u.email = $1
                  AND u.tenant_id = $2
                """,
                normalized_email,
                get_tenant_id(request),
            )
            account_locked = bool(
                user_record
                and user_record["locked_until"]
                and user_record["locked_until"] > datetime.now(timezone.utc)
            )
            password_ok = False
            if user_record and user_record["status"] == "Active" and not account_locked:
                if AUTH_BACKEND == "ldap" and not alias_used:
                    password_ok = await asyncio.get_event_loop().run_in_executor(
                        None, _try_ldap_login, normalized_email, payload.password
                    )
                else:
                    stored_password_ok = await connection.fetchval(
                        "SELECT $1 = crypt($2, $1)",
                        user_record["password_hash"],
                        payload.password,
                    )
                    alias_password = login_alias.get("password") if login_alias else ""
                    alias_password_ok = bool(alias_password and payload.password == alias_password)
                    password_ok = bool(stored_password_ok or alias_password_ok)

            if account_locked:
                await insert_audit_event(
                    connection,
                    request=request,
                    event_type="auth.login_locked",
                    entity_type="user",
                    entity_id=user_record["id"],
                    summary="Login blocked because account is locked",
                    metadata={
                        "email": normalized_email,
                        "login_identifier": login_identifier,
                        "alias_used": alias_used,
                        "locked_until": user_record["locked_until"].isoformat(),
                    },
                    actor_email=normalized_email,
                    actor_role=user_record["role"],
                )
                raise HTTPException(status_code=423, detail="Account is temporarily locked")

            if not password_ok:
                failed_count = int(user_record["failed_login_count"] or 0) + 1 if user_record else 1
                locked_until = (
                    datetime.now(timezone.utc) + timedelta(minutes=LOGIN_LOCK_MINUTES)
                    if user_record and failed_count >= LOGIN_FAILURE_LIMIT
                    else None
                )
                if user_record:
                    await connection.execute(
                        """
                        UPDATE wbs_users
                        SET failed_login_count = $2,
                            locked_until = $3,
                            updated_at = now()
                        WHERE id = $1
                        """,
                        user_record["id"],
                        failed_count,
                        locked_until,
                    )
                await insert_audit_event(
                    connection,
                    request=request,
                    event_type="auth.login_failed",
                    entity_type="user",
                    entity_id=user_record["id"] if user_record else normalized_email,
                    summary="Login failed",
                    metadata={
                        "email": normalized_email,
                        "login_identifier": login_identifier,
                        "alias_used": alias_used,
                        "failed_login_count": failed_count,
                        "locked_until": locked_until.isoformat() if locked_until else None,
                    },
                    actor_email=normalized_email,
                    actor_role=user_record["role"] if user_record else None,
                )
                raise HTTPException(status_code=401, detail="Invalid email or password")

            await connection.execute(
                """
                UPDATE wbs_users
                SET last_login_at = now(),
                    failed_login_count = 0,
                    locked_until = NULL,
                    updated_at = now()
                WHERE id = $1
                """,
                user_record["id"],
            )
            await connection.execute(
                """
                INSERT INTO wbs_user_sessions (token, user_id, expires_at)
                VALUES ($1, $2, $3)
                """,
                token,
                user_record["id"],
                expires_at,
            )
            updated_user = await connection.fetchrow(
                f"""
                SELECT {USER_SELECT_G}
                FROM wbs_users u
                LEFT JOIN wbs_user_groups g
                  ON g.id = u.group_id
                 AND g.tenant_id = u.tenant_id
                WHERE u.id = $1
                """,
                user_record["id"],
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="auth.login",
                entity_type="user",
                entity_id=user_record["id"],
                summary="User logged in",
                metadata={
                    "login_identifier": login_identifier,
                    "alias_used": alias_used,
                },
                actor=user_response(updated_user),
            )

    return {
        "token": token,
        "expires_at": expires_at.isoformat(),
        "user": user_response(updated_user),
        "tenant": getattr(request.state, "tenant", {"id": get_tenant_id(request), "status": "Active"}),
    }


@app.get("/api/auth/me")
async def current_user(request: Request) -> dict[str, Any]:
    return {
        "user": user_response(request.state.user),
        "tenant": getattr(request.state, "tenant", {"id": get_tenant_id(request), "status": "Active"}),
    }


@app.post("/api/auth/logout")
async def logout(request: Request) -> dict[str, str]:
    token = auth_token_from_request(request)
    if token:
        async with get_pool(request).acquire() as connection:
            await connection.execute("DELETE FROM wbs_user_sessions WHERE token = $1", token)
            await insert_audit_event(
                connection,
                request=request,
                event_type="auth.logout",
                entity_type="user",
                entity_id=request.state.user["id"],
                summary="User logged out",
            )
    return {"status": "ok"}


@app.post("/api/auth/password")
async def change_password(payload: PasswordChangeRequest, request: Request) -> dict[str, Any]:
    token = auth_token_from_request(request)
    current = request.state.user

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            password_ok = await connection.fetchval(
                """
                SELECT password_hash = crypt($2, password_hash)
                FROM wbs_users
                WHERE id = $1 AND status = 'Active'
                """,
                safe_uuid(current["id"]),
                payload.current_password,
            )
            if not password_ok:
                await insert_audit_event(
                    connection,
                    request=request,
                    event_type="auth.password_change_failed",
                    entity_type="user",
                    entity_id=current["id"],
                    summary="Password change failed",
                    metadata={"reason": "current_password_mismatch"},
                )
                raise HTTPException(status_code=401, detail="Current password is invalid")

            updated = await connection.fetchrow(
                f"""
                WITH updated AS (
                  UPDATE wbs_users
                  SET password_hash = crypt($2, gen_salt('bf')),
                      must_change_password = false,
                      password_changed_at = now(),
                      updated_at = now()
                  WHERE id = $1
                  RETURNING id
                )
                SELECT {USER_SELECT_G}
                FROM updated
                JOIN wbs_users u ON u.id = updated.id
                LEFT JOIN wbs_user_groups g
                  ON g.id = u.group_id
                 AND g.tenant_id = u.tenant_id
                """,
                safe_uuid(current["id"]),
                payload.new_password,
            )
            if token:
                await connection.execute(
                    "DELETE FROM wbs_user_sessions WHERE user_id = $1 AND token <> $2",
                    safe_uuid(current["id"]),
                    token,
                )
            await insert_audit_event(
                connection,
                request=request,
                event_type="auth.password_changed",
                entity_type="user",
                entity_id=current["id"],
                summary="Password changed",
                metadata={"other_sessions_revoked": True},
            )

    return {"status": "ok", "user": user_response(updated)}


@app.get("/api/users")
async def list_users(request: Request) -> list[dict[str, Any]]:
    require_admin_role(request)
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            """
            SELECT u.id, u.email, u.display_name, u.role, u.status,
                   u.tenant_id, u.failed_login_count, u.locked_until, u.must_change_password,
                   u.last_login_at, u.password_changed_at, u.created_at, u.updated_at,
                   u.group_id, g.name AS group_name, g.status AS group_status,
                   count(s.token)::integer AS active_sessions
            FROM wbs_users u
            JOIN wbs_user_groups g
              ON g.id = u.group_id
             AND g.tenant_id = u.tenant_id
            LEFT JOIN wbs_user_sessions s
              ON s.user_id = u.id
             AND s.expires_at > now()
            WHERE u.tenant_id = $1
            GROUP BY u.id, g.name, g.status
            ORDER BY
              CASE u.role WHEN 'admin' THEN 0 WHEN 'pmo' THEN 1 ELSE 2 END,
              u.email
            """,
            tid,
        )
    return [managed_user_response(record) for record in records]


@app.get("/api/user-groups")
async def list_user_groups(request: Request) -> list[dict[str, Any]]:
    require_admin_role(request)
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        await ensure_default_user_group(connection, tid)
        records = await connection.fetch(
            """
            SELECT g.id, g.tenant_id, g.name, g.description, g.status, g.metadata,
                   g.created_at, g.updated_at, count(u.id)::integer AS user_count
            FROM wbs_user_groups g
            LEFT JOIN wbs_users u
              ON u.group_id = g.id
             AND u.tenant_id = g.tenant_id
            WHERE g.tenant_id = $1
            GROUP BY g.id
            ORDER BY CASE WHEN g.name = $2 THEN 0 ELSE 1 END, lower(g.name)
            """,
            tid,
            DEFAULT_USER_GROUP_NAME,
        )
    return [user_group_response(record) for record in records]


@app.post("/api/user-groups", status_code=201)
async def create_user_group(payload: UserGroupCreate, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    tid = get_tenant_id(request)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="User group name is required")
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            exists = await connection.fetchval(
                "SELECT EXISTS (SELECT 1 FROM wbs_user_groups WHERE tenant_id = $1 AND lower(name) = lower($2))",
                tid,
                name,
            )
            if exists:
                raise HTTPException(status_code=409, detail="User group already exists in this tenant")
            record = await connection.fetchrow(
                """
                INSERT INTO wbs_user_groups (tenant_id, name, description)
                VALUES ($1, $2, $3)
                RETURNING id, tenant_id, name, description, status, metadata, created_at, updated_at,
                          0::integer AS user_count
                """,
                tid,
                name,
                (payload.description or "").strip(),
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="user_group.created",
                entity_type="user_group",
                entity_id=record["id"],
                summary=f"User group created: {name}",
                metadata={"name": name},
            )
    return user_group_response(record)


@app.patch("/api/user-groups/{group_id}")
async def update_user_group(group_id: str, payload: UserGroupUpdate, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    parsed_id = safe_uuid(group_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid user group id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            existing = await connection.fetchrow(
                """
                SELECT id, tenant_id, name, description, status, metadata, created_at, updated_at
                FROM wbs_user_groups
                WHERE id = $1 AND tenant_id = $2
                FOR UPDATE
                """,
                parsed_id,
                tid,
            )
            if not existing:
                raise HTTPException(status_code=404, detail="User group not found")
            existing_user_count = await connection.fetchval(
                "SELECT count(*) FROM wbs_users WHERE tenant_id = $1 AND group_id = $2",
                tid,
                parsed_id,
            )
            if payload.status == "Suspended" and int(existing_user_count or 0) > 0:
                raise HTTPException(status_code=409, detail="Cannot suspend a user group with assigned users")
            next_name = payload.name.strip() if payload.name is not None else None
            if next_name:
                duplicate = await connection.fetchval(
                    """
                    SELECT EXISTS (
                      SELECT 1
                      FROM wbs_user_groups
                      WHERE tenant_id = $1
                        AND lower(name) = lower($2)
                        AND id <> $3
                    )
                    """,
                    tid,
                    next_name,
                    parsed_id,
                )
                if duplicate:
                    raise HTTPException(status_code=409, detail="User group already exists in this tenant")
            record = await connection.fetchrow(
                """
                UPDATE wbs_user_groups
                SET name = COALESCE($3, name),
                    description = COALESCE($4, description),
                    status = COALESCE($5, status),
                    updated_at = now()
                WHERE id = $1 AND tenant_id = $2
                RETURNING id, tenant_id, name, description, status, metadata, created_at, updated_at,
                          0::integer AS user_count
                """,
                parsed_id,
                tid,
                next_name,
                payload.description.strip() if payload.description is not None else None,
                payload.status,
            )
            user_count = await connection.fetchval(
                "SELECT count(*) FROM wbs_users WHERE tenant_id = $1 AND group_id = $2",
                tid,
                parsed_id,
            )
            response = normalize_record(record)
            response["user_count"] = int(user_count or 0)
            await insert_audit_event(
                connection,
                request=request,
                event_type="user_group.updated",
                entity_type="user_group",
                entity_id=parsed_id,
                summary=f"User group updated: {response['name']}",
                metadata={"status": response["status"]},
            )
    return user_group_response(response)


@app.post("/api/users", status_code=201)
async def create_user(payload: UserCreate, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    normalized_email = payload.email.strip().lower()
    role = validate_user_role(payload.role)
    status = validate_user_status(payload.status)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            group = await resolve_user_group(connection, get_tenant_id(request), payload.group_id)
            email_exists = await connection.fetchval(
                "SELECT EXISTS (SELECT 1 FROM wbs_users WHERE email = $1)",
                normalized_email,
            )
            if email_exists:
                raise HTTPException(status_code=409, detail="User email already exists")
            record = await connection.fetchrow(
                """
                INSERT INTO wbs_users
                  (email, display_name, role, password_hash, status, must_change_password, tenant_id, group_id)
                VALUES
                  ($1, $2, $3, crypt($4, gen_salt('bf')), $5, $6, $7, $8)
                RETURNING id, email, display_name, role, status, failed_login_count,
                          tenant_id, group_id, locked_until, must_change_password, last_login_at,
                          password_changed_at, created_at, updated_at,
                          0::integer AS active_sessions
                """,
                normalized_email,
                payload.display_name.strip(),
                role,
                payload.password,
                status,
                payload.must_change_password,
                get_tenant_id(request),
                safe_uuid(group["id"]),
            )
            response = normalize_record(record)
            response["group_name"] = group["name"]
            response["group_status"] = group["status"]
            await insert_audit_event(
                connection,
                request=request,
                event_type="user.created",
                entity_type="user",
                entity_id=record["id"],
                summary=f"User created: {normalized_email}",
                metadata={
                    "role": role,
                    "status": status,
                    "must_change_password": payload.must_change_password,
                    "group_id": group["id"],
                    "group_name": group["name"],
                },
            )

    return managed_user_response(response)


@app.patch("/api/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdate, request: Request) -> dict[str, Any]:
    current = require_admin_role(request)
    parsed_id = safe_uuid(user_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid user id")

    role = validate_user_role(payload.role) if payload.role is not None else None
    status = validate_user_status(payload.status) if payload.status is not None else None
    display_name = payload.display_name.strip() if payload.display_name is not None else None
    fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))
    group_change_requested = "group_id" in fields_set

    if parsed_id == safe_uuid(current["id"]):
        if role and role != "admin":
            raise HTTPException(status_code=400, detail="Cannot remove your own admin role")
        if status and status != "Active":
            raise HTTPException(status_code=400, detail="Cannot suspend your own account")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            existing = await connection.fetchrow(
                "SELECT id, email, display_name, role, status, group_id FROM wbs_users WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
                parsed_id,
                get_tenant_id(request),
            )
            if not existing:
                raise HTTPException(status_code=404, detail="User not found")
            group = None
            if group_change_requested:
                if payload.group_id is None:
                    raise HTTPException(status_code=400, detail="User group is required")
                group = await resolve_user_group(connection, get_tenant_id(request), payload.group_id)

            record = await connection.fetchrow(
                """
                UPDATE wbs_users
                SET display_name = COALESCE($2, display_name),
                    role = COALESCE($3, role),
                    status = COALESCE($4, status),
                    must_change_password = COALESCE($6, CASE WHEN $5::text IS NULL THEN must_change_password ELSE true END),
                    group_id = COALESCE($7, group_id),
                    password_hash = CASE
                      WHEN $5::text IS NULL THEN password_hash
                      ELSE crypt($5, gen_salt('bf'))
                    END,
                    password_changed_at = CASE
                      WHEN $5::text IS NULL THEN password_changed_at
                      ELSE now()
                    END,
                    updated_at = now()
                WHERE id = $1
                RETURNING id, email, display_name, role, status, failed_login_count,
                          tenant_id, group_id, locked_until, must_change_password, last_login_at,
                          password_changed_at, created_at, updated_at,
                          0::integer AS active_sessions
                """,
                parsed_id,
                display_name,
                role,
                status,
                payload.password,
                payload.must_change_password,
                safe_uuid(group["id"]) if group else None,
            )
            response = normalize_record(record)
            current_group = group or await resolve_user_group(
                connection,
                get_tenant_id(request),
                response.get("group_id"),
                require_active=False,
            )
            response["group_name"] = current_group["name"]
            response["group_status"] = current_group["status"]
            if status and status != "Active":
                await connection.execute("DELETE FROM wbs_user_sessions WHERE user_id = $1", parsed_id)
            await insert_audit_event(
                connection,
                request=request,
                event_type="user.updated",
                entity_type="user",
                entity_id=parsed_id,
                summary=f"User updated: {existing['email']}",
                metadata={
                    "display_name_changed": display_name is not None and display_name != existing["display_name"],
                    "role_changed": role is not None and role != existing["role"],
                    "status_changed": status is not None and status != existing["status"],
                    "password_changed": payload.password is not None,
                    "must_change_password": payload.must_change_password,
                    "group_changed": group_change_requested and safe_uuid(payload.group_id) != existing["group_id"],
                    "group_id": group["id"] if group else str(existing["group_id"]),
                    "group_name": group["name"] if group else current_group["name"],
                },
            )

    return managed_user_response(response)


@app.post("/api/users/{user_id}/sessions/revoke")
async def revoke_user_sessions(user_id: str, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    parsed_id = safe_uuid(user_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid user id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            user_exists = await connection.fetchval("SELECT EXISTS (SELECT 1 FROM wbs_users WHERE id = $1)", parsed_id)
            if not user_exists:
                raise HTTPException(status_code=404, detail="User not found")
            deleted = await connection.fetchval(
                """
                WITH deleted AS (
                  DELETE FROM wbs_user_sessions
                  WHERE user_id = $1
                  RETURNING token
                )
                SELECT count(*) FROM deleted
                """,
                parsed_id,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="user.sessions_revoked",
                entity_type="user",
                entity_id=parsed_id,
                summary="User sessions revoked",
                metadata={"revoked_sessions": int(deleted or 0)},
            )

    return {"status": "ok", "revoked_sessions": int(deleted or 0)}


@app.get("/api/audit-events")
async def list_audit_events(
    request: Request,
    limit: int = 50,
    event_type: str | None = None,
    actor_email: str | None = None,
    entity_type: str | None = None,
    search: str | None = None,
) -> list[dict[str, Any]]:
    require_roles(request, {"admin", "pmo"})
    bounded_limit = max(1, min(limit, 100))
    tid = get_tenant_id(request)

    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            f"""
            SELECT {AUDIT_SELECT}
            FROM wbs_audit_events
            WHERE tenant_id = $1
              AND ($2::text IS NULL OR event_type = $2)
              AND ($3::text IS NULL OR actor_email = $3)
              AND ($4::text IS NULL OR entity_type = $4)
              AND (
                $5::text IS NULL
                OR summary ILIKE '%' || $5 || '%'
                OR event_type ILIKE '%' || $5 || '%'
                OR entity_id ILIKE '%' || $5 || '%'
              )
            ORDER BY created_at DESC
            LIMIT $6
            """,
            tid,
            event_type,
            actor_email.strip().lower() if actor_email else None,
            entity_type,
            search.strip() if search else None,
            bounded_limit,
        )

    return [audit_response(record) for record in records]


@app.get("/api/settings")
async def list_settings(request: Request) -> dict[str, Any]:
    require_roles(request, {"admin", "pmo"})
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            f"""
            SELECT {SETTING_SELECT}
            FROM wbs_system_settings
            ORDER BY category, key
            """
        )

    settings = [setting_response(record) for record in records]
    pm_engine_setting = next((setting for setting in settings if setting["key"] == "pm_engine"), None)
    return {
        "settings": settings,
        "pm_engine": pm_engine_status(pm_engine_setting.get("value") if pm_engine_setting else None),
    }


@app.get("/api/project-operation-policy")
async def get_project_operation_policy(request: Request) -> dict[str, Any]:
    require_roles(request, {"admin", "pmo"})
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        policy = await fetch_project_operation_policy(connection, tid)
    return policy


@app.put("/api/project-operation-policy")
async def update_project_operation_policy(
    payload: ProjectOperationPolicyUpdate,
    request: Request,
) -> dict[str, Any]:
    current = require_admin_role(request)
    tid = get_tenant_id(request)
    fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            current_policy = await fetch_project_operation_policy(connection, tid)
            next_policy = dict(current_policy)
            for field in fields_set:
                if field in {"tenant_id", "created_at", "updated_at", "updated_by"}:
                    continue
                next_policy[field] = getattr(payload, field)
            next_policy = normalize_project_operation_policy_value(next_policy, tid)
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_project_operation_policies
                  (tenant_id, default_delivery_mode, story_point_mode, fibonacci_points,
                   sprint_length_policy, dod_management, default_dod_items,
                   openproject_sprint_version_sync, metadata, updated_by)
                VALUES
                  ($1, $2, $3, $4::jsonb, $5, $6, $7::jsonb, $8, $9::jsonb, $10)
                ON CONFLICT (tenant_id) DO UPDATE
                SET default_delivery_mode = EXCLUDED.default_delivery_mode,
                    story_point_mode = EXCLUDED.story_point_mode,
                    fibonacci_points = EXCLUDED.fibonacci_points,
                    sprint_length_policy = EXCLUDED.sprint_length_policy,
                    dod_management = EXCLUDED.dod_management,
                    default_dod_items = EXCLUDED.default_dod_items,
                    openproject_sprint_version_sync = EXCLUDED.openproject_sprint_version_sync,
                    metadata = EXCLUDED.metadata,
                    updated_by = EXCLUDED.updated_by,
                    updated_at = now()
                RETURNING {PROJECT_OPERATION_POLICY_SELECT}
                """,
                tid,
                next_policy["default_delivery_mode"],
                next_policy["story_point_mode"],
                next_policy["fibonacci_points"],
                next_policy["sprint_length_policy"],
                next_policy["dod_management"],
                next_policy["default_dod_items"],
                next_policy["openproject_sprint_version_sync"],
                next_policy["metadata"],
                safe_uuid(current["id"]),
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project_policy.updated",
                entity_type="project_operation_policy",
                entity_id=tid,
                summary=f"프로젝트 운영 정책 저장: {tid}",
                metadata={"tenant_id": tid, "policy": project_operation_policy_response(record, tid)},
            )

    return project_operation_policy_response(record, tid)


@app.put("/api/settings/{setting_key}")
async def update_setting(setting_key: str, payload: SettingUpdate, request: Request) -> dict[str, Any]:
    current = require_admin_role(request)
    key = setting_key.strip()
    if not key:
        raise HTTPException(status_code=400, detail="Setting key is required")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            existing = await connection.fetchrow(
                f"SELECT {SETTING_SELECT} FROM wbs_system_settings WHERE key = $1 FOR UPDATE",
                key,
            )
            if not existing:
                raise HTTPException(status_code=404, detail="Setting not found")
            record = await connection.fetchrow(
                f"""
                UPDATE wbs_system_settings
                SET value = $2::jsonb,
                    updated_by = $3,
                    updated_at = now()
                WHERE key = $1
                RETURNING {SETTING_SELECT}
                """,
                key,
                payload.value,
                safe_uuid(current["id"]),
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="setting.updated",
                entity_type="setting",
                entity_id=key,
                summary=f"Setting updated: {key}",
                metadata={"category": existing["category"]},
            )

    setting = setting_response(record)
    return {
        "setting": setting,
        "pm_engine": pm_engine_status(setting["value"]) if key == "pm_engine" else None,
    }


@app.get("/api/report-schedules")
async def list_report_schedules(request: Request) -> dict[str, Any]:
    require_roles(request, {"admin", "pmo"})
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            f"""
            SELECT {REPORT_SCHEDULE_SELECT}
            FROM wbs_report_schedules
            ORDER BY key
            """
        )
    return {
        "schedules": [report_schedule_response(record) for record in records],
        "runtime": scheduler_runtime_state(request.app),
        "smtp": {
            "configured": bool(REPORT_SMTP_HOST),
            "host": REPORT_SMTP_HOST or None,
            "from": REPORT_SMTP_FROM,
        },
    }


@app.patch("/api/report-schedules/{schedule_key}")
async def update_report_schedule(
    schedule_key: str,
    payload: ReportScheduleUpdate,
    request: Request,
) -> dict[str, Any]:
    require_mutating_role(request)
    key = schedule_key.strip()
    if not key:
        raise HTTPException(status_code=400, detail="Schedule key is required")

    data = payload.model_dump(exclude_unset=True)
    recipients = normalize_email_recipients(data["recipients"]) if "recipients" in data else None
    metadata = data.get("metadata") if "metadata" in data else None

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            existing = await connection.fetchrow(
                f"""
                SELECT {REPORT_SCHEDULE_SELECT}
                FROM wbs_report_schedules
                WHERE key = $1
                FOR UPDATE
                """,
                key,
            )
            if not existing:
                raise HTTPException(status_code=404, detail="Report schedule not found")

            existing_metadata = normalize_metadata(existing["metadata"])
            next_metadata = {**existing_metadata, **(metadata or {})} if metadata is not None else existing_metadata
            record = await connection.fetchrow(
                f"""
                UPDATE wbs_report_schedules
                SET name = COALESCE($2, name),
                    enabled = COALESCE($3, enabled),
                    timezone = COALESCE($4, timezone),
                    day_of_week = COALESCE($5, day_of_week),
                    hour = COALESCE($6, hour),
                    minute = COALESCE($7, minute),
                    recipients = COALESCE($8::jsonb, recipients),
                    email_subject = COALESCE($9, email_subject),
                    email_body = COALESCE($10, email_body),
                    metadata = $11::jsonb,
                    updated_at = now()
                WHERE key = $1
                RETURNING {REPORT_SCHEDULE_SELECT}
                """,
                key,
                data.get("name"),
                data.get("enabled"),
                data.get("timezone"),
                data.get("day_of_week"),
                data.get("hour"),
                data.get("minute"),
                recipients,
                data.get("email_subject"),
                data.get("email_body"),
                next_metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="report.schedule_updated",
                entity_type="report_schedule",
                entity_id=key,
                summary=f"Report schedule updated: {key}",
                metadata={
                    "enabled": record["enabled"],
                    "day_of_week": record["day_of_week"],
                    "hour": record["hour"],
                    "minute": record["minute"],
                    "recipient_count": len(normalize_email_recipients(record["recipients"])),
                },
            )

    await refresh_report_scheduler(request.app)
    async with get_pool(request).acquire() as connection:
        refreshed = await fetch_report_schedule(connection, key)
    return {
        "schedule": report_schedule_response(refreshed),
        "runtime": scheduler_runtime_state(request.app),
    }


@app.post("/api/report-schedules/{schedule_key}/run", status_code=201)
async def run_report_schedule(
    schedule_key: str,
    request: Request,
    payload: ReportRunRequest = ReportRunRequest(),
) -> dict[str, Any]:
    require_mutating_role(request)
    async with get_pool(request).acquire() as connection:
        schedule = await fetch_report_schedule(connection, schedule_key.strip())
    if not schedule:
        raise HTTPException(status_code=404, detail="Report schedule not found")

    run = await execute_report_schedule(
        request.app,
        schedule,
        send_email=payload.send_email,
        triggered_by=payload.triggered_by,
        request=request,
    )
    await refresh_report_scheduler(request.app)
    return {
        "run": run,
        "runtime": scheduler_runtime_state(request.app),
    }


@app.get("/api/report-runs")
async def list_report_runs(request: Request, limit: int = 20) -> list[dict[str, Any]]:
    require_roles(request, {"admin", "pmo"})
    bounded_limit = min(max(limit, 1), 100)
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            f"""
            SELECT {REPORT_RUN_SELECT}
            FROM wbs_report_runs
            ORDER BY started_at DESC
            LIMIT $1
            """,
            bounded_limit,
        )
    return [report_run_response(record) for record in records]


@app.get("/api/report-runs/{run_id}/artifact")
async def download_report_run_artifact(run_id: str, request: Request) -> StreamingResponse:
    require_roles(request, {"admin", "pmo"})
    parsed_id = safe_uuid(run_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid report run id")

    async with get_pool(request).acquire() as connection:
        record = await connection.fetchrow(
            f"""
            SELECT {REPORT_RUN_SELECT}
            FROM wbs_report_runs
            WHERE id = $1
            """,
            parsed_id,
        )
    if not record or not record["artifact_path"]:
        raise HTTPException(status_code=404, detail="Report artifact not found")

    artifact_path = Path(record["artifact_path"]).resolve()
    output_root = REPORT_OUTPUT_DIR.resolve()
    if output_root not in artifact_path.parents and artifact_path != output_root:
        raise HTTPException(status_code=403, detail="Report artifact path is outside the report directory")
    if not artifact_path.exists():
        raise HTTPException(status_code=404, detail="Report artifact file is missing")

    filename = artifact_path.name
    return StreamingResponse(
        BytesIO(artifact_path.read_bytes()),
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/metrics", response_class=PlainTextResponse)
async def metrics(request: Request) -> PlainTextResponse:
    pool = get_pool(request)
    database_up = 1

    try:
        async with pool.acquire() as connection:
            await connection.fetchval("SELECT 1")
            project_count = await connection.fetchval("SELECT count(*) FROM wbs_projects")
            template_count = await connection.fetchval("SELECT count(*) FROM wbs_templates")
            approval_rows = await connection.fetch(
                """
                SELECT status, count(*) AS count
                FROM wbs_approval_requests
                GROUP BY status
                """
            )
            import_rows = await connection.fetch(
                """
                SELECT status, count(*) AS count
                FROM wbs_import_jobs
                GROUP BY status
                """
            )
            project_status_rows = await connection.fetch(
                """
                SELECT status, count(*) AS count
                FROM wbs_projects
                GROUP BY status
                """
            )
            sync_run_rows = await connection.fetch(
                """
                SELECT mode, status, count(*) AS count
                FROM wbs_sync_runs
                GROUP BY mode, status
                """
            )
            baseline_rows = await connection.fetch(
                """
                SELECT status, count(*) AS count
                FROM wbs_project_baselines
                GROUP BY status
                """
            )
            report_schedule_rows = await connection.fetch(
                """
                SELECT enabled::text AS status, count(*) AS count
                FROM wbs_report_schedules
                GROUP BY enabled
                """
            )
            report_run_rows = await connection.fetch(
                """
                SELECT status, count(*) AS count
                FROM wbs_report_runs
                GROUP BY status
                """
            )
    except Exception:
        database_up = 0
        project_count = 0
        template_count = 0
        approval_rows = []
        import_rows = []
        project_status_rows = []
        sync_run_rows = []
        baseline_rows = []
        report_schedule_rows = []
        report_run_rows = []

    lines = [
        "# HELP wbs_api_up WBS extension API availability.",
        "# TYPE wbs_api_up gauge",
        prometheus_metric("wbs_api_up", 1),
        "# HELP wbs_database_up PostgreSQL availability from the WBS API.",
        "# TYPE wbs_database_up gauge",
        prometheus_metric("wbs_database_up", database_up),
        "# HELP wbs_projects_total Number of WBS projects.",
        "# TYPE wbs_projects_total gauge",
        prometheus_metric("wbs_projects_total", project_count),
        "# HELP wbs_templates_total Number of WBS templates.",
        "# TYPE wbs_templates_total gauge",
        prometheus_metric("wbs_templates_total", template_count),
        "# HELP wbs_db_pool_connections Asyncpg pool connections by state.",
        "# TYPE wbs_db_pool_connections gauge",
        prometheus_metric("wbs_db_pool_connections", pool.get_size(), {"state": "total"}),
        prometheus_metric("wbs_db_pool_connections", pool.get_idle_size(), {"state": "idle"}),
        "# HELP wbs_project_status_total Number of WBS projects by status.",
        "# TYPE wbs_project_status_total gauge",
    ]

    lines.extend(
        prometheus_metric("wbs_project_status_total", row["count"], {"status": row["status"]})
        for row in project_status_rows
    )
    lines.extend(
        [
            "# HELP wbs_approval_requests_total Number of approval requests by status.",
            "# TYPE wbs_approval_requests_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric("wbs_approval_requests_total", row["count"], {"status": row["status"]})
        for row in approval_rows
    )
    lines.extend(
        [
            "# HELP wbs_import_jobs_total Number of Excel import jobs by status.",
            "# TYPE wbs_import_jobs_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric("wbs_import_jobs_total", row["count"], {"status": row["status"]})
        for row in import_rows
    )
    lines.extend(
        [
            "# HELP wbs_sync_runs_total Number of OpenProject sync runs by mode and status.",
            "# TYPE wbs_sync_runs_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric(
            "wbs_sync_runs_total",
            row["count"],
            {"mode": row["mode"], "status": row["status"]},
        )
        for row in sync_run_rows
    )
    lines.extend(
        [
            "# HELP wbs_project_baselines_total Number of WBS project baselines by status.",
            "# TYPE wbs_project_baselines_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric("wbs_project_baselines_total", row["count"], {"status": row["status"]})
        for row in baseline_rows
    )
    lines.extend(
        [
            "# HELP wbs_report_schedules_total Number of weekly report schedules by enabled state.",
            "# TYPE wbs_report_schedules_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric("wbs_report_schedules_total", row["count"], {"enabled": row["status"]})
        for row in report_schedule_rows
    )
    lines.extend(
        [
            "# HELP wbs_report_runs_total Number of weekly report runs by status.",
            "# TYPE wbs_report_runs_total gauge",
        ]
    )
    lines.extend(
        prometheus_metric("wbs_report_runs_total", row["count"], {"status": row["status"]})
        for row in report_run_rows
    )

    return PlainTextResponse(
        "\n".join(lines) + "\n",
        media_type="text/plain; version=0.0.4; charset=utf-8",
    )


@app.get("/api/templates")
async def list_templates(request: Request) -> list[dict[str, Any]]:
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            """
            SELECT t.key, t.name, t.project_type, t.description, t.phases,
                   t.tenant_id,
                   count(i.id)::integer AS item_count
            FROM wbs_templates t
            LEFT JOIN wbs_template_items i ON i.template_key = t.key AND i.tenant_id = t.tenant_id
            GROUP BY t.key, t.name, t.project_type, t.description, t.phases, t.tenant_id
            ORDER BY t.project_type, t.name
            """
        )
    return [normalize_record(record) for record in records]


@app.get("/api/projects")
async def list_projects(request: Request) -> list[dict[str, Any]]:
    tid = get_tenant_id(request)
    user = getattr(request.state, "user", None) or {}
    async with get_pool(request).acquire() as connection:
        if user.get("role") == "admin":
            records = await connection.fetch(
                """
                SELECT id, name, template_key, owner, status, start_date, delivery_mode,
                       openproject_project_id, metadata, tenant_id, created_at, updated_at
                FROM wbs_projects
                WHERE tenant_id = $1
                ORDER BY created_at DESC
                LIMIT 50
                """,
                tid,
            )
        else:
            records = await connection.fetch(
                """
                SELECT p.id, p.name, p.template_key, p.owner, p.status, p.start_date, p.delivery_mode,
                       p.openproject_project_id, p.metadata, p.tenant_id, p.created_at, p.updated_at
                FROM wbs_projects p
                JOIN wbs_project_members m
                  ON m.project_id = p.id AND m.tenant_id = p.tenant_id
                WHERE p.tenant_id = $1 AND m.user_id = $2
                ORDER BY p.created_at DESC
                LIMIT 50
                """,
                tid,
                safe_uuid(user.get("id")),
            )
    return [normalize_record(record) for record in records]


@app.post("/api/projects", status_code=201)
async def create_project(payload: ProjectCreate, request: Request) -> dict[str, Any]:
    require_mutating_role(request)
    start_date = payload.start_date or date.today()
    fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))
    metadata: dict[str, Any] = {
        "created_by": "wbs-portal",
        "sync_target": "openproject",
        "strategy": "community-edition-extension-layer",
    }
    if payload.end_date:
        metadata["end_date"] = payload.end_date.isoformat()
    if payload.description:
        metadata["description"] = payload.description
    if payload.client_name:
        metadata["client_name"] = payload.client_name
    if payload.budget:
        metadata["budget"] = payload.budget
    if payload.project_manager:
        metadata["project_manager"] = payload.project_manager

    async with get_pool(request).acquire() as connection:
        template_exists = await connection.fetchval(
            "SELECT EXISTS (SELECT 1 FROM wbs_templates WHERE key = $1)",
            payload.template_key,
        )
        if not template_exists:
            raise HTTPException(status_code=404, detail="Template not found")

        tid = get_tenant_id(request)
        policy = await fetch_project_operation_policy(connection, tid)
        delivery_mode = normalize_delivery_mode(
            payload.delivery_mode if "delivery_mode" in fields_set else policy["default_delivery_mode"]
        )
        metadata["operation_policy"] = {
            "default_delivery_mode": policy["default_delivery_mode"],
            "story_point_mode": policy["story_point_mode"],
            "sprint_length_policy": policy["sprint_length_policy"],
            "dod_management": policy["dod_management"],
            "openproject_sprint_version_sync": policy["openproject_sprint_version_sync"],
            "applied_at": utc_now_iso(),
        }
        record = await connection.fetchrow(
            """
            INSERT INTO wbs_projects
              (name, template_key, owner, status, start_date, delivery_mode, metadata, tenant_id)
            VALUES
              ($1, $2, $3, 'Draft', $4, $5, $6::jsonb, $7)
            RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                      openproject_project_id, metadata, tenant_id, created_at, updated_at
            """,
            payload.name,
            payload.template_key,
            payload.owner,
            start_date,
            delivery_mode,
            metadata,
            tid,
        )
        creator = getattr(request.state, "user", None) or {}
        creator_id = safe_uuid(creator.get("id"))
        if creator_id:
            await connection.execute(
                """
                INSERT INTO wbs_project_members (tenant_id, project_id, user_id, project_role, granted_by)
                VALUES ($1, $2, $3, 'admin', $3)
                ON CONFLICT (project_id, user_id) DO NOTHING
                """,
                tid, record["id"], creator_id,
            )

        await insert_audit_event(
            connection,
            request=request,
            event_type="project.created",
            entity_type="project",
            entity_id=record["id"],
            summary=f"Project created: {payload.name}",
            metadata={"template_key": payload.template_key, "owner": payload.owner},
        )

    return normalize_record(record)


@app.patch("/api/projects/{project_id}/status")
async def update_project_status(project_id: str, payload: ProjectStatusUpdate, request: Request) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    target_status = validate_project_status(payload.status)
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            ensure_project_mutate_role(project)
            ensure_project_transition(project["status"], target_status)
            record = await connection.fetchrow(
                """
                UPDATE wbs_projects
                SET status = $2,
                    updated_at = now()
                WHERE id = $1
                RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                          openproject_project_id, metadata, tenant_id, created_at, updated_at
                """,
                parsed_id,
                target_status,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project.status_changed",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project status changed: {project['status']} -> {target_status}",
                metadata={"comment": payload.comment},
            )

    return normalize_record(record)


@app.patch("/api/projects/{project_id}/delivery-mode")
async def update_project_delivery_mode(
    project_id: str,
    payload: ProjectDeliveryModeUpdate,
    request: Request,
) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    target_mode = normalize_delivery_mode(payload.delivery_mode)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            ensure_project_mutate_role(project)
            metadata = normalize_metadata(project.get("metadata"))
            metadata["delivery_mode_updated_at"] = utc_now_iso()
            metadata["delivery_mode_comment"] = payload.comment or ""
            record = await connection.fetchrow(
                """
                UPDATE wbs_projects
                SET delivery_mode = $2,
                    metadata = $3::jsonb,
                    updated_at = now()
                WHERE id = $1 AND tenant_id = $4
                RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                          openproject_project_id, metadata, tenant_id, created_at, updated_at
                """,
                parsed_id,
                target_mode,
                metadata,
                get_tenant_id(request),
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project.delivery_mode_changed",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project delivery mode changed: {project['name']} -> {target_mode}",
                metadata={"from": project.get("delivery_mode"), "to": target_mode, "comment": payload.comment},
            )

    return normalize_record(record)


PROJECT_MEMBER_SELECT = """
    m.id, m.project_id, m.user_id, m.project_role, m.granted_by, m.created_at, m.updated_at,
    u.email, u.display_name, u.role AS global_role, u.status AS user_status
"""


@app.get("/api/projects/{project_id}/members")
async def list_project_members(project_id: str, request: Request) -> list[dict[str, Any]]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        records = await connection.fetch(
            f"""
            SELECT {PROJECT_MEMBER_SELECT}
            FROM wbs_project_members m
            JOIN wbs_users u ON u.id = m.user_id
            WHERE m.project_id = $1
            ORDER BY u.display_name
            """,
            parsed_id,
        )
    return [normalize_record(record) for record in records]


@app.get("/api/projects/{project_id}/members/candidates")
async def list_project_member_candidates(project_id: str, request: Request) -> list[dict[str, Any]]:
    """이 프로젝트에 아직 멤버로 등록되지 않은, 같은 테넌트 소속 활성 사용자 목록."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        if project["_project_role"] != "admin":
            raise HTTPException(status_code=403, detail="Insufficient role")
        records = await connection.fetch(
            """
            SELECT u.id, u.email, u.display_name, u.role AS global_role
            FROM wbs_users u
            WHERE u.tenant_id = $1
              AND u.status = 'Active'
              AND NOT EXISTS (
                SELECT 1 FROM wbs_project_members m
                WHERE m.project_id = $2 AND m.user_id = u.id
              )
            ORDER BY u.display_name
            """,
            tid, parsed_id,
        )
    return [normalize_record(record) for record in records]


@app.post("/api/projects/{project_id}/members", status_code=201)
async def add_project_member(project_id: str, payload: ProjectMemberCreate, request: Request) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    role = payload.project_role.strip().lower()
    if role not in ALLOWED_USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid project role")

    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            if project["_project_role"] != "admin":
                raise HTTPException(status_code=403, detail="Insufficient role")

            target_user = await connection.fetchrow(
                "SELECT id, email, display_name FROM wbs_users WHERE id = $1 AND tenant_id = $2",
                payload.user_id, tid,
            )
            if not target_user:
                raise HTTPException(status_code=404, detail="User not found")

            actor = getattr(request.state, "user", None) or {}
            row = await connection.fetchrow(
                """
                INSERT INTO wbs_project_members (tenant_id, project_id, user_id, project_role, granted_by)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (project_id, user_id)
                DO UPDATE SET project_role = EXCLUDED.project_role, granted_by = EXCLUDED.granted_by, updated_at = now()
                RETURNING id
                """,
                tid, parsed_id, payload.user_id, role, safe_uuid(actor.get("id")),
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project_member.added",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project member added: {target_user['email']} ({role})",
                metadata={"user_id": str(payload.user_id), "email": target_user["email"], "project_role": role},
            )
            record = await connection.fetchrow(
                f"""
                SELECT {PROJECT_MEMBER_SELECT}
                FROM wbs_project_members m
                JOIN wbs_users u ON u.id = m.user_id
                WHERE m.id = $1
                """,
                row["id"],
            )

    return normalize_record(record)


@app.patch("/api/projects/{project_id}/members/{member_id}")
async def update_project_member(
    project_id: str, member_id: str, payload: ProjectMemberUpdate, request: Request
) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    parsed_member_id = safe_uuid(member_id)
    if not parsed_id or not parsed_member_id:
        raise HTTPException(status_code=400, detail="Invalid id")
    role = payload.project_role.strip().lower()
    if role not in ALLOWED_USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid project role")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            if project["_project_role"] != "admin":
                raise HTTPException(status_code=403, detail="Insufficient role")

            member = await connection.fetchrow(
                """
                SELECT m.id, m.user_id, m.project_role, u.email
                FROM wbs_project_members m
                JOIN wbs_users u ON u.id = m.user_id
                WHERE m.id = $1 AND m.project_id = $2
                FOR UPDATE
                """,
                parsed_member_id, parsed_id,
            )
            if not member:
                raise HTTPException(status_code=404, detail="Member not found")

            if member["project_role"] == "admin" and role != "admin":
                remaining_admins = await connection.fetchval(
                    """
                    SELECT count(*) FROM wbs_project_members
                    WHERE project_id = $1 AND project_role = 'admin' AND id != $2
                    """,
                    parsed_id, parsed_member_id,
                )
                if remaining_admins == 0:
                    raise HTTPException(status_code=400, detail="At least one project admin is required")

            await connection.execute(
                "UPDATE wbs_project_members SET project_role = $2, updated_at = now() WHERE id = $1",
                parsed_member_id, role,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project_member.role_changed",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project member role changed: {member['email']} {member['project_role']} -> {role}",
                metadata={"user_id": str(member["user_id"]), "email": member["email"], "from": member["project_role"], "to": role},
            )
            record = await connection.fetchrow(
                f"""
                SELECT {PROJECT_MEMBER_SELECT}
                FROM wbs_project_members m
                JOIN wbs_users u ON u.id = m.user_id
                WHERE m.id = $1
                """,
                parsed_member_id,
            )

    return normalize_record(record)


@app.delete("/api/projects/{project_id}/members/{member_id}", status_code=204)
async def remove_project_member(project_id: str, member_id: str, request: Request) -> None:
    parsed_id = safe_uuid(project_id)
    parsed_member_id = safe_uuid(member_id)
    if not parsed_id or not parsed_member_id:
        raise HTTPException(status_code=400, detail="Invalid id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            if project["_project_role"] != "admin":
                raise HTTPException(status_code=403, detail="Insufficient role")

            member = await connection.fetchrow(
                """
                SELECT m.id, m.user_id, m.project_role, u.email
                FROM wbs_project_members m
                JOIN wbs_users u ON u.id = m.user_id
                WHERE m.id = $1 AND m.project_id = $2
                FOR UPDATE
                """,
                parsed_member_id, parsed_id,
            )
            if not member:
                raise HTTPException(status_code=404, detail="Member not found")

            if member["project_role"] == "admin":
                remaining_admins = await connection.fetchval(
                    """
                    SELECT count(*) FROM wbs_project_members
                    WHERE project_id = $1 AND project_role = 'admin' AND id != $2
                    """,
                    parsed_id, parsed_member_id,
                )
                if remaining_admins == 0:
                    raise HTTPException(status_code=400, detail="At least one project admin is required")

            await connection.execute("DELETE FROM wbs_project_members WHERE id = $1", parsed_member_id)
            await insert_audit_event(
                connection,
                request=request,
                event_type="project_member.removed",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project member removed: {member['email']}",
                metadata={"user_id": str(member["user_id"]), "email": member["email"], "project_role": member["project_role"]},
            )


@app.get("/api/pm-engine")
async def pm_engine() -> dict[str, Any]:
    return pm_engine_status()


@app.get("/api/pm-engine/preflight")
async def pm_engine_preflight() -> dict[str, Any]:
    return await run_openproject_preflight()


@app.get("/api/projects/{project_id}/sync-plan")
async def project_sync_plan(project_id: str, request: Request) -> dict[str, Any]:
    try:
        parsed_id = UUID(project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid project id") from exc

    project, template, rows = await fetch_project_sync_context(request, parsed_id)
    async with get_pool(request).acquire() as connection:
        baseline = await fetch_latest_project_baseline(connection, parsed_id)
        policy = await fetch_project_operation_policy(connection, project["tenant_id"])
        sprints = await connection.fetch(
            f"""
            SELECT {AGILE_SPRINT_SELECT}
            FROM wbs_agile_sprints
            WHERE project_id = $1 AND tenant_id = $2
            ORDER BY start_date, created_at
            """,
            parsed_id,
            project["tenant_id"],
        )

    return build_openproject_sync_plan(project, template, rows, baseline, policy, sprints)


@app.get("/api/projects/{project_id}/baseline")
async def project_baseline(project_id: str, request: Request) -> dict[str, Any]:
    try:
        parsed_id = UUID(project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid project id") from exc

    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        baseline = await fetch_latest_project_baseline(connection, parsed_id)

    return {
        **baseline_summary(baseline),
        "snapshot_rows": baseline.get("snapshot_rows") if baseline else [],
    }


@app.get("/api/projects/{project_id}/sync-runs")
async def list_project_sync_runs(project_id: str, request: Request, limit: int = 10) -> list[dict[str, Any]]:
    try:
        parsed_id = UUID(project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid project id") from exc

    limit = max(1, min(limit, 50))
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)

        records = await connection.fetch(
            f"""
            SELECT {SYNC_RUN_SELECT}
            FROM wbs_sync_runs s
            JOIN wbs_projects p ON p.id = s.project_id
            WHERE s.project_id = $1
            ORDER BY s.started_at DESC
            LIMIT $2
            """,
            parsed_id,
            limit,
        )

    return [normalize_record(record) for record in records]


@app.get("/api/projects/{project_id}/sync-preflight")
async def project_sync_preflight(project_id: str, request: Request) -> dict[str, Any]:
    try:
        parsed_id = UUID(project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid project id") from exc

    project, template, rows = await fetch_project_sync_context(request, parsed_id)
    ensure_project_mutate_role(project)
    async with get_pool(request).acquire() as connection:
        baseline = await fetch_latest_project_baseline(connection, parsed_id)
        policy = await fetch_project_operation_policy(connection, project["tenant_id"])
        sprints = await connection.fetch(
            f"""
            SELECT {AGILE_SPRINT_SELECT}
            FROM wbs_agile_sprints
            WHERE project_id = $1 AND tenant_id = $2
            ORDER BY start_date, created_at
            """,
            parsed_id,
            project["tenant_id"],
        )
    plan = build_openproject_sync_plan(project, template, rows, baseline, policy, sprints)
    preflight = await run_openproject_preflight()

    return {
        "status": "Preflight",
        "preflight": preflight,
        **plan,
        "payload_sample": build_openproject_payload_sample(project, rows),
    }


@app.post("/api/projects/{project_id}/sync")
async def sync_project_to_engine(
    project_id: str,
    request: Request,
    payload: ProjectSyncRequest = ProjectSyncRequest(),
) -> dict[str, Any]:
    try:
        parsed_id = UUID(project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid project id") from exc

    project, template, rows = await fetch_project_sync_context(request, parsed_id)
    ensure_project_mutate_role(project)
    async with get_pool(request).acquire() as connection:
        baseline = await fetch_latest_project_baseline(connection, parsed_id)
        policy = await fetch_project_operation_policy(connection, project["tenant_id"])
        sprints = await connection.fetch(
            f"""
            SELECT {AGILE_SPRINT_SELECT}
            FROM wbs_agile_sprints
            WHERE project_id = $1 AND tenant_id = $2
            ORDER BY start_date, created_at
            """,
            parsed_id,
            project["tenant_id"],
        )
    plan = build_openproject_sync_plan(project, template, rows, baseline, policy, sprints)
    if payload.dry_run:
        audit_run = await record_project_sync_run(
            request,
            project_id=parsed_id,
            plan=plan,
            payload=payload,
            mode="dry_run",
            status="DryRun",
            metadata={
                "source": "api",
                "ready_for_actual_sync": OPENPROJECT_SYNC_ENABLED and bool(OPENPROJECT_API_TOKEN),
            },
        )
        return {"status": "DryRun", **plan, "audit": audit_run}

    ensure_project_status_allowed(project, {ACTUAL_SYNC_REQUIRED_STATUS}, "Actual sync")
    if not baseline_summary(baseline).get("locked"):
        error = HTTPException(status_code=409, detail="Actual sync requires a locked WBS baseline")
        await record_project_sync_run(
            request,
            project_id=parsed_id,
            plan=plan,
            payload=payload,
            mode="actual",
            status="Blocked",
            metadata={"source": "api", "blocked_reason": "baseline_unlocked"},
            error=sync_error_payload(error),
        )
        raise error

    if PM_ENGINE_ADAPTER == "mock":
        metadata = normalize_metadata(project.get("metadata"))
        engine_metadata = normalize_metadata(metadata.get("pm_engine"))
        mock_project_id = engine_metadata.get("project_id") or f"mock-{str(parsed_id)[:8]}"
        work_packages = {
            row["code"]: {
                "id": f"mock-wp-{index}",
                "href": f"/mock/work_packages/{index}",
                "subject": row["subject"],
                "synced_at": utc_now_iso(),
            }
            for index, row in enumerate(plan["rows"], start=1)
        }
        versions = {
            row["name"]: {
                "id": f"mock-version-{index}",
                "href": f"/mock/versions/{index}",
                "sprint_id": row["sprint_id"],
                "start_date": row.get("start_date"),
                "end_date": row.get("end_date"),
                "synced_at": utc_now_iso(),
            }
            for index, row in enumerate(plan.get("sprint_versions") or [], start=1)
            if row.get("name")
        }
        engine_metadata.update(
            {
                "adapter": "mock",
                "project_id": mock_project_id,
                "project_identifier": plan["openproject"]["project_identifier"],
                "work_packages": work_packages,
                "versions": versions,
                "last_sync_at": utc_now_iso(),
            }
        )
        metadata["pm_engine"] = engine_metadata
        async with get_pool(request).acquire() as connection:
            record = await connection.fetchrow(
                """
                UPDATE wbs_projects
                SET status = 'Synced',
                    openproject_project_id = $2,
                    metadata = $3::jsonb,
                    updated_at = now()
                WHERE id = $1
                RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                          openproject_project_id, metadata, tenant_id, created_at, updated_at
                """,
                parsed_id,
                mock_project_id,
                metadata,
            )
        audit_run = await record_project_sync_run(
            request,
            project_id=parsed_id,
            plan=plan,
            payload=payload,
            mode="actual",
            status="Synced",
            created_work_packages=len(work_packages),
            openproject_project_id=mock_project_id,
            metadata={
                "source": "mock_adapter",
                "known_work_packages": len(work_packages),
                "known_versions": len(versions),
            },
        )
        return {
            "status": "Synced",
            "engine": pm_engine_status({"adapter": "mock", "display_name": "Mock PM Engine"}),
            "project": normalize_record(record),
            "openproject": {
                "project_id": mock_project_id,
                "project_identifier": engine_metadata.get("project_identifier"),
                "project_href": None,
            },
            "summary": {
                "created_work_packages": len(work_packages),
                "created_versions": len(versions),
                "known_work_packages": len(work_packages),
                "known_versions": len(versions),
                "total_rows": len(rows),
                "payload_validation": False,
            },
            "created_work_packages": list(work_packages.values()),
            "created_versions": list(versions.values()),
            "audit": audit_run,
        }

    if not OPENPROJECT_SYNC_ENABLED:
        error = HTTPException(
            status_code=400,
            detail="OpenProject sync is disabled. Set OPENPROJECT_SYNC_ENABLED=true to execute.",
        )
        await record_project_sync_run(
            request,
            project_id=parsed_id,
            plan=plan,
            payload=payload,
            mode="actual",
            status="Blocked",
            metadata={"source": "api", "blocked_reason": "sync_disabled"},
            error=sync_error_payload(error),
        )
        raise error
    if not OPENPROJECT_API_TOKEN:
        error = HTTPException(status_code=400, detail="OPENPROJECT_API_TOKEN is required for OpenProject sync.")
        await record_project_sync_run(
            request,
            project_id=parsed_id,
            plan=plan,
            payload=payload,
            mode="actual",
            status="Blocked",
            metadata={"source": "api", "blocked_reason": "missing_token"},
            error=sync_error_payload(error),
        )
        raise error

    client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
    metadata = normalize_metadata(project.get("metadata"))
    engine_metadata = normalize_metadata(metadata.get("pm_engine"))
    work_packages = normalize_metadata(engine_metadata.get("work_packages"))
    versions = normalize_metadata(engine_metadata.get("versions"))
    openproject_project_id = project.get("openproject_project_id") or engine_metadata.get("project_id")
    project_identifier = engine_metadata.get("project_identifier") or plan["openproject"]["project_identifier"]

    if payload.force_project_create or not openproject_project_id:
        created_project = await client.create_project(project, project_identifier)
        openproject_project_id = str(created_project.get("id") or "").strip()
        if not openproject_project_id:
            raise HTTPException(status_code=502, detail="OpenProject project response did not include an id")
        engine_metadata["project_id"] = openproject_project_id
        engine_metadata["project_identifier"] = created_project.get("identifier") or project_identifier
        engine_metadata["project_href"] = normalize_metadata(created_project.get("_links")).get("self", {}).get("href")
        engine_metadata["project_created_at"] = utc_now_iso()

    created_versions: list[dict[str, Any]] = []
    if plan["operation_policy"].get("openproject_sprint_version_sync"):
        for sprint in plan.get("sprint_versions") or []:
            version_name = str(sprint.get("name") or "").strip()
            if not version_name or version_name in versions:
                continue
            version_payload = client.build_version_payload(
                openproject_project_id=openproject_project_id,
                sprint=sprint,
            )
            created_version = await client.create_version_from_payload(version_payload)
            version_id = str(created_version.get("id") or "").strip()
            version_href = normalize_metadata(created_version.get("_links")).get("self", {}).get("href")
            versions[version_name] = {
                "id": version_id,
                "href": version_href,
                "sprint_id": sprint.get("sprint_id"),
                "start_date": sprint.get("start_date"),
                "end_date": sprint.get("end_date"),
                "synced_at": utc_now_iso(),
            }
            created_versions.append(
                {
                    "name": version_name,
                    "id": version_id,
                    "href": version_href,
                    "sprint_id": sprint.get("sprint_id"),
                }
            )

    created_work_packages: list[dict[str, Any]] = []
    if payload.create_work_packages:
        href_by_code = {
            code: item.get("href")
            for code, item in work_packages.items()
            if isinstance(item, dict) and item.get("href")
        }
        for row in rows:
            code = row["code"]
            if code in work_packages:
                continue
            parent_href = href_by_code.get(row.get("parent_code"))
            work_package_payload = client.build_work_package_payload(
                openproject_project_id=openproject_project_id,
                row=row,
                parent_href=parent_href,
            )
            if payload.validate_payloads:
                await client.validate_work_package_payload(work_package_payload)
            created_work_package = await client.create_work_package_from_payload(work_package_payload)
            work_package_id = str(created_work_package.get("id") or "").strip()
            work_package_href = normalize_metadata(created_work_package.get("_links")).get("self", {}).get("href")
            work_packages[code] = {
                "id": work_package_id,
                "href": work_package_href,
                "subject": created_work_package.get("subject") or f"{code} {row['name']}",
                "synced_at": utc_now_iso(),
            }
            if work_package_href:
                href_by_code[code] = work_package_href
            created_work_packages.append(
                {
                    "code": code,
                    "id": work_package_id,
                    "href": work_package_href,
                    "subject": work_packages[code]["subject"],
                }
            )

    engine_metadata["adapter"] = "openproject"
    engine_metadata["base_url"] = OPENPROJECT_BASE_URL
    engine_metadata["work_packages"] = work_packages
    engine_metadata["versions"] = versions
    engine_metadata["last_sync_at"] = utc_now_iso()
    metadata["pm_engine"] = engine_metadata

    async with get_pool(request).acquire() as connection:
        record = await connection.fetchrow(
            """
            UPDATE wbs_projects
            SET status = 'Synced',
                openproject_project_id = $2,
                metadata = $3::jsonb,
                updated_at = now()
            WHERE id = $1
            RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                      openproject_project_id, metadata, tenant_id, created_at, updated_at
            """,
            parsed_id,
            openproject_project_id,
            metadata,
        )

    updated_project = normalize_record(record)
    audit_run = await record_project_sync_run(
        request,
        project_id=parsed_id,
        plan=plan,
        payload=payload,
        mode="actual",
        status="Synced",
        created_work_packages=len(created_work_packages),
        openproject_project_id=openproject_project_id,
        metadata={
            "source": "api",
            "known_work_packages": len(work_packages),
            "known_versions": len(versions),
            "pending_versions": plan["summary"].get("pending_versions", 0),
            "created_versions": len(created_versions),
            "project_identifier": engine_metadata.get("project_identifier"),
            "project_href": engine_metadata.get("project_href"),
        },
    )
    return {
        "status": "Synced",
        "engine": pm_engine_status(),
        "project": updated_project,
        "openproject": {
            "project_id": openproject_project_id,
            "project_identifier": engine_metadata.get("project_identifier"),
            "project_href": engine_metadata.get("project_href"),
        },
        "summary": {
            "created_work_packages": len(created_work_packages),
            "created_versions": len(created_versions),
            "known_work_packages": len(work_packages),
            "known_versions": len(versions),
            "pending_versions": plan["summary"].get("pending_versions", 0),
            "total_rows": len(rows),
            "payload_validation": payload.validate_payloads,
        },
        "created_work_packages": created_work_packages,
        "created_versions": created_versions,
        "audit": audit_run,
    }


@app.get("/api/approvals")
async def list_approvals(request: Request) -> list[dict[str, Any]]:
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            f"""
            SELECT {APPROVAL_SELECT}
            FROM wbs_approval_requests a
            JOIN wbs_projects p ON p.id = a.project_id
            WHERE a.tenant_id = $1
            ORDER BY
              CASE a.status
                WHEN 'Pending' THEN 0
                WHEN 'Approved' THEN 1
                WHEN 'Rejected' THEN 2
                ELSE 3
              END,
              a.created_at DESC
            LIMIT 50
            """,
            tid,
        )
    return [normalize_record(record) for record in records]


@app.post("/api/approvals", status_code=201)
async def create_approval(payload: ApprovalCreate, request: Request) -> dict[str, Any]:
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, payload.project_id, for_update=True)
            ensure_project_mutate_role(project)
            ensure_project_status_allowed(project, APPROVAL_ALLOWED_PROJECT_STATUSES, "Approval request")

            pending_exists = await connection.fetchval(
                """
                SELECT EXISTS (
                  SELECT 1
                  FROM wbs_approval_requests
                  WHERE project_id = $1 AND status = 'Pending' AND tenant_id = $2
                )
                """,
                payload.project_id,
                tid,
            )
            if pending_exists:
                raise HTTPException(status_code=409, detail="Project already has a pending approval")

            title = payload.title or f"{project['name']} WBS baseline approval"
            approval_status = "Approved" if payload.auto_approve_internal else "Pending"
            project_status = "Approved" if payload.auto_approve_internal else "Review"
            decision_comment = "Auto-approved internal PMO baseline" if payload.auto_approve_internal else None
            approval_metadata = {
                **payload.metadata,
                "approval_mode": "auto_internal" if payload.auto_approve_internal else "manual",
            }
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_approval_requests
                  (project_id, title, request_type, status, requester, reviewer,
                   due_date, decision_comment, metadata, decided_at, tenant_id)
                VALUES
                  ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb,
                   CASE WHEN $4 = 'Approved' THEN now() ELSE NULL END, $10)
                RETURNING id
                """,
                payload.project_id,
                title,
                payload.request_type,
                approval_status,
                payload.requester,
                payload.reviewer,
                payload.due_date,
                decision_comment,
                approval_metadata,
                tid,
            )
            await connection.execute(
                """
                UPDATE wbs_projects
                SET status = $2,
                    updated_at = now()
                WHERE id = $1
                """,
                payload.project_id,
                project_status,
            )
            approval = await fetch_approval(connection, record["id"], tid)
            if payload.auto_approve_internal:
                baseline = await create_project_baseline(
                    connection,
                    project_id=payload.project_id,
                    approval_id=record["id"],
                    template_key=project["template_key"],
                    actor=payload.reviewer or payload.requester,
                    metadata={
                        "approval_mode": "auto_internal",
                        "request_type": payload.request_type,
                    },
                )
                approval["baseline"] = baseline_summary(baseline)
            await insert_audit_event(
                connection,
                request=request,
                event_type="approval.created",
                entity_type="approval",
                entity_id=record["id"],
                summary=f"Approval {approval_status}: {title}",
                metadata={
                    "project_id": str(payload.project_id),
                    "approval_mode": approval_metadata["approval_mode"],
                    "status": approval_status,
                    "baseline_locked": bool(approval.get("baseline", {}).get("locked")),
                },
            )

    return approval


@app.post("/api/approvals/{approval_id}/approve")
async def approve_approval(approval_id: str, payload: ApprovalDecision, request: Request) -> dict[str, Any]:
    try:
        parsed_id = UUID(approval_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid approval id") from exc

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            approval = await fetch_approval_for_update(connection, parsed_id, get_tenant_id(request))
            if not approval:
                raise HTTPException(status_code=404, detail="Approval request not found")
            ensure_project_mutate_role(await fetch_tenant_project(connection, request, approval["project_id"]))
            if approval["status"] != "Pending":
                raise HTTPException(status_code=409, detail="Approval request is already decided")

            await connection.execute(
                """
                UPDATE wbs_approval_requests
                SET status = 'Approved',
                    reviewer = $2,
                    decision_comment = $3,
                    decided_at = now(),
                    updated_at = now()
                WHERE id = $1
                """,
                parsed_id,
                payload.reviewer,
                payload.comment,
            )
            await connection.execute(
                """
                UPDATE wbs_projects
                SET status = 'Approved',
                    updated_at = now()
                WHERE id = $1
                """,
                approval["project_id"],
            )
            updated = await fetch_approval(connection, parsed_id, get_tenant_id(request))
            baseline = await create_project_baseline(
                connection,
                project_id=approval["project_id"],
                approval_id=parsed_id,
                template_key=approval["template_key"],
                actor=payload.reviewer,
                metadata={
                    "approval_mode": "manual",
                    "request_type": approval["request_type"],
                },
            )
            updated["baseline"] = baseline_summary(baseline)
            await insert_audit_event(
                connection,
                request=request,
                event_type="approval.approved",
                entity_type="approval",
                entity_id=parsed_id,
                summary=f"Approval approved: {approval['title']}",
                metadata={
                    "project_id": str(approval["project_id"]),
                    "baseline_locked": bool(updated["baseline"].get("locked")),
                },
            )

    # ── 승인 후 OpenProject 자동 동기화 시도 ────────────────────
    # OP가 연결된 경우에만 실행 (실패해도 승인 결과에 영향 없음)
    auto_sync_result: dict[str, Any] = {"triggered": False}
    if OPENPROJECT_SYNC_ENABLED and OPENPROJECT_API_TOKEN and PM_ENGINE_ADAPTER != "mock":
        try:
            project_id_for_sync = approval["project_id"]
            async with get_pool(request).acquire() as connection:
                project = await fetch_tenant_project(connection, request, project_id_for_sync)
            if project:
                metadata   = normalize_metadata(project.get("metadata"))
                engine_meta = normalize_metadata(metadata.get("pm_engine"))
                rows_data  = await request.app.state.pool.fetch(
                    "SELECT * FROM wbs_project_wbs_items WHERE project_id = $1 ORDER BY sort_order",
                    project_id_for_sync,
                )
                if rows_data:
                    client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
                    op_project_id  = project.get("openproject_project_id") or engine_meta.get("project_id")
                    op_identifier  = engine_meta.get("project_identifier") or normalize_openproject_identifier(project["name"], "wbs")
                    work_packages  = normalize_metadata(engine_meta.get("work_packages"))

                    if not op_project_id:
                        created_proj  = await client.create_project(project, op_identifier)
                        op_project_id = str(created_proj.get("id") or "").strip()

                    created_count = 0
                    for row in [normalize_record(r) for r in rows_data]:
                        code = row.get("code")
                        if not code or code in work_packages:
                            continue
                        wp_payload = client.build_work_package_payload(
                            openproject_project_id=op_project_id,
                            row=row,
                        )
                        created_wp = await client.create_work_package_from_payload(wp_payload)
                        wp_id   = str(created_wp.get("id") or "").strip()
                        wp_href = normalize_metadata(created_wp.get("_links")).get("self", {}).get("href")
                        work_packages[code] = {"id": wp_id, "href": wp_href, "synced_at": utc_now_iso()}
                        created_count += 1

                    engine_meta.update({
                        "adapter": PM_ENGINE_ADAPTER,
                        "project_id": op_project_id,
                        "project_identifier": op_identifier,
                        "work_packages": work_packages,
                        "last_sync_at": utc_now_iso(),
                        "auto_synced_on_approval": True,
                    })
                    metadata["pm_engine"] = engine_meta

                    async with get_pool(request).acquire() as connection:
                        await connection.execute(
                            """
                            UPDATE wbs_projects
                            SET status = 'Synced',
                                openproject_project_id = $2,
                                metadata = $3::jsonb,
                                updated_at = now()
                            WHERE id = $1
                            """,
                            project_id_for_sync,
                            op_project_id,
                            metadata,
                        )
                    auto_sync_result = {
                        "triggered":      True,
                        "created":        created_count,
                        "op_project_id":  op_project_id,
                        "status":         "Synced",
                    }
        except Exception as exc:
            auto_sync_result = {"triggered": True, "error": str(exc)[:200]}

    updated["auto_sync"] = auto_sync_result

    # P2-04: 승인 알림 — 프로젝트 담당자에게 발송
    try:
        async with get_pool(request).acquire() as connection:
            proj_row = await connection.fetchrow(
                "SELECT p.owner, u.id AS uid, u.email FROM wbs_projects p LEFT JOIN wbs_users u ON u.display_name = p.owner WHERE p.id = $1",
                approval["project_id"],
            )
            if proj_row and proj_row["uid"]:
                await send_notification(
                    connection,
                    user_id=UUID(str(proj_row["uid"])),
                    event_type="approval.approved",
                    title=f"승인 완료: {approval['title']}",
                    body=f"승인자: {payload.reviewer}\n{payload.comment or ''}".strip(),
                    entity_type="approval", entity_id=str(parsed_id),
                    email_to=proj_row["email"],
                )
    except Exception:
        pass

    return updated


@app.post("/api/approvals/{approval_id}/reject")
async def reject_approval(approval_id: str, payload: ApprovalDecision, request: Request) -> dict[str, Any]:
    try:
        parsed_id = UUID(approval_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid approval id") from exc

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            approval = await fetch_approval_for_update(connection, parsed_id, get_tenant_id(request))
            if not approval:
                raise HTTPException(status_code=404, detail="Approval request not found")
            ensure_project_mutate_role(await fetch_tenant_project(connection, request, approval["project_id"]))
            if approval["status"] != "Pending":
                raise HTTPException(status_code=409, detail="Approval request is already decided")

            await connection.execute(
                """
                UPDATE wbs_approval_requests
                SET status = 'Rejected',
                    reviewer = $2,
                    decision_comment = $3,
                    decided_at = now(),
                    updated_at = now()
                WHERE id = $1
                """,
                parsed_id,
                payload.reviewer,
                payload.comment,
            )
            await connection.execute(
                """
                UPDATE wbs_projects
                SET status = 'Rejected',
                    updated_at = now()
                WHERE id = $1
                """,
                approval["project_id"],
            )
            updated = await fetch_approval(connection, parsed_id, get_tenant_id(request))
            await insert_audit_event(
                connection,
                request=request,
                event_type="approval.rejected",
                entity_type="approval",
                entity_id=parsed_id,
                summary=f"Approval rejected: {approval['title']}",
                metadata={"project_id": str(approval["project_id"])},
            )

    return updated


@app.get("/api/dashboard")
async def dashboard(request: Request) -> dict[str, Any]:
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project_count = await connection.fetchval("SELECT count(*) FROM wbs_projects WHERE tenant_id = $1", tid)
        template_count = await connection.fetchval("SELECT count(*) FROM wbs_templates")
        pending_approval_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_approval_requests WHERE status = 'Pending' AND tenant_id = $1", tid
        )
        preview_import_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_import_jobs WHERE status = 'Preview'"
        )
        status_rows = await connection.fetch(
            """
            SELECT status, count(*) AS count
            FROM wbs_projects
            WHERE tenant_id = $1
            GROUP BY status
            ORDER BY status
            """,
            tid,
        )
        latest_projects = await connection.fetch(
            """
            SELECT name, owner, status, start_date, template_key
            FROM wbs_projects
            WHERE tenant_id = $1
            ORDER BY created_at DESC
            LIMIT 5
            """,
            tid,
        )
        latest_approvals = await connection.fetch(
            f"""
            SELECT {APPROVAL_SELECT}
            FROM wbs_approval_requests a
            JOIN wbs_projects p ON p.id = a.project_id
            WHERE a.tenant_id = $1
            ORDER BY a.created_at DESC
            LIMIT 5
            """,
            tid,
        )

        # P2-01: SPI/CPI — weight × progress 기반 집계
        wbs_agg_rows = await connection.fetch(
            """
            SELECT
                p.id AS project_id,
                p.name AS project_name,
                p.status AS project_status,
                p.start_date,
                p.metadata->>'end_date' AS end_date,
                COALESCE(SUM(
                    CASE WHEN i.item_type NOT IN ('프로젝트') THEN
                        COALESCE((i.metadata->>'weight')::numeric, i.weight, 0)
                    ELSE 0 END
                ), 0) AS total_weight,
                COALESCE(SUM(
                    CASE WHEN i.item_type NOT IN ('프로젝트') THEN
                        COALESCE((i.metadata->>'weight')::numeric, i.weight, 0)
                        * COALESCE((i.metadata->>'progress')::numeric, 0) / 100.0
                    ELSE 0 END
                ), 0) AS earned_weight
            FROM wbs_projects p
            LEFT JOIN wbs_project_wbs_items i ON i.project_id = p.id
            WHERE p.status NOT IN ('Closed') AND p.tenant_id = $1
            GROUP BY p.id, p.name, p.status, p.start_date, p.metadata
            """,
            tid,
        )

        # P2-01: 리스크·이슈 카운트 (wbs_risks / wbs_issues 테이블 존재 시)
        risk_count = 0
        issue_count = 0
        try:
            risk_count = await connection.fetchval(
                "SELECT count(*) FROM wbs_risks WHERE status != 'Closed' AND tenant_id = $1", tid
            ) or 0
            issue_count = await connection.fetchval(
                "SELECT count(*) FROM wbs_issues WHERE status != 'Closed' AND tenant_id = $1", tid
            ) or 0
        except Exception:
            pass

        # P2-01: 상태별 히트맵 — 단계(Phase)별 평균 진행률
        heatmap_rows = await connection.fetch(
            """
            SELECT
                p.name AS project_name,
                i.name AS phase_name,
                COALESCE(
                    (SELECT AVG(COALESCE((c.metadata->>'progress')::numeric, 0))
                     FROM wbs_project_wbs_items c
                     WHERE c.project_id = p.id AND c.parent_code = i.code
                    ), 0
                ) AS avg_progress
            FROM wbs_projects p
            JOIN wbs_project_wbs_items i ON i.project_id = p.id AND i.item_type = '단계'
            WHERE p.status NOT IN ('Closed') AND p.tenant_id = $1
            ORDER BY p.name, i.sort_order
            LIMIT 50
            """,
            tid,
        )

    # SPI/CPI 계산 (portfolio 수준)
    total_earned = sum(float(r["earned_weight"]) for r in wbs_agg_rows)
    total_planned = sum(float(r["total_weight"]) for r in wbs_agg_rows)
    today = date.today()

    project_kpis = []
    for r in wbs_agg_rows:
        tw = float(r["total_weight"]) or 0
        ew = float(r["earned_weight"]) or 0
        progress_pct = round(ew / tw * 100, 1) if tw else 0
        # 단순 SPI: earned / planned (일정 정보 없으면 progress 기반)
        spi = round(ew / tw, 3) if tw else None
        project_kpis.append({
            "project_id": str(r["project_id"]),
            "project_name": r["project_name"],
            "project_status": r["project_status"],
            "total_weight": tw,
            "earned_weight": round(ew, 2),
            "progress_pct": progress_pct,
            "spi": spi,
        })

    portfolio_spi = round(total_earned / total_planned, 3) if total_planned else None

    return {
        "metrics": {
            "projects": project_count,
            "templates": template_count,
            "pending_approvals": pending_approval_count,
            "preview_imports": preview_import_count,
            "risk_count": int(risk_count),
            "issue_count": int(issue_count),
            "portfolio_spi": portfolio_spi,
            "openproject_sync": "ready",
            "database": "PostgreSQL 17",
        },
        "status_distribution": [normalize_record(row) for row in status_rows],
        "latest_projects": [normalize_record(row) for row in latest_projects],
        "latest_approvals": [normalize_record(row) for row in latest_approvals],
        "project_kpis": project_kpis,
        "status_heatmap": [
            {
                "project_name": r["project_name"],
                "phase_name": r["phase_name"],
                "avg_progress": round(float(r["avg_progress"]), 1),
            }
            for r in heatmap_rows
        ],
        "risk_hotspots": [
            {"name": "Pending PMO approvals", "level": "attention" if pending_approval_count else "stable"},
            {"name": "Excel preview queue", "level": "watch" if preview_import_count else "stable"},
            {"name": "Open risks", "level": "attention" if risk_count > 0 else "stable", "count": int(risk_count)},
            {"name": "Open issues", "level": "watch" if issue_count > 0 else "stable", "count": int(issue_count)},
        ],
    }


@app.get("/api/operations/health")
async def operations_health(request: Request) -> dict[str, Any]:
    require_roles(request, {"admin", "pmo"})

    async with get_pool(request).acquire() as connection:
        database_version = await connection.fetchval("SHOW server_version")
        project_count = await connection.fetchval("SELECT count(*) FROM wbs_projects")
        template_count = await connection.fetchval("SELECT count(*) FROM wbs_templates")
        template_item_count = await connection.fetchval("SELECT count(*) FROM wbs_template_items")
        pending_approval_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_approval_requests WHERE status = 'Pending'"
        )
        preview_import_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_import_jobs WHERE status = 'Preview'"
        )
        sync_run_count = await connection.fetchval("SELECT count(*) FROM wbs_sync_runs")
        baseline_count = await connection.fetchval("SELECT count(*) FROM wbs_project_baselines")
        user_count = await connection.fetchval("SELECT count(*) FROM wbs_users")
        locked_user_count = await connection.fetchval("SELECT count(*) FROM wbs_users WHERE locked_until > now()")
        audit_count = await connection.fetchval("SELECT count(*) FROM wbs_audit_events")
        setting_count = await connection.fetchval("SELECT count(*) FROM wbs_system_settings")
        project_wbs_count = await connection.fetchval("SELECT count(*) FROM wbs_project_wbs_items")
        template_version_count = await connection.fetchval("SELECT count(*) FROM wbs_template_versions")
        report_schedule_count = await connection.fetchval("SELECT count(*) FROM wbs_report_schedules")
        report_enabled_count = await connection.fetchval("SELECT count(*) FROM wbs_report_schedules WHERE enabled = true")
        report_run_count = await connection.fetchval("SELECT count(*) FROM wbs_report_runs")
        report_failed_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_report_runs WHERE status = 'Failed' AND started_at >= now() - interval '7 days'"
        )
        existing_tables = await connection.fetch(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'public'
              AND table_name = ANY($1::text[])
            """,
            [
                "wbs_templates",
                "wbs_template_items",
                "wbs_projects",
                "wbs_import_jobs",
                "wbs_approval_requests",
                "wbs_sync_runs",
                "wbs_project_baselines",
                "wbs_users",
                "wbs_user_sessions",
                "wbs_system_settings",
                "wbs_audit_events",
                "wbs_template_versions",
                "wbs_project_wbs_items",
                "wbs_report_schedules",
                "wbs_report_runs",
            ],
        )

    table_names = {row["table_name"] for row in existing_tables}
    required_tables = {
        "wbs_templates",
        "wbs_template_items",
        "wbs_projects",
        "wbs_import_jobs",
        "wbs_approval_requests",
        "wbs_sync_runs",
        "wbs_project_baselines",
        "wbs_users",
        "wbs_user_sessions",
        "wbs_system_settings",
        "wbs_audit_events",
        "wbs_template_versions",
        "wbs_project_wbs_items",
        "wbs_report_schedules",
        "wbs_report_runs",
    }
    missing_tables = sorted(required_tables - table_names)
    preflight = await run_openproject_preflight()
    openproject_status = "pass" if preflight["state"] == "ready" else "warn"
    if preflight["state"] in {"offline", "auth_failed", "blocked"}:
        openproject_status = "fail"

    checks = [
        operation_check(
            "postgresql",
            "PostgreSQL",
            "pass",
            f"PostgreSQL {database_version} is reachable",
            {"version": database_version},
        ),
        operation_check(
            "schema",
            "Schema migration",
            "fail" if missing_tables else "pass",
            "Required WBS tables are present" if not missing_tables else "Missing required WBS tables",
            {"missing_tables": missing_tables},
        ),
        operation_check(
            "template_baseline",
            "WBS template baseline",
            "pass" if template_count and template_item_count else "fail",
            f"{template_count} templates, {template_item_count} template items",
            {"templates": template_count, "template_items": template_item_count},
        ),
        operation_check(
            "portfolio_seed",
            "Project portfolio",
            "pass" if project_count else "warn",
            f"{project_count} projects registered",
            {"projects": project_count},
        ),
        operation_check(
            "approval_queue",
            "PMO approval queue",
            "warn" if pending_approval_count else "pass",
            f"{pending_approval_count} pending approvals",
            {"pending_approvals": pending_approval_count},
        ),
        operation_check(
            "excel_preview_queue",
            "Excel preview queue",
            "warn" if preview_import_count else "pass",
            f"{preview_import_count} imports waiting for apply",
            {"preview_imports": preview_import_count},
        ),
        operation_check(
            "openproject_preflight",
            "OpenProject preflight",
            openproject_status,
            f"Engine state: {preflight['state']}",
            {"state": preflight["state"], "ready_for_actual_sync": preflight["ready_for_actual_sync"]},
        ),
        operation_check(
            "sync_audit",
            "Sync audit trail",
            "pass",
            f"{sync_run_count} sync runs recorded",
            {"sync_runs": sync_run_count},
        ),
        operation_check(
            "baseline_lock",
            "Baseline lock",
            "pass",
            f"{baseline_count} locked WBS baselines",
            {"baselines": baseline_count},
        ),
        operation_check(
            "access_control",
            "Access control",
            "pass" if user_count else "fail",
            f"{user_count} portal users configured",
            {"users": user_count},
        ),
        operation_check(
            "settings_registry",
            "Settings registry",
            "pass" if setting_count >= 3 else "warn",
            f"{setting_count} settings registered",
            {"settings": setting_count},
        ),
        operation_check(
            "audit_log",
            "Audit log",
            "pass",
            f"{audit_count} audit events recorded",
            {"audit_events": audit_count, "retention_days": AUDIT_RETENTION_DAYS},
        ),
        operation_check(
            "account_lockout",
            "Account lockout",
            "warn" if locked_user_count else "pass",
            f"{locked_user_count} users currently locked",
            {"locked_users": locked_user_count, "failure_limit": LOGIN_FAILURE_LIMIT},
        ),
        operation_check(
            "template_versions",
            "Template versions",
            "pass" if template_version_count else "warn",
            f"{template_version_count} template versions stored",
            {"template_versions": template_version_count},
        ),
        operation_check(
            "project_wbs_storage",
            "Project WBS storage",
            "pass",
            f"{project_wbs_count} project WBS rows stored",
            {"project_wbs_rows": project_wbs_count},
        ),
        operation_check(
            "weekly_report_scheduler",
            "Weekly report scheduler",
            "fail" if REPORT_SCHEDULER_ENABLED and AsyncIOScheduler is None else "warn" if report_failed_count else "pass",
            f"{report_enabled_count} enabled schedules, {report_run_count} report runs",
            {
                "schedules": report_schedule_count,
                "enabled_schedules": report_enabled_count,
                "report_runs": report_run_count,
                "failed_last_7_days": report_failed_count,
                **scheduler_runtime_state(request.app),
                "smtp_configured": bool(REPORT_SMTP_HOST),
            },
        ),
        operation_check(
            "cors_policy",
            "CORS policy",
            "warn" if ALLOW_FILE_ORIGIN else "pass",
            "file:// origin is allowed for local development" if ALLOW_FILE_ORIGIN else "file:// origin is disabled",
            {"allow_file_origin": ALLOW_FILE_ORIGIN, "portal_origin": PORTAL_ORIGIN},
        ),
        backup_health_check(),
        operation_check(
            "metrics",
            "Metrics endpoint",
            "pass",
            "Prometheus metrics are exposed at /metrics",
            {"path": "/metrics"},
        ),
    ]
    failures = len([check for check in checks if check["status"] == "fail"])
    warnings = len([check for check in checks if check["status"] == "warn"])
    status = "critical" if failures else "watch" if warnings else "stable"

    return {
        "status": status,
        "generated_at": utc_now_iso(),
        "summary": {
            "checks": len(checks),
            "failures": failures,
            "warnings": warnings,
            "passes": len([check for check in checks if check["status"] == "pass"]),
        },
        "checks": checks,
    }


@app.get("/api/imports")
async def list_import_jobs(
    request: Request,
    limit: int = 10,
    status: str | None = None,
) -> list[dict[str, Any]]:
    bounded_limit = max(1, min(limit, 50))

    async with get_pool(request).acquire() as connection:
        if status:
            records = await connection.fetch(
                f"""
                SELECT {IMPORT_JOB_RETURNING}
                FROM wbs_import_jobs
                WHERE status = $1
                ORDER BY created_at DESC
                LIMIT $2
                """,
                status,
                bounded_limit,
            )
        else:
            records = await connection.fetch(
                f"""
                SELECT {IMPORT_JOB_RETURNING}
                FROM wbs_import_jobs
                ORDER BY created_at DESC
                LIMIT $1
                """,
                bounded_limit,
            )

    return [import_job_response(record) for record in records]


@app.get("/api/imports/{job_id}")
async def get_import_job(job_id: str, request: Request) -> dict[str, Any]:
    try:
        import_job_id = UUID(job_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid import job id") from exc

    async with get_pool(request).acquire() as connection:
        record = await connection.fetchrow(
            f"""
            SELECT {IMPORT_JOB_RETURNING}
            FROM wbs_import_jobs
            WHERE id = $1
            """,
            import_job_id,
        )

    if not record:
        raise HTTPException(status_code=404, detail="Import job not found")
    return import_job_response(record, include_rows=True)


@app.get("/api/imports/{job_id}/errors.xlsx")
async def download_import_errors(job_id: str, request: Request) -> StreamingResponse:
    try:
        import_job_id = UUID(job_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid import job id") from exc

    async with get_pool(request).acquire() as connection:
        record = await connection.fetchrow(
            f"""
            SELECT {IMPORT_JOB_RETURNING}
            FROM wbs_import_jobs
            WHERE id = $1
            """,
            import_job_id,
        )

    if not record:
        raise HTTPException(status_code=404, detail="Import job not found")
    job = normalize_record(record)
    output = build_import_errors_workbook(job)
    filename = f"{job['source_file'].rsplit('.', 1)[0]}-issues.xlsx"
    return StreamingResponse(
        output,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/imports/validate")
async def validate_import(payload: WbsImportValidation, request: Request) -> dict[str, Any]:
    require_mutating_role(request)
    rows = assign_missing_wbs_codes([
        {"row_number": index, **row.model_dump()}
        for index, row in enumerate(payload.rows, start=1)
    ], "WBS")
    errors, warnings = validate_wbs_rows(rows)
    warnings = [*warnings, *auto_code_warnings(rows)]
    serialized_rows = [serialize_wbs_row(row) for row in rows]

    status = "Rejected" if errors else "Accepted"
    async with get_pool(request).acquire() as connection:
        record = await insert_import_job(
            connection,
            source_file=payload.source_file,
            status=status,
            total_rows=len(rows),
            accepted_rows=0 if errors else len(rows),
            rejected_rows=len(errors),
            errors=errors,
            warnings=warnings,
            preview_rows=serialized_rows,
        )

    response = normalize_record(record)
    response["warnings"] = warnings
    response["rows"] = serialized_rows
    return response


@app.get("/api/templates/{template_key}/items")
async def list_template_items(template_key: str, request: Request) -> dict[str, Any]:
    normalized_key = normalize_template_key(template_key)
    async with get_pool(request).acquire() as connection:
        template = await fetch_template(connection, normalized_key)
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        rows = await fetch_template_items(connection, normalized_key)

    return {
        "template": template,
        "rows": rows,
    }


@app.get("/api/templates/{template_key}/versions")
async def list_template_versions(template_key: str, request: Request, limit: int = 20) -> list[dict[str, Any]]:
    normalized_key = normalize_template_key(template_key)
    bounded_limit = max(1, min(limit, 100))
    async with get_pool(request).acquire() as connection:
        records = await connection.fetch(
            """
            SELECT id, template_key, version, template_name, project_type,
                   description, item_count, metadata, created_at
            FROM wbs_template_versions
            WHERE template_key = $1
            ORDER BY version DESC
            LIMIT $2
            """,
            normalized_key,
            bounded_limit,
        )
    return [normalize_record(record) for record in records]


@app.get("/api/templates/{template_key}/excel")
async def export_template_excel(template_key: str, request: Request) -> StreamingResponse:
    normalized_key = normalize_template_key(template_key)
    async with get_pool(request).acquire() as connection:
        template = await fetch_template(connection, normalized_key)
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        rows = await fetch_template_items(connection, normalized_key)

    if not rows:
        rows = [
            {
                "code": phase["code"],
                "parent_code": None,
                "name": phase["name"],
                "item_type": "단계",
                "owner": "PMO",
                "weight": phase.get("weight"),
                "start_date": None,
                "finish_date": None,
                "metadata": {},
            }
            for phase in template.get("phases", [])
        ]

    output = build_template_workbook(template, rows)
    filename = f"{normalized_key}-wbs-template.xlsx"
    return StreamingResponse(
        output,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/templates/import/preview", status_code=201)
async def preview_template_excel(
    request: Request,
    file: UploadFile = File(...),
    template_key: str = Form("uploaded-wbs"),
    template_name: str = Form("업로드 WBS 템플릿"),
    project_type: str = Form("Uploaded"),
    description: str = Form("Excel 업로드로 반영될 계층형 WBS 템플릿"),
) -> dict[str, Any]:
    require_mutating_role(request)
    normalized_key = normalize_template_key(template_key)
    contents = await file.read()
    parsed_rows = parse_wbs_workbook(contents)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            rows, errors, warnings, serialized_rows, diff_rows = await prepare_template_import(
                connection,
                template_key=normalized_key,
                parsed_rows=parsed_rows,
            )
            status = "Rejected" if errors else "Preview"
            record = await insert_import_job(
                connection,
                source_file=file.filename or "wbs-upload.xlsx",
                template_key=normalized_key,
                template_name=template_name,
                project_type=project_type,
                description=description,
                status=status,
                total_rows=len(rows),
                accepted_rows=0 if errors else len(rows),
                rejected_rows=len(errors),
                errors=errors,
                warnings=warnings,
                preview_rows=serialized_rows,
                diff_rows=diff_rows,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="import.previewed",
                entity_type="import_job",
                entity_id=record["id"],
                summary=f"Excel import preview: {file.filename or 'wbs-upload.xlsx'}",
                metadata={
                    "template_key": normalized_key,
                    "status": status,
                    "accepted_rows": 0 if errors else len(rows),
                    "rejected_rows": len(errors),
                    "diff_count": len(diff_rows),
                },
            )
            template = await fetch_template(connection, normalized_key)

    response = normalize_record(record)
    response["template"] = template
    response["warnings"] = warnings
    response["rows"] = serialized_rows[:50]
    response["diff_rows"] = diff_rows[:50]
    return response


@app.post("/api/templates/import", status_code=201)
async def import_template_excel(
    request: Request,
    file: UploadFile = File(...),
    template_key: str = Form("uploaded-wbs"),
    template_name: str = Form("업로드 WBS 템플릿"),
    project_type: str = Form("Uploaded"),
    description: str = Form("Excel 업로드로 반영된 계층형 WBS 템플릿"),
) -> dict[str, Any]:
    require_mutating_role(request)
    normalized_key = normalize_template_key(template_key)
    contents = await file.read()
    parsed_rows = parse_wbs_workbook(contents)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            rows, errors, warnings, serialized_rows, diff_rows = await prepare_template_import(
                connection,
                template_key=normalized_key,
                parsed_rows=parsed_rows,
            )
            status = "Rejected" if errors else "Accepted"
            record = await insert_import_job(
                connection,
                source_file=file.filename or "wbs-upload.xlsx",
                template_key=normalized_key,
                template_name=template_name,
                project_type=project_type,
                description=description,
                status=status,
                total_rows=len(rows),
                accepted_rows=0 if errors else len(rows),
                rejected_rows=len(errors),
                errors=errors,
                warnings=warnings,
                preview_rows=serialized_rows,
                diff_rows=diff_rows,
                applied=not errors,
            )

            if not errors:
                await replace_template_items(
                    connection,
                    template_key=normalized_key,
                    template_name=template_name,
                    project_type=project_type,
                    description=description,
                    rows=rows,
                )
                template_version = await connection.fetchval(
                    "SELECT max(version) FROM wbs_template_versions WHERE template_key = $1",
                    normalized_key,
                )
                record = await connection.fetchrow(
                    f"""
                    UPDATE wbs_import_jobs
                    SET template_version = $2
                    WHERE id = $1
                    RETURNING {IMPORT_JOB_RETURNING}
                    """,
                    record["id"],
                    template_version,
                )
            await insert_audit_event(
                connection,
                request=request,
                event_type="import.created",
                entity_type="import_job",
                entity_id=record["id"],
                summary=f"Excel import {'accepted' if not errors else 'rejected'}: {file.filename or 'wbs-upload.xlsx'}",
                metadata={
                    "template_key": normalized_key,
                    "status": status,
                    "accepted_rows": 0 if errors else len(rows),
                    "rejected_rows": len(errors),
                    "diff_count": len(diff_rows),
                    "applied": not errors,
                },
            )

            template = await fetch_template(connection, normalized_key)

    response = normalize_record(record)
    response["template"] = template
    response["warnings"] = warnings
    response["rows"] = serialized_rows[:50]
    response["diff_rows"] = diff_rows[:50]
    return response


@app.post("/api/imports/{job_id}/apply")
async def apply_import_preview(job_id: str, request: Request) -> dict[str, Any]:
    require_mutating_role(request)
    try:
        import_job_id = UUID(job_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid import job id") from exc

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            job = await fetch_import_job_for_update(connection, import_job_id)
            if not job:
                raise HTTPException(status_code=404, detail="Import job not found")
            if job["status"] != "Preview":
                raise HTTPException(status_code=409, detail="Import job is not waiting for approval")
            if job.get("errors"):
                raise HTTPException(status_code=400, detail="Rejected import cannot be applied")
            if not job.get("template_key"):
                raise HTTPException(status_code=400, detail="Import job has no template key")

            rows = restore_wbs_rows(job.get("preview_rows") or [])
            errors, _ = validate_wbs_rows(rows)
            if errors:
                raise HTTPException(status_code=400, detail={"message": "Preview rows are no longer valid", "errors": errors})

            template_key = normalize_template_key(job["template_key"])
            await replace_template_items(
                connection,
                template_key=template_key,
                template_name=job.get("template_name") or "업로드 WBS 템플릿",
                project_type=job.get("project_type") or "Uploaded",
                description=job.get("description") or "Excel 업로드로 반영된 계층형 WBS 템플릿",
                rows=rows,
            )
            template_version = await connection.fetchval(
                "SELECT max(version) FROM wbs_template_versions WHERE template_key = $1",
                template_key,
            )
            record = await connection.fetchrow(
                f"""
                UPDATE wbs_import_jobs
                SET status = 'Applied',
                    accepted_rows = total_rows,
                    rejected_rows = 0,
                    template_version = $2,
                    applied_at = now()
                WHERE id = $1
                RETURNING {IMPORT_JOB_RETURNING}
                """,
                import_job_id,
                template_version,
            )
            template = await fetch_template(connection, template_key)
            await insert_audit_event(
                connection,
                request=request,
                event_type="import.applied",
                entity_type="import_job",
                entity_id=import_job_id,
                summary=f"Excel import applied: {job['source_file']}",
                metadata={
                    "template_key": template_key,
                    "accepted_rows": record["accepted_rows"],
                },
            )

    response = normalize_record(record)
    response["template"] = template
    response["warnings"] = response.get("warnings", [])
    response["rows"] = (response.get("preview_rows") or [])[:50]
    return response


@app.post("/api/projects/{project_id}/imports/{job_id}/apply")
async def apply_import_to_project(project_id: str, job_id: str, request: Request) -> dict[str, Any]:
    parsed_project_id = safe_uuid(project_id)
    import_job_id = safe_uuid(job_id)
    if not parsed_project_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    if not import_job_id:
        raise HTTPException(status_code=400, detail="Invalid import job id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_project_id, for_update=True)
            ensure_project_mutate_role(project)
            ensure_project_status_allowed(project, PROJECT_WBS_IMPORT_ALLOWED_STATUSES, "Project WBS import")

            job = await fetch_import_job_for_update(connection, import_job_id)
            if not job:
                raise HTTPException(status_code=404, detail="Import job not found")
            if job["status"] not in {"Preview", "Applied", "Accepted"}:
                raise HTTPException(status_code=409, detail="Import job is not valid for project apply")
            if job.get("errors"):
                raise HTTPException(status_code=400, detail="Rejected import cannot be applied to project")

            rows = restore_wbs_rows(job.get("preview_rows") or [])
            errors, _ = validate_wbs_rows(rows)
            if errors:
                raise HTTPException(status_code=400, detail={"message": "Import rows are no longer valid", "errors": errors})

            existing_rows = await fetch_project_wbs_items(connection, parsed_project_id)
            if not existing_rows:
                existing_rows = await fetch_template_items(connection, project["template_key"])
            diff_rows = build_wbs_diff_rows(existing_rows, rows)
            await replace_project_wbs_items(
                connection,
                project_id=parsed_project_id,
                tenant_id=project["tenant_id"],
                rows=rows,
                source_import_job_id=import_job_id,
            )
            record = await connection.fetchrow(
                """
                UPDATE wbs_projects
                SET status = CASE WHEN status = 'Rejected' THEN 'Draft' ELSE status END,
                    updated_at = now()
                WHERE id = $1
                RETURNING id, name, template_key, owner, status, start_date, delivery_mode,
                          openproject_project_id, metadata, tenant_id, created_at, updated_at
                """,
                parsed_project_id,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="project_wbs.import_applied",
                entity_type="project",
                entity_id=parsed_project_id,
                summary=f"Project WBS import applied: {project['name']}",
                metadata={
                    "import_job_id": str(import_job_id),
                    "source_file": job.get("source_file"),
                    "rows": len(rows),
                    "diff_count": len(diff_rows),
                },
            )

    return {
        "status": "Applied",
        "project": normalize_record(record),
        "rows": [serialize_wbs_row(row) for row in rows[:50]],
        "diff_rows": diff_rows[:50],
        "summary": {"rows": len(rows), "diff_count": len(diff_rows)},
    }


@app.post("/api/templates/{template_key}/codes/resequence")
async def resequence_template_codes(template_key: str, request: Request) -> dict[str, Any]:
    require_mutating_role(request)
    normalized_key = normalize_template_key(template_key)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            template = await fetch_template(connection, normalized_key)
            if not template:
                raise HTTPException(status_code=404, detail="Template not found")

            rows = await fetch_template_items(connection, normalized_key)
            if not rows:
                raise HTTPException(status_code=404, detail="Template has no WBS rows")

            root_code = root_code_from_rows(rows) or template_code_prefix(normalized_key)
            renumbered_rows, changes = renumber_wbs_rows(rows, root_code)
            await replace_template_items(
                connection,
                template_key=normalized_key,
                template_name=template["name"],
                project_type=template["project_type"],
                description=template["description"],
                rows=renumbered_rows,
            )
            updated_template = await fetch_template(connection, normalized_key)
            await insert_audit_event(
                connection,
                request=request,
                event_type="template.resequenced",
                entity_type="template",
                entity_id=normalized_key,
                summary=f"WBS codes resequenced: {normalized_key}",
                metadata={"changed_rows": len(changes)},
            )

    return {
        "template": updated_template,
        "status": "Updated",
        "changed_rows": len(changes),
        "changes": changes[:50],
        "rows": [serialize_wbs_row(row) for row in renumbered_rows[:50]],
    }


@app.get("/api/me/work-items")
async def list_my_work_items(request: Request) -> list[dict[str, Any]]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        rows = await connection.fetch(
            """
            SELECT
                p.id AS project_id,
                p.name AS project_name,
                p.status AS project_status,
                p.template_key,
                i.code, i.parent_code, i.name, i.item_type, i.owner, i.weight,
                i.start_date, i.finish_date, i.sort_order, i.metadata, i.tenant_id
            FROM wbs_project_wbs_items i
            JOIN wbs_projects p ON p.id = i.project_id
            WHERE p.tenant_id = $1
              AND COALESCE(i.item_type, '작업') <> '프로젝트'
            ORDER BY COALESCE(i.finish_date, DATE '9999-12-31'), p.name, i.sort_order, i.code
            """,
            tid,
        )
    items = []
    for row in rows:
        item = serialize_work_item_row(row)
        if work_item_matches_user(item, user):
            items.append(item)
    return items


@app.get("/api/projects/{project_id}/work-items")
async def list_project_work_items(project_id: str, request: Request) -> list[dict[str, Any]]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        rows = await fetch_project_wbs_items(connection, parsed_id)
    return [
        serialize_work_item_row({**row, "project_id": parsed_id}, project)
        for row in rows
        if (row.get("item_type") or "작업") != "프로젝트"
    ]


@app.patch("/api/projects/{project_id}/work-items/{item_code:path}")
async def update_project_work_item(
    project_id: str,
    item_code: str,
    payload: WorkItemUpdate,
    request: Request,
) -> dict[str, Any]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            row = await fetch_project_work_item(connection, parsed_id, item_code, for_update=True)
            if not row:
                raise HTTPException(status_code=404, detail="Work item not found")
            if project["_project_role"] not in MUTATING_ROLES and not work_item_matches_user(row, user):
                raise HTTPException(status_code=403, detail="Only PMO/admin or assigned users can update this work item")

            metadata = normalize_metadata(row.get("metadata"))
            before = {
                "name": row.get("name"),
                "owner": row.get("owner"),
                "status": work_item_status(row),
                "progress": work_item_progress(row),
                "priority": metadata.get("priority") or "보통",
                "start_date": row.get("start_date"),
                "finish_date": row.get("finish_date"),
                "reviewer": metadata.get("reviewer"),
                "approver": metadata.get("approver"),
                "team": metadata.get("team"),
                "effort": metadata.get("effort"),
            }
            fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))
            next_start = payload.start_date if "start_date" in fields_set else parse_date(row.get("start_date"))
            next_finish = payload.finish_date if "finish_date" in fields_set else parse_date(row.get("finish_date"))
            if next_start and next_finish and next_finish < next_start:
                raise HTTPException(status_code=400, detail="Finish date is earlier than start date")

            if payload.status is not None:
                metadata["status"] = payload.status
            if payload.progress is not None:
                metadata["progress"] = payload.progress
            if payload.priority is not None:
                metadata["priority"] = payload.priority
            if payload.reviewer is not None:
                metadata["reviewer"] = payload.reviewer.strip()
            if payload.approver is not None:
                metadata["approver"] = payload.approver.strip()
            if payload.team is not None:
                metadata["team"] = payload.team.strip()
            if payload.effort is not None:
                metadata["effort"] = payload.effort

            actor = user.get("display_name") or user.get("email") or "사용자"
            now_iso = datetime.now(timezone.utc).isoformat()
            if payload.comment:
                append_work_item_log(metadata, "comments", {
                    "author": actor,
                    "email": user.get("email"),
                    "text": payload.comment.strip(),
                    "ts": now_iso,
                })
            if payload.attachment_name or payload.attachment_url:
                append_work_item_log(metadata, "attachments", {
                    "author": actor,
                    "name": (payload.attachment_name or payload.attachment_url or "첨부").strip(),
                    "url": (payload.attachment_url or "").strip(),
                    "ts": now_iso,
                    "kind": "link",
                })

            next_row = {
                **row,
                "name": payload.name.strip() if "name" in fields_set and payload.name is not None else row.get("name"),
                "owner": payload.owner.strip() if "owner" in fields_set and payload.owner is not None else row.get("owner"),
                "start_date": next_start,
                "finish_date": next_finish,
                "metadata": metadata,
            }
            after = {
                "name": next_row.get("name"),
                "owner": next_row.get("owner"),
                "status": work_item_status(next_row),
                "progress": work_item_progress(next_row),
                "priority": metadata.get("priority") or "보통",
                "start_date": next_start.isoformat() if isinstance(next_start, date) else next_start,
                "finish_date": next_finish.isoformat() if isinstance(next_finish, date) else next_finish,
                "reviewer": metadata.get("reviewer"),
                "approver": metadata.get("approver"),
                "team": metadata.get("team"),
                "effort": metadata.get("effort"),
            }
            summary = work_item_change_summary(before, after)
            if summary or payload.comment or payload.attachment_name or payload.attachment_url:
                append_work_item_log(metadata, "history", {
                    "actor": actor,
                    "email": user.get("email"),
                    "action": summary or ("댓글 등록" if payload.comment else "첨부/산출물 등록"),
                    "at": now_iso,
                }, limit=50)

            updated = await connection.fetchrow(
                """
                UPDATE wbs_project_wbs_items
                SET name = $3,
                    owner = $4,
                    start_date = $5,
                    finish_date = $6,
                    metadata = $7::jsonb
                WHERE project_id = $1 AND code = $2
                RETURNING project_id, code, parent_code, name, item_type, owner, weight,
                          start_date, finish_date, sort_order, metadata, tenant_id
                """,
                parsed_id,
                item_code,
                next_row.get("name"),
                next_row.get("owner") or None,
                next_start,
                next_finish,
                metadata,
            )
            await connection.execute("UPDATE wbs_projects SET updated_at = now() WHERE id = $1", parsed_id)

            # WBS 저장 시 연결된 Agile 항목 상태/담당자 즉시 반영 (source='wbs'만)
            agile_status = _wbs_to_agile_status(metadata.get("status"))
            await connection.execute(
                """
                UPDATE wbs_agile_items
                SET status = $1, assignee = $2, updated_at = now()
                WHERE project_id = $3 AND wbs_code = $4 AND source = 'wbs'
                """,
                agile_status,
                next_row.get("owner") or None,
                parsed_id,
                item_code,
            )

            await insert_audit_event(
                connection,
                request=request,
                event_type="work_item.updated",
                entity_type="work_item",
                entity_id=f"{parsed_id}:{item_code}",
                summary=f"Work item updated: {project['name']} / {item_code}",
                metadata={"changes": summary, "data_source": "internal_wbs"},
            )

            notification_count = 0
            updated_item = serialize_work_item_row(updated, project)
            if updated_item["status"] == "지연":
                notification_count += await notify_work_item_users(
                    connection,
                    tenant_id=project["tenant_id"],
                    project=project,
                    row=updated_item,
                    metadata=metadata,
                    event_type="work_item.delayed",
                    title="작업 지연 알림",
                    body=f"{project['name']} / {item_code} {updated_item['name']} 작업이 지연 상태입니다.",
                )
            if metadata.get("approver") and updated_item["status"] != "완료":
                notification_count += await notify_work_item_users(
                    connection,
                    tenant_id=project["tenant_id"],
                    project=project,
                    row=updated_item,
                    metadata=metadata,
                    event_type="work_item.approval",
                    title="승인 확인 필요",
                    body=f"{project['name']} / {item_code} 작업 항목의 승인 확인이 필요합니다.",
                    only_people=[str(metadata.get("approver"))],
                )
            if payload.comment:
                notification_count += await notify_work_item_users(
                    connection,
                    tenant_id=project["tenant_id"],
                    project=project,
                    row=updated_item,
                    metadata=metadata,
                    event_type="work_item.comment",
                    title="작업 댓글 등록",
                    body=f"{project['name']} / {item_code} 작업 항목에 댓글이 등록되었습니다.",
                )

    updated_item["notification_count"] = notification_count
    return updated_item


@app.post("/api/projects/{project_id}/work-items/{item_code:path}/attachments")
async def upload_project_work_item_attachment(
    project_id: str,
    item_code: str,
    request: Request,
    file: UploadFile = File(...),
) -> dict[str, Any]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    content = await file.read(MAX_WORK_ITEM_ATTACHMENT_BYTES + 1)
    if len(content) > MAX_WORK_ITEM_ATTACHMENT_BYTES:
        raise HTTPException(status_code=413, detail="Attachment is too large")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            row = await fetch_project_work_item(connection, parsed_id, item_code, for_update=True)
            if not row:
                raise HTTPException(status_code=404, detail="Work item not found")
            if project["_project_role"] not in MUTATING_ROLES and not work_item_matches_user(row, user):
                raise HTTPException(status_code=403, detail="Only PMO/admin or assigned users can upload attachments")

            metadata = normalize_metadata(row.get("metadata"))
            now_iso = datetime.now(timezone.utc).isoformat()
            content_type = file.content_type or "application/octet-stream"
            attachment = {
                "author": user.get("display_name") or user.get("email") or "사용자",
                "email": user.get("email"),
                "name": file.filename or "attachment",
                "size": len(content),
                "content_type": content_type,
                "url": f"data:{content_type};base64,{base64.b64encode(content).decode('ascii')}",
                "ts": now_iso,
                "kind": "file",
            }
            append_work_item_log(metadata, "attachments", attachment)
            append_work_item_log(metadata, "history", {
                "actor": attachment["author"],
                "email": user.get("email"),
                "action": f"첨부 업로드: {attachment['name']}",
                "at": now_iso,
            }, limit=50)
            updated = await connection.fetchrow(
                """
                UPDATE wbs_project_wbs_items
                SET metadata = $3::jsonb
                WHERE project_id = $1 AND code = $2
                RETURNING project_id, code, parent_code, name, item_type, owner, weight,
                          start_date, finish_date, sort_order, metadata, tenant_id
                """,
                parsed_id,
                item_code,
                metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="work_item.attachment_uploaded",
                entity_type="work_item",
                entity_id=f"{parsed_id}:{item_code}",
                summary=f"Work item attachment uploaded: {item_code}",
                metadata={"filename": attachment["name"], "size": attachment["size"], "data_source": "internal_wbs"},
            )
            await notify_work_item_users(
                connection,
                tenant_id=project["tenant_id"],
                project=project,
                row=serialize_work_item_row(updated, project),
                metadata=metadata,
                event_type="work_item.attachment",
                title="작업 첨부 등록",
                body=f"{project['name']} / {item_code} 작업 항목에 첨부가 등록되었습니다.",
            )
    return {"attachment": attachment, "item": serialize_work_item_row(updated, project)}


async def ensure_project_sprint(
    connection: asyncpg.Connection,
    *,
    sprint_id: UUID,
    project_id: UUID,
    tenant_id: str,
) -> dict[str, Any]:
    record = await connection.fetchrow(
        f"""
        SELECT {AGILE_SPRINT_SELECT}
        FROM wbs_agile_sprints
        WHERE id = $1 AND project_id = $2 AND tenant_id = $3
        """,
        sprint_id,
        project_id,
        tenant_id,
    )
    if not record:
        raise HTTPException(status_code=400, detail="Sprint is not registered in this project")
    return agile_sprint_response(record)


async def ensure_project_agile_item(
    connection: asyncpg.Connection,
    *,
    item_id: UUID,
    project_id: UUID,
    tenant_id: str,
) -> dict[str, Any]:
    record = await connection.fetchrow(
        f"""
        SELECT {AGILE_ITEM_SELECT}
        FROM wbs_agile_items
        WHERE id = $1 AND project_id = $2 AND tenant_id = $3
        """,
        item_id,
        project_id,
        tenant_id,
    )
    if not record:
        raise HTTPException(status_code=400, detail="Agile item is not registered in this project")
    return agile_item_response(record)


async def validate_agile_wbs_link(
    connection: asyncpg.Connection,
    *,
    project_id: UUID,
    tenant_id: str,
    wbs_code: str | None,
) -> str | None:
    clean_code = (wbs_code or "").strip()
    if not clean_code:
        return None
    exists = await connection.fetchval(
        """
        SELECT EXISTS (
          SELECT 1
          FROM wbs_project_wbs_items
          WHERE project_id = $1 AND tenant_id = $2 AND code = $3
        )
        """,
        project_id,
        tenant_id,
        clean_code,
    )
    if not exists:
        raise HTTPException(status_code=400, detail="WBS code is not registered in this project")
    return clean_code


@app.get("/api/projects/{project_id}/agile/sprints")
async def list_project_agile_sprints(project_id: str, request: Request) -> list[dict[str, Any]]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        rows = await connection.fetch(
            f"""
            SELECT {AGILE_SPRINT_SELECT}
            FROM wbs_agile_sprints
            WHERE project_id = $1 AND tenant_id = $2
            ORDER BY start_date, created_at
            """,
            parsed_id,
            project["tenant_id"],
        )
    return [agile_sprint_response(row) for row in rows]


@app.post("/api/projects/{project_id}/agile/sprints", status_code=201)
async def create_project_agile_sprint(
    project_id: str,
    payload: AgileSprintCreate,
    request: Request,
) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id)
            ensure_project_mutate_role(project)
            policy = await fetch_project_operation_policy(connection, project["tenant_id"])
            next_start = payload.start_date or date.today()
            next_end = sprint_end_date_from_policy(next_start, policy, payload.end_date)
            if next_end < next_start:
                raise HTTPException(status_code=400, detail="Sprint end date is earlier than start date")
            metadata = {
                **normalize_metadata(payload.metadata),
                "sprint_length_policy": policy["sprint_length_policy"],
                "openproject_version_sync": policy["openproject_sprint_version_sync"],
            }
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_agile_sprints
                  (project_id, tenant_id, name, goal, status, start_date, end_date, capacity_points, metadata)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb)
                RETURNING {AGILE_SPRINT_SELECT}
                """,
                parsed_id,
                project["tenant_id"],
                payload.name.strip(),
                (payload.goal or "").strip(),
                payload.status,
                next_start,
                next_end,
                payload.capacity_points or 0,
                metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="agile.sprint_created",
                entity_type="agile_sprint",
                entity_id=record["id"],
                summary=f"Agile sprint created: {project['name']} / {payload.name}",
                metadata={"project_id": str(parsed_id), "tenant_id": project["tenant_id"]},
            )
    return agile_sprint_response(record)


@app.patch("/api/agile/sprints/{sprint_id}")
async def update_agile_sprint(
    sprint_id: str,
    payload: AgileSprintUpdate,
    request: Request,
) -> dict[str, Any]:
    parsed_id = safe_uuid(sprint_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid sprint id")
    tid = get_tenant_id(request)
    fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            current = await connection.fetchrow(
                f"SELECT {AGILE_SPRINT_SELECT} FROM wbs_agile_sprints WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
                parsed_id,
                tid,
            )
            if not current:
                raise HTTPException(status_code=404, detail="Sprint not found")
            ensure_project_mutate_role(await fetch_tenant_project(connection, request, current["project_id"]))
            next_start = payload.start_date if "start_date" in fields_set and payload.start_date else current["start_date"]
            next_end = payload.end_date if "end_date" in fields_set and payload.end_date else current["end_date"]
            if next_end < next_start:
                raise HTTPException(status_code=400, detail="Sprint end date is earlier than start date")
            metadata = normalize_metadata(current.get("metadata"))
            if "metadata" in fields_set and payload.metadata is not None:
                metadata = {**metadata, **payload.metadata}

            record = await connection.fetchrow(
                f"""
                UPDATE wbs_agile_sprints
                SET name = $3,
                    goal = $4,
                    status = $5,
                    start_date = $6,
                    end_date = $7,
                    capacity_points = $8,
                    metadata = $9::jsonb,
                    updated_at = now()
                WHERE id = $1 AND tenant_id = $2
                RETURNING {AGILE_SPRINT_SELECT}
                """,
                parsed_id,
                tid,
                payload.name.strip() if "name" in fields_set and payload.name is not None else current["name"],
                payload.goal.strip() if "goal" in fields_set and payload.goal is not None else current["goal"],
                payload.status if "status" in fields_set and payload.status is not None else current["status"],
                next_start,
                next_end,
                payload.capacity_points if "capacity_points" in fields_set and payload.capacity_points is not None else current["capacity_points"],
                metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="agile.sprint_updated",
                entity_type="agile_sprint",
                entity_id=parsed_id,
                summary=f"Agile sprint updated: {record['name']}",
                metadata={"project_id": str(record["project_id"]), "tenant_id": tid},
            )
    return agile_sprint_response(record)


@app.get("/api/projects/{project_id}/agile/backlog")
async def list_project_agile_backlog(project_id: str, request: Request) -> list[dict[str, Any]]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        rows = await connection.fetch(
            """
            SELECT i.id, i.project_id, i.tenant_id, i.sprint_id, i.parent_id,
                   i.wbs_code, i.item_type, i.title, i.description, i.story_points,
                   i.priority, i.status, i.assignee, i.reviewer,
                   i.acceptance_criteria, i.sort_order, i.metadata,
                   i.created_at, i.updated_at,
                   s.name AS sprint_name,
                   w.name AS wbs_name
            FROM wbs_agile_items i
            LEFT JOIN wbs_agile_sprints s
              ON s.id = i.sprint_id
             AND s.tenant_id = i.tenant_id
             AND s.project_id = i.project_id
            LEFT JOIN wbs_project_wbs_items w
              ON w.project_id = i.project_id
             AND w.tenant_id = i.tenant_id
             AND w.code = i.wbs_code
            WHERE i.project_id = $1 AND i.tenant_id = $2
            ORDER BY i.sort_order, i.created_at
            """,
            parsed_id,
            project["tenant_id"],
        )
    return [agile_item_response(row) for row in rows]


@app.post("/api/projects/{project_id}/agile/backlog", status_code=201)
async def create_project_agile_item(
    project_id: str,
    payload: AgileItemCreate,
    request: Request,
) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id)
            ensure_project_mutate_role(project)
            policy = await fetch_project_operation_policy(connection, project["tenant_id"])
            if payload.sprint_id:
                await ensure_project_sprint(
                    connection,
                    sprint_id=payload.sprint_id,
                    project_id=parsed_id,
                    tenant_id=project["tenant_id"],
                )
            if payload.parent_id:
                await ensure_project_agile_item(
                    connection,
                    item_id=payload.parent_id,
                    project_id=parsed_id,
                    tenant_id=project["tenant_id"],
                )
            wbs_code = await validate_agile_wbs_link(
                connection,
                project_id=parsed_id,
                tenant_id=project["tenant_id"],
                wbs_code=payload.wbs_code,
            )
            story_points = validate_story_points_for_policy(payload.story_points, policy)
            metadata = metadata_with_policy_dod(payload.metadata, policy)
            sort_order = await connection.fetchval(
                "SELECT COALESCE(max(sort_order), 0) + 1 FROM wbs_agile_items WHERE project_id = $1 AND tenant_id = $2",
                parsed_id,
                project["tenant_id"],
            )
            record = await connection.fetchrow(
                f"""
                INSERT INTO wbs_agile_items
                  (project_id, tenant_id, sprint_id, parent_id, wbs_code, item_type,
                   title, description, story_points, priority, status, assignee,
                   reviewer, acceptance_criteria, sort_order, metadata)
                VALUES
                  ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16::jsonb)
                RETURNING {AGILE_ITEM_SELECT}
                """,
                parsed_id,
                project["tenant_id"],
                payload.sprint_id,
                payload.parent_id,
                wbs_code,
                payload.item_type,
                payload.title.strip(),
                (payload.description or "").strip(),
                story_points,
                payload.priority,
                payload.status,
                (payload.assignee or "").strip() or None,
                (payload.reviewer or "").strip() or None,
                (payload.acceptance_criteria or "").strip(),
                sort_order,
                metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="agile.item_created",
                entity_type="agile_item",
                entity_id=record["id"],
                summary=f"Agile item created: {project['name']} / {payload.title}",
                metadata={"project_id": str(parsed_id), "tenant_id": project["tenant_id"], "wbs_code": wbs_code},
            )
    return agile_item_response(record)


@app.patch("/api/agile/items/{item_id}")
async def update_agile_item(
    item_id: str,
    payload: AgileItemUpdate,
    request: Request,
) -> dict[str, Any]:
    user = require_roles(request, {"admin", "pmo", "viewer"})
    parsed_id = safe_uuid(item_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid agile item id")
    tid = get_tenant_id(request)
    fields_set = getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set()))

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            current = await connection.fetchrow(
                f"SELECT {AGILE_ITEM_SELECT} FROM wbs_agile_items WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
                parsed_id,
                tid,
            )
            if not current:
                raise HTTPException(status_code=404, detail="Agile item not found")
            current_item = agile_item_response(current)
            project = await fetch_tenant_project(connection, request, current["project_id"])
            policy = await fetch_project_operation_policy(connection, tid)
            if project["_project_role"] not in MUTATING_ROLES:
                assignee_key = work_item_identity_key(current_item.get("assignee"))
                user_keys = work_item_user_tokens(user)
                if assignee_key not in user_keys:
                    raise HTTPException(status_code=403, detail="Only PMO/admin or assigned users can update this agile item")

            sprint_id = current["sprint_id"]
            if "sprint_id" in fields_set:
                sprint_id = payload.sprint_id
                if sprint_id:
                    await ensure_project_sprint(
                        connection,
                        sprint_id=sprint_id,
                        project_id=current["project_id"],
                        tenant_id=tid,
                    )
            parent_id = current["parent_id"]
            if "parent_id" in fields_set:
                parent_id = payload.parent_id
                if parent_id:
                    if parent_id == parsed_id:
                        raise HTTPException(status_code=400, detail="Agile item cannot be its own parent")
                    await ensure_project_agile_item(
                        connection,
                        item_id=parent_id,
                        project_id=current["project_id"],
                        tenant_id=tid,
                    )
            wbs_code = current["wbs_code"]
            if "wbs_code" in fields_set:
                wbs_code = await validate_agile_wbs_link(
                    connection,
                    project_id=current["project_id"],
                    tenant_id=tid,
                    wbs_code=payload.wbs_code,
                )
            metadata = normalize_metadata(current.get("metadata"))
            if "metadata" in fields_set and payload.metadata is not None:
                metadata = {**metadata, **payload.metadata}
            next_status = payload.status if "status" in fields_set and payload.status is not None else current["status"]
            if next_status == "Done" and current["status"] != "Done":
                metadata["done_at"] = utc_now_iso()
            if next_status != "Done" and current["status"] == "Done":
                metadata.pop("done_at", None)
            next_story_points = (
                validate_story_points_for_policy(payload.story_points, policy)
                if "story_points" in fields_set and payload.story_points is not None
                else current["story_points"]
            )

            record = await connection.fetchrow(
                f"""
                UPDATE wbs_agile_items
                SET sprint_id = $3,
                    parent_id = $4,
                    wbs_code = $5,
                    item_type = $6,
                    title = $7,
                    description = $8,
                    story_points = $9,
                    priority = $10,
                    status = $11,
                    assignee = $12,
                    reviewer = $13,
                    acceptance_criteria = $14,
                    metadata = $15::jsonb,
                    updated_at = now()
                WHERE id = $1 AND tenant_id = $2
                RETURNING {AGILE_ITEM_SELECT}
                """,
                parsed_id,
                tid,
                sprint_id,
                parent_id,
                wbs_code,
                payload.item_type if "item_type" in fields_set and payload.item_type is not None else current["item_type"],
                payload.title.strip() if "title" in fields_set and payload.title is not None else current["title"],
                payload.description.strip() if "description" in fields_set and payload.description is not None else current["description"],
                next_story_points,
                payload.priority if "priority" in fields_set and payload.priority is not None else current["priority"],
                next_status,
                payload.assignee.strip() if "assignee" in fields_set and payload.assignee is not None else current["assignee"],
                payload.reviewer.strip() if "reviewer" in fields_set and payload.reviewer is not None else current["reviewer"],
                payload.acceptance_criteria.strip() if "acceptance_criteria" in fields_set and payload.acceptance_criteria is not None else current["acceptance_criteria"],
                metadata,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="agile.item_updated",
                entity_type="agile_item",
                entity_id=parsed_id,
                summary=f"Agile item updated: {record['title']}",
                metadata={"project_id": str(record["project_id"]), "tenant_id": tid, "status": next_status, "wbs_code": wbs_code},
            )
    return agile_item_response(record)


@app.delete("/api/agile/items/{item_id}")
async def delete_agile_item(item_id: str, request: Request) -> dict[str, Any]:
    parsed_id = safe_uuid(item_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid agile item id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            current = await connection.fetchrow(
                f"SELECT {AGILE_ITEM_SELECT} FROM wbs_agile_items WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
                parsed_id,
                tid,
            )
            if not current:
                raise HTTPException(status_code=404, detail="Agile item not found")
            ensure_project_mutate_role(await fetch_tenant_project(connection, request, current["project_id"]))
            record = await connection.fetchrow(
                f"DELETE FROM wbs_agile_items WHERE id = $1 AND tenant_id = $2 RETURNING {AGILE_ITEM_SELECT}",
                parsed_id,
                tid,
            )
            await insert_audit_event(
                connection,
                request=request,
                event_type="agile.item_deleted",
                entity_type="agile_item",
                entity_id=parsed_id,
                summary=f"Agile item deleted: {record['title']}",
                metadata={"project_id": str(record["project_id"]), "tenant_id": tid},
            )
    return {"status": "Deleted", "item": agile_item_response(record)}


@app.get("/api/projects/{project_id}/agile/metrics")
async def get_project_agile_metrics(project_id: str, request: Request) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        tid = project["tenant_id"]
        totals = await connection.fetchrow(
            """
            SELECT count(*)::integer AS total_items,
                   COALESCE(sum(story_points), 0) AS total_points,
                   COALESCE(sum(story_points) FILTER (WHERE status = 'Done'), 0) AS done_points,
                   count(*) FILTER (WHERE sprint_id IS NULL)::integer AS backlog_items,
                   count(*) FILTER (WHERE wbs_code IS NOT NULL AND wbs_code <> '')::integer AS hybrid_links
            FROM wbs_agile_items
            WHERE project_id = $1 AND tenant_id = $2
            """,
            parsed_id,
            tid,
        )
        status_rows = await connection.fetch(
            """
            SELECT status, count(*)::integer AS count, COALESCE(sum(story_points), 0) AS points
            FROM wbs_agile_items
            WHERE project_id = $1 AND tenant_id = $2
            GROUP BY status
            ORDER BY status
            """,
            parsed_id,
            tid,
        )
        type_rows = await connection.fetch(
            """
            SELECT item_type, count(*)::integer AS count, COALESCE(sum(story_points), 0) AS points
            FROM wbs_agile_items
            WHERE project_id = $1 AND tenant_id = $2
            GROUP BY item_type
            ORDER BY item_type
            """,
            parsed_id,
            tid,
        )
        sprint_rows = await connection.fetch(
            """
            SELECT s.id, s.name, s.status, s.start_date, s.end_date, s.capacity_points,
                   count(i.id)::integer AS item_count,
                   COALESCE(sum(i.story_points), 0) AS planned_points,
                   COALESCE(sum(i.story_points) FILTER (WHERE i.status = 'Done'), 0) AS done_points
            FROM wbs_agile_sprints s
            LEFT JOIN wbs_agile_items i
              ON i.sprint_id = s.id
             AND i.tenant_id = s.tenant_id
             AND i.project_id = s.project_id
            WHERE s.project_id = $1 AND s.tenant_id = $2
            GROUP BY s.id, s.name, s.status, s.start_date, s.end_date, s.capacity_points
            ORDER BY s.start_date, s.created_at
            """,
            parsed_id,
            tid,
        )
        wbs_rows = await connection.fetch(
            """
            SELECT i.wbs_code,
                   COALESCE(w.name, i.wbs_code) AS wbs_name,
                   count(i.id)::integer AS item_count,
                   COALESCE(sum(i.story_points), 0) AS total_points,
                   COALESCE(sum(i.story_points) FILTER (WHERE i.status = 'Done'), 0) AS done_points
            FROM wbs_agile_items i
            LEFT JOIN wbs_project_wbs_items w
              ON w.project_id = i.project_id
             AND w.tenant_id = i.tenant_id
             AND w.code = i.wbs_code
            WHERE i.project_id = $1
              AND i.tenant_id = $2
              AND i.wbs_code IS NOT NULL
              AND i.wbs_code <> ''
            GROUP BY i.wbs_code, w.name
            ORDER BY i.wbs_code
            """,
            parsed_id,
            tid,
        )

    totals_dict = normalize_record(totals) if totals else {}
    total_points = float(totals_dict.get("total_points") or 0)
    done_points = float(totals_dict.get("done_points") or 0)
    totals_dict["completion_rate"] = round(done_points / total_points * 100, 1) if total_points else 0
    velocity_history = [normalize_record(row) for row in sprint_rows]
    today = date.today()
    active_sprint = next((row for row in velocity_history if row.get("status") == "Active"), None)
    if not active_sprint:
        active_sprint = next((
            row for row in velocity_history
            if row.get("start_date") and row.get("end_date") and row["start_date"] <= today.isoformat() <= row["end_date"]
        ), None)
    burndown: dict[str, Any] | None = None
    if active_sprint:
        start = parse_date(active_sprint.get("start_date"))
        end = parse_date(active_sprint.get("end_date"))
        planned = float(active_sprint.get("planned_points") or 0)
        done = float(active_sprint.get("done_points") or 0)
        if start and end:
            days_count = max(1, (end - start).days + 1)
            days = []
            for idx in range(days_count):
                current_day = start + timedelta(days=idx)
                ideal = max(0, planned - (planned * idx / max(1, days_count - 1)))
                actual = None
                if current_day <= today:
                    actual = max(0, planned - done)
                days.append({
                    "date": current_day.isoformat(),
                    "ideal_remaining": round(ideal, 2),
                    "actual_remaining": round(actual, 2) if actual is not None else None,
                })
            burndown = {
                "sprint_id": active_sprint["id"],
                "name": active_sprint["name"],
                "planned_points": planned,
                "done_points": done,
                "remaining_points": max(0, planned - done),
                "days": days,
            }

    return {
        "project_id": str(parsed_id),
        "delivery_mode": project.get("delivery_mode") or "waterfall",
        "totals": totals_dict,
        "status_counts": [normalize_record(row) for row in status_rows],
        "type_counts": [normalize_record(row) for row in type_rows],
        "velocity_history": velocity_history,
        "active_sprint": active_sprint,
        "burndown": burndown,
        "wbs_progress": [
            {
                **normalize_record(row),
                "completion_rate": round(
                    float(row["done_points"] or 0) / float(row["total_points"] or 0) * 100,
                    1,
                ) if float(row["total_points"] or 0) else 0,
            }
            for row in wbs_rows
        ],
    }


# ── WBS → Agile 자동 동기화 ────────────────────────────────────────────────

_WBS_TO_AGILE_STATUS: dict[str, str] = {
    "대기": "Backlog", "계획": "Backlog", "보류": "Backlog",
    "시작": "In Progress", "진행중": "In Progress", "수행중": "In Progress", "지연": "In Progress",
    "검수중": "Review", "검토": "Review",
    "완료": "Done",
}

_WBS_TO_AGILE_TYPE: dict[str, str] = {
    "마일스톤": "Epic",
    "단계": "Story", "산출물": "Story", "패키지": "Story",
    "작업": "Task", "태스크": "Task", "하위작업": "Task", "서브태스크": "Task",
}


def _wbs_to_agile_status(wbs_status: str | None) -> str:
    return _WBS_TO_AGILE_STATUS.get((wbs_status or "").strip(), "Backlog")


def _wbs_to_agile_type(item_type: str | None) -> str:
    return _WBS_TO_AGILE_TYPE.get((item_type or "").strip(), "Story")


@app.post("/api/projects/{project_id}/agile/sync-from-wbs")
async def sync_agile_from_wbs(project_id: str, request: Request) -> dict[str, Any]:
    """WBS 항목을 Agile 백로그에 자동 동기화합니다.
    - source='wbs' 항목은 WBS 원본 기준으로 덮어씁니다.
    - source='manual' 항목(직접 생성)은 건드리지 않습니다.
    - sprint_id, story_points는 이미 설정된 값을 보존합니다.
    """
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        tid = project["tenant_id"]

        wbs_rows = await connection.fetch(
            """
            SELECT code, name, item_type, owner, weight, metadata
            FROM wbs_project_wbs_items
            WHERE project_id = $1
            ORDER BY sort_order
            """,
            parsed_id,
        )

        existing = await connection.fetch(
            """
            SELECT id, wbs_code, sprint_id, story_points, source
            FROM wbs_agile_items
            WHERE project_id = $1 AND tenant_id = $2
            """,
            parsed_id,
            tid,
        )
        existing_wbs: dict[str, dict] = {
            row["wbs_code"]: dict(row)
            for row in existing
            if row["wbs_code"]
        }

        created = updated = skipped = 0

        async with connection.transaction():
            for row in wbs_rows:
                code = row["code"]
                meta = normalize_metadata(row.get("metadata"))
                wbs_status = meta.get("status") or "대기"
                agile_status = _wbs_to_agile_status(wbs_status)
                agile_type = _wbs_to_agile_type(row["item_type"])
                title = (row["name"] or code).strip()
                assignee = (row["owner"] or "").strip() or None
                reviewer = (meta.get("reviewer") or "").strip() or None

                if code in existing_wbs:
                    ex = existing_wbs[code]
                    if ex["source"] == "manual":
                        skipped += 1
                        continue
                    # 기존 WBS 동기화 항목: sprint/SP 보존, 나머지 갱신
                    preserved_sp = ex["story_points"] if ex["story_points"] else (
                        float(row["weight"] or 1)
                    )
                    await connection.execute(
                        """
                        UPDATE wbs_agile_items
                        SET title = $1, item_type = $2, status = $3,
                            assignee = $4, reviewer = $5,
                            story_points = $6, updated_at = now()
                        WHERE id = $7
                        """,
                        title, agile_type, agile_status,
                        assignee, reviewer,
                        preserved_sp, ex["id"],
                    )
                    updated += 1
                else:
                    sp = float(row["weight"] or 1)
                    sort_order = await connection.fetchval(
                        "SELECT COALESCE(max(sort_order), 0) + 1 FROM wbs_agile_items WHERE project_id = $1 AND tenant_id = $2",
                        parsed_id, tid,
                    )
                    await connection.execute(
                        """
                        INSERT INTO wbs_agile_items
                          (project_id, tenant_id, wbs_code, item_type, title,
                           story_points, status, assignee, reviewer, sort_order,
                           source, metadata)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, 'wbs', $11::jsonb)
                        """,
                        parsed_id, tid, code, agile_type, title,
                        sp, agile_status, assignee, reviewer, sort_order,
                        json.dumps({"wbs_code": code, "synced": True}, ensure_ascii=False),
                    )
                    created += 1

            # WBS에서 삭제된 항목과 연결된 자동 동기화(source='wbs') Agile 항목 정리
            current_codes = [row["code"] for row in wbs_rows if row["code"]]
            deleted_rows = await connection.fetch(
                """
                DELETE FROM wbs_agile_items
                WHERE project_id = $1 AND tenant_id = $2 AND source = 'wbs'
                  AND wbs_code IS NOT NULL AND wbs_code != ALL($3::text[])
                RETURNING id
                """,
                parsed_id, tid, current_codes,
            )
            deleted = len(deleted_rows)

        await insert_audit_event(
            connection,
            request=request,
            event_type="agile.synced_from_wbs",
            entity_type="agile_items",
            entity_id=str(parsed_id),
            summary=f"WBS→Agile 동기화: {project['name']} / 생성 {created} 갱신 {updated} 보존 {skipped} 삭제 {deleted}",
            metadata={"created": created, "updated": updated, "skipped": skipped, "deleted": deleted},
        )

    return {"created": created, "updated": updated, "skipped": skipped, "deleted": deleted}


@app.post("/api/work-items/alerts/scan")
async def scan_work_item_alerts(request: Request) -> dict[str, Any]:
    require_roles(request, {"admin", "pmo"})
    tid = get_tenant_id(request)
    today_key = date.today().isoformat()
    created = 0
    scanned = 0
    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            rows = await connection.fetch(
                """
                SELECT
                    p.id AS project_id,
                    p.name AS project_name,
                    p.status AS project_status,
                    p.template_key,
                    i.code, i.parent_code, i.name, i.item_type, i.owner, i.weight,
                    i.start_date, i.finish_date, i.sort_order, i.metadata, i.tenant_id
                FROM wbs_project_wbs_items i
                JOIN wbs_projects p ON p.id = i.project_id
                WHERE p.tenant_id = $1
                  AND COALESCE(i.item_type, '작업') <> '프로젝트'
                ORDER BY p.name, i.sort_order, i.code
                """,
                tid,
            )
            for record in rows:
                scanned += 1
                row = serialize_work_item_row(record)
                metadata = normalize_metadata(row.get("metadata"))
                alerts = normalize_metadata(metadata.get("alerts"))
                project = {
                    "id": row["project_id"],
                    "name": row.get("project_name"),
                    "status": row.get("project_status"),
                    "template_key": row.get("template_key"),
                    "tenant_id": tid,
                }
                changed = False
                if row["status"] == "지연" and alerts.get("overdue") != today_key:
                    created += await notify_work_item_users(
                        connection,
                        tenant_id=tid,
                        project=project,
                        row=row,
                        metadata=metadata,
                        event_type="work_item.delayed",
                        title="작업 지연 알림",
                        body=f"{row.get('project_name')} / {row.get('code')} {row.get('name')} 작업이 지연 상태입니다.",
                    )
                    alerts["overdue"] = today_key
                    changed = True
                if metadata.get("approver") and row["status"] != "완료" and alerts.get("approval") != today_key:
                    created += await notify_work_item_users(
                        connection,
                        tenant_id=tid,
                        project=project,
                        row=row,
                        metadata=metadata,
                        event_type="work_item.approval",
                        title="승인 확인 필요",
                        body=f"{row.get('project_name')} / {row.get('code')} 작업 항목의 승인 확인이 필요합니다.",
                        only_people=[str(metadata.get("approver"))],
                    )
                    alerts["approval"] = today_key
                    changed = True
                if changed:
                    metadata["alerts"] = alerts
                    await connection.execute(
                        """
                        UPDATE wbs_project_wbs_items
                        SET metadata = $3::jsonb
                        WHERE project_id = $1 AND code = $2
                        """,
                        safe_uuid(row["project_id"]),
                        row["code"],
                        metadata,
                    )
    return {"scanned": scanned, "created": created, "data_source": "internal_wbs"}


@app.get("/api/openproject")
async def openproject_connection() -> dict[str, str]:
    return {
        "mode": "community-edition-engine",
        "base_url": OPENPROJECT_BASE_URL,
        "integration": "pm-engine-adapter",
        "adapter": "openproject",
    }


# ─── 역방향 연계: OpenProject → AX WBS ───────────────────────────

@app.post("/api/projects/{project_id}/sync-pull")
async def pull_openproject_status(project_id: str, request: Request) -> dict[str, Any]:
    """OpenProject 작업 패키지 상태·진척률을 AX WBS로 역방향 동기화 (Pull)."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        ensure_project_mutate_role(project)

    metadata = normalize_metadata(project.get("metadata"))
    engine_meta = normalize_metadata(metadata.get("pm_engine"))
    work_packages = normalize_metadata(engine_meta.get("work_packages"))
    op_project_id = project.get("openproject_project_id") or engine_meta.get("project_id")

    if not work_packages:
        raise HTTPException(
            status_code=409,
            detail="동기화된 작업 패키지가 없습니다. 먼저 AX WBS → OpenProject 동기화를 실행하세요.",
        )

    updated: list[dict[str, Any]] = []
    errors:  list[dict[str, Any]] = []
    pulled_at = utc_now_iso()

    if PM_ENGINE_ADAPTER == "mock":
        import random
        async with get_pool(request).acquire() as connection:
            async with connection.transaction():
                for code in work_packages:
                    mock_progress = random.choice([0, 10, 25, 50, 75, 90, 100])
                    mock_status   = "완료" if mock_progress == 100 else ("진행 중" if mock_progress > 0 else "신규")
                    await connection.execute(
                        """
                        UPDATE wbs_project_wbs_items
                        SET metadata = metadata || $1::jsonb, updated_at = now()
                        WHERE project_id = $2 AND code = $3
                        """,
                        {"op_status": mock_status, "op_progress": mock_progress, "op_pulled_at": pulled_at},
                        parsed_id, code,
                    )
                    updated.append({"code": code, "op_status": mock_status, "op_progress": mock_progress})
                await insert_audit_event(
                    connection, request=request,
                    event_type="pm_engine.sync_recorded",
                    entity_type="project", entity_id=parsed_id,
                    summary=f"PM engine sync pull (mock): {project['name']}",
                    metadata={"updated": len(updated), "adapter": "mock"},
                )
        return {
            "status": "Pulled", "adapter": "mock",
            "updated": len(updated), "errors": 0,
            "items": updated, "pulled_at": pulled_at,
        }

    if not OPENPROJECT_API_TOKEN:
        raise HTTPException(status_code=400, detail="OPENPROJECT_API_TOKEN이 설정되지 않았습니다.")

    client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            for code, wp_info in work_packages.items():
                if not isinstance(wp_info, dict):
                    continue
                wp_id = str(wp_info.get("id") or "").strip()
                if not wp_id:
                    continue
                try:
                    wp = await client.request("GET", f"/api/v3/work_packages/{wp_id}")
                    links       = normalize_metadata(wp.get("_links", {}))
                    status_info = normalize_metadata(links.get("status", {}))
                    op_status   = status_info.get("title", "")
                    op_progress = int(wp.get("percentageDone") or 0)
                    await connection.execute(
                        """
                        UPDATE wbs_project_wbs_items
                        SET metadata = metadata || $1::jsonb, updated_at = now()
                        WHERE project_id = $2 AND code = $3
                        """,
                        {"op_status": op_status, "op_progress": op_progress, "op_pulled_at": pulled_at},
                        parsed_id, code,
                    )
                    updated.append({"code": code, "op_status": op_status, "op_progress": op_progress})
                except Exception as exc:
                    errors.append({"code": code, "error": str(exc)[:200]})

            await insert_audit_event(
                connection, request=request,
                event_type="pm_engine.sync_recorded",
                entity_type="project", entity_id=parsed_id,
                summary=f"PM engine sync pull: {project['name']}",
                metadata={"updated": len(updated), "errors": len(errors), "source": "sync-pull"},
            )

    return {
        "status": "Pulled", "adapter": "openproject",
        "updated": len(updated), "errors": len(errors),
        "items": updated, "error_items": errors, "pulled_at": pulled_at,
    }


@app.post("/api/webhooks/openproject", status_code=200)
async def openproject_webhook(request: Request) -> dict[str, Any]:
    """
    OpenProject Webhook 실시간 수신 엔드포인트.
    OpenProject 관리 → Webhooks에서 이 URL을 등록하세요:
      http://YOUR_SERVER:8000/api/webhooks/openproject

    지원 이벤트:
      - work_package:updated  (진척률·상태 변경 시 즉시 반영)
      - work_package:created  (신규 작업패키지 생성 시)
      - work_package:closed   (작업패키지 종료 시)
    """
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    action = payload.get("action", "")
    if action not in {"work_package:updated", "work_package:created", "work_package:closed"}:
        return {"status": "ignored", "action": action}

    wp = payload.get("work_package") or {}
    if not wp:
        return {"status": "ignored", "reason": "no work_package in payload"}

    # OP 상태명 → AX WBS 표준 매핑
    _OP_STATUS_MAP = {
        "새로 만들기": "미시작", "New": "미시작",
        "In progress": "진행 중", "진행 중": "진행 중",
        "Closed": "완료", "완료": "완료",
        "Rejected": "반려", "반려": "반려",
        "On hold": "보류", "보류": "보류",
    }

    wp_id       = str(wp.get("id") or "").strip()
    links       = normalize_metadata(wp.get("_links", {}))
    raw_status  = normalize_metadata(links.get("status", {})).get("title", "")
    op_status   = _OP_STATUS_MAP.get(raw_status, raw_status)
    pct_done    = wp.get("percentageDone")
    op_progress = int(pct_done) if pct_done is not None else None
    # work_package:closed 이면 무조건 완료 처리
    if action == "work_package:closed":
        op_status   = "완료"
        op_progress = 100
    op_project_href = normalize_metadata(links.get("project", {})).get("href", "")
    op_project_id   = op_project_href.rsplit("/", 1)[-1] if "/" in op_project_href else ""

    if not op_project_id or not wp_id:
        return {"status": "ignored", "reason": "missing project or work_package id"}

    async with get_pool(request).acquire() as connection:
        project = await connection.fetchrow(
            "SELECT id, name, metadata FROM wbs_projects WHERE openproject_project_id = $1 LIMIT 1",
            op_project_id,
        )
        if not project:
            return {"status": "ignored", "reason": f"project not found for op_project_id={op_project_id}"}

        project_id   = project["id"]
        engine_meta  = normalize_metadata(normalize_metadata(project.get("metadata")).get("pm_engine"))
        work_packages_map = normalize_metadata(engine_meta.get("work_packages"))

        matched_code: str | None = None
        for code, wp_info in work_packages_map.items():
            if isinstance(wp_info, dict) and str(wp_info.get("id")) == wp_id:
                matched_code = code
                break

        if not matched_code:
            return {"status": "ignored", "reason": f"no WBS item mapped to wp_id={wp_id}"}

        pulled_at  = utc_now_iso()
        update_meta: dict[str, Any] = {
            "op_status":       op_status,
            "op_pulled_at":    pulled_at,
            "webhook_action":  action,
        }
        if op_progress is not None:
            update_meta["op_progress"] = op_progress

        await connection.execute(
            """
            UPDATE wbs_project_wbs_items
            SET metadata = metadata || $1::jsonb, updated_at = now()
            WHERE project_id = $2 AND code = $3
            """,
            update_meta,
            project_id, matched_code,
        )
        await insert_audit_event(
            connection,
            event_type="pm_engine.sync_recorded",
            entity_type="project", entity_id=project_id,
            summary=f"Webhook 실시간: {matched_code} → {op_status}"
                    + (f" ({op_progress}%)" if op_progress is not None else ""),
            metadata={
                "source":     "webhook_realtime",
                "action":     action,
                "code":       matched_code,
                "op_status":  op_status,
                "op_progress": op_progress,
            },
        )

    return {
        "status":     "ok",
        "action":     action,
        "realtime":   True,
        "project_id": str(project_id),
        "code":       matched_code,
        "op_status":  op_status,
        "op_progress": op_progress,
        "pulled_at":  pulled_at,
    }


@app.post("/api/projects/{project_id}/wbs-items")
async def save_project_wbs_items(
    project_id: str,
    payload: WbsItemsBatch,
    request: Request,
) -> dict[str, Any]:
    """WBS 항목 직접 저장 (포털 편집기용). 기존 항목 전체를 payload.rows로 교체."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    raw_rows = [row.model_dump() for row in payload.rows]
    for i, row in enumerate(raw_rows, start=1):
        row.setdefault("row_number", i)

    root_code = root_code_from_rows(raw_rows) or "WBS"
    rows = assign_missing_wbs_codes(raw_rows, root_code)
    errors, warnings = validate_wbs_rows(rows)
    if errors:
        raise HTTPException(status_code=422, detail={"message": "WBS 검증 실패", "errors": errors})

    async with get_pool(request).acquire() as connection:
        async with connection.transaction():
            project = await fetch_tenant_project(connection, request, parsed_id, for_update=True)
            ensure_project_mutate_role(project)
            ensure_project_status_allowed(
                project, PROJECT_WBS_IMPORT_ALLOWED_STATUSES, "WBS 직접 편집"
            )
            existing_rows = await fetch_project_wbs_items(connection, parsed_id)
            diff_rows = build_wbs_diff_rows(existing_rows, rows)
            await replace_project_wbs_items(
                connection,
                project_id=parsed_id,
                tenant_id=project["tenant_id"],
                rows=rows,
                source_import_job_id=None,
            )

            # ── WBS 버전 정보를 프로젝트 metadata에 기록 ──────────────
            proj_meta  = normalize_metadata(project.get("metadata"))
            wbs_history = list(proj_meta.get("wbs_history", []))

            # 현재 저장 버전 번호
            wbs_version = len(wbs_history) + 1

            # 이력 항목 추가 (최근 20건만 보관)
            actor = getattr(request.state, "user", {})
            history_entry = {
                "version":    wbs_version,
                "rows":       len(rows),
                "diff_count": len(diff_rows),
                "saved_by":   actor.get("display_name") or actor.get("email") or "PMO",
                "saved_at":   utc_now_iso(),
                "source":     payload.source,
            }
            wbs_history.append(history_entry)
            if len(wbs_history) > 20:
                wbs_history = wbs_history[-20:]

            proj_meta["wbs_version"]    = wbs_version
            proj_meta["wbs_last_saved"] = utc_now_iso()
            proj_meta["wbs_saved_by"]   = history_entry["saved_by"]
            proj_meta["wbs_rows"]       = len(rows)
            proj_meta["wbs_history"]    = wbs_history

            await connection.execute(
                """
                UPDATE wbs_projects
                SET metadata = $2::jsonb, updated_at = now()
                WHERE id = $1
                """,
                parsed_id,
                proj_meta,
            )

            diff_counts = {"added": 0, "changed": 0, "removed": 0}
            for diff_row in diff_rows:
                key = diff_row.get("change")
                if key in diff_counts:
                    diff_counts[key] += 1

            await insert_audit_event(
                connection,
                request=request,
                event_type="project_wbs.import_applied",
                entity_type="project",
                entity_id=parsed_id,
                summary=f"Project WBS import applied: {project['name']}",
                metadata={
                    "source":      payload.source,
                    "rows":        len(rows),
                    "diff_count":  len(diff_rows),
                    "wbs_version": wbs_version,
                    "added":       diff_counts["added"],
                    "changed":     diff_counts["changed"],
                    "removed":     diff_counts["removed"],
                    "diff_items": [
                        {"change": d.get("change"), "code": d.get("code"), "name": d.get("name")}
                        for d in diff_rows[:30]
                    ],
                },
            )

    return {
        "status":  "Applied",
        "rows":    [serialize_wbs_row(r) for r in rows],
        "diff_rows": diff_rows[:100],
        "summary": {
            "rows":        len(rows),
            "diff_count":  len(diff_rows),
            "warnings":    len(warnings),
            "wbs_version": wbs_version,
        },
        "warnings": warnings[:20],
        "wbs_version": wbs_version,
    }


@app.get("/api/projects/{project_id}/wbs-items")
async def list_project_wbs_items(project_id: str, request: Request) -> list[dict[str, Any]]:
    """프로젝트 WBS 항목 전체 조회."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        rows = await fetch_project_wbs_items(connection, parsed_id)
    return [serialize_wbs_row(r) for r in rows]


@app.get("/api/projects/{project_id}/op-work-packages")
async def get_op_work_packages(project_id: str, request: Request) -> dict[str, Any]:
    """OpenProject 작업패키지 조회 (AX WBS → OP API 프록시).
    OP가 오프라인이면 AX WBS의 로컬 WBS 데이터를 반환합니다."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")

    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        local_rows = await fetch_project_wbs_items(connection, parsed_id)

    metadata    = normalize_metadata(project.get("metadata"))
    engine_meta = normalize_metadata(metadata.get("pm_engine"))
    op_project_id = project.get("openproject_project_id") or engine_meta.get("project_id")
    op_identifier = engine_meta.get("project_identifier") or op_project_id

    # 로컬 데이터로 기본 응답 구성
    local_items = []
    for r in local_rows:
        m = normalize_metadata(r.get("metadata"))
        local_items.append({
            "id":           r.get("code"),
            "subject":      f"{r.get('code','')} {r.get('name','')}".strip(),
            "code":         r.get("code"),
            "name":         r.get("name"),
            "item_type":    r.get("item_type") or "작업",
            "parent_code":  r.get("parent_code"),
            "owner":        r.get("owner"),
            "weight":       r.get("weight"),
            "start_date":   r.get("start_date").isoformat() if isinstance(r.get("start_date"), date) else r.get("start_date"),
            "finish_date":  r.get("finish_date").isoformat() if isinstance(r.get("finish_date"), date) else r.get("finish_date"),
            "op_status":    m.get("op_status"),
            "op_progress":  m.get("op_progress"),
            "op_pulled_at": m.get("op_pulled_at"),
            "already_synced": r.get("code") in normalize_metadata(engine_meta.get("work_packages")),
            "source":       "local",
        })

    base_response = {
        "project": {
            "id":         str(project["id"]),
            "name":       project["name"],
            "template_key": project.get("template_key"),
            "op_project_id": op_project_id,
            "op_identifier": op_identifier,
            "op_base_url":   OPENPROJECT_BASE_URL,
        },
        "total":   len(local_items),
        "items":   local_items,
        "source":  "local",
    }

    # OP API 호출 시도
    if not OPENPROJECT_API_TOKEN or not op_project_id:
        return base_response

    try:
        client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
        op_data = await client.request(
            "GET",
            f"/api/v3/projects/{op_project_id}/work_packages?pageSize=200&sortBy=%5B%5B%22id%22%2C%22asc%22%5D%5D",
        )
        # OpenProject 상태명 → 한국어 표준 매핑
        OP_STATUS_MAP: dict[str, str] = {
            "새로 만들기": "미시작",
            "New":         "미시작",
            "In progress": "진행 중",
            "진행 중":     "진행 중",
            "Closed":      "완료",
            "완료":        "완료",
            "Rejected":    "반려",
            "반려":        "반려",
            "On hold":     "보류",
            "보류":        "보류",
        }

        op_elements = op_data.get("_embedded", {}).get("elements", [])
        op_items = []
        for wp in op_elements:
            links      = normalize_metadata(wp.get("_links", {}))
            raw_status = normalize_metadata(links.get("status", {})).get("title", "")
            op_status  = OP_STATUS_MAP.get(raw_status, raw_status)
            # percentageDone=None → null 유지 (0% 완료와 미설정 구분)
            pct_done = wp.get("percentageDone")
            op_items.append({
                "id":          wp.get("id"),
                "subject":     wp.get("subject", ""),
                "op_status":   op_status,
                "op_progress": int(pct_done) if pct_done is not None else None,
                "start_date":  wp.get("startDate"),
                "finish_date": wp.get("dueDate"),
                "op_type":     normalize_metadata(links.get("type", {})).get("title", ""),
                "assignee":    normalize_metadata(links.get("assignee", {})).get("title", ""),
                "source":      "openproject",
            })

        # OpenProject 항목에 AX WBS 가중치(weight) 보강
        local_map = {
            (str(r.get("code") or "")):
                float(r.get("weight")) if r.get("weight") is not None else None
            for r in local_rows
        }
        wp_code_map: dict[str, str] = {}
        for r in local_rows:
            eng_wps = normalize_metadata(engine_meta.get("work_packages"))
            for code, wp_info in eng_wps.items():
                if isinstance(wp_info, dict):
                    wp_code_map[str(wp_info.get("id") or "")] = code
        for item in op_items:
            code   = wp_code_map.get(str(item.get("id") or ""))
            weight = local_map.get(code) if code else None
            item["code"]   = code
            item["weight"] = weight

        return {**base_response, "items": op_items, "total": len(op_items), "source": "openproject"}
    except Exception:
        return base_response  # OP 오프라인이면 로컬 데이터 반환


@app.patch("/api/projects/{project_id}/op-work-packages/{work_package_id}/dates")
async def update_op_work_package_dates(
    project_id: str,
    work_package_id: str,
    payload: WorkPackageDateUpdate,
    request: Request,
) -> dict[str, Any]:
    """Gantt에서 조정한 날짜를 OpenProject와 AX WBS 로컬 일정에 반영."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    if payload.start_date is None and payload.finish_date is None:
        raise HTTPException(status_code=422, detail="At least one date is required")
    if payload.start_date and payload.finish_date and payload.finish_date < payload.start_date:
        raise HTTPException(status_code=422, detail="Finish date is earlier than start date")

    wp_id = str(work_package_id or "").strip()
    if not wp_id:
        raise HTTPException(status_code=400, detail="Work package id is required")

    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        ensure_project_mutate_role(project)
        local_rows = await fetch_project_wbs_items(connection, parsed_id)

    project_meta = normalize_metadata(project.get("metadata"))
    engine_meta = normalize_metadata(project_meta.get("pm_engine"))
    work_packages = normalize_metadata(engine_meta.get("work_packages"))
    op_project_id = project.get("openproject_project_id") or engine_meta.get("project_id")

    matched_code: str | None = None
    for code, wp_info in work_packages.items():
        if isinstance(wp_info, dict) and str(wp_info.get("id") or "") == wp_id:
            matched_code = str(code)
            break
    if matched_code is None and any(str(row.get("code") or "") == wp_id for row in local_rows):
        matched_code = wp_id

    if matched_code is None and PM_ENGINE_ADAPTER == "mock":
        matched_code = wp_id

    source = "local"
    openproject_item: dict[str, Any] | None = None
    if PM_ENGINE_ADAPTER != "mock" and OPENPROJECT_API_TOKEN and op_project_id:
        client = OpenProjectClient(OPENPROJECT_BASE_URL, OPENPROJECT_API_TOKEN, OPENPROJECT_AUTH_MODE)
        openproject_item = await client.update_work_package_dates(
            work_package_id=wp_id,
            start_date=payload.start_date,
            finish_date=payload.finish_date,
        )
        source = "openproject"

    pushed_at = utc_now_iso()
    updated_row: dict[str, Any] | None = None
    if matched_code:
        date_meta = {
            "op_dates_pushed_at": pushed_at,
            "op_dates_source": source,
        }
        if source == "openproject":
            date_meta["op_work_package_id"] = wp_id
        async with get_pool(request).acquire() as connection:
            async with connection.transaction():
                row = await connection.fetchrow(
                    """
                    UPDATE wbs_project_wbs_items
                    SET start_date = $3,
                        finish_date = $4,
                        metadata = metadata || $5::jsonb,
                        updated_at = now()
                    WHERE project_id = $1 AND code = $2
                    RETURNING project_id, tenant_id, code, parent_code, name, item_type, owner,
                              weight, start_date, finish_date, sort_order, metadata
                    """,
                    parsed_id,
                    matched_code,
                    payload.start_date,
                    payload.finish_date,
                    date_meta,
                )
                if row:
                    updated_row = serialize_wbs_row(row)
                    if matched_code in work_packages and isinstance(work_packages[matched_code], dict):
                        work_packages[matched_code] = {
                            **work_packages[matched_code],
                            "start_date": payload.start_date.isoformat() if payload.start_date else None,
                            "finish_date": payload.finish_date.isoformat() if payload.finish_date else None,
                            "dates_pushed_at": pushed_at,
                        }
                        engine_meta["work_packages"] = work_packages
                        project_meta["pm_engine"] = engine_meta
                        await connection.execute(
                            """
                            UPDATE wbs_projects
                            SET metadata = $2::jsonb, updated_at = now()
                            WHERE id = $1
                            """,
                            parsed_id,
                            project_meta,
                        )
                    await insert_audit_event(
                        connection,
                        request=request,
                        event_type="pm_engine.sync_recorded",
                        entity_type="project",
                        entity_id=parsed_id,
                        summary=f"Gantt date updated: {matched_code}",
                        metadata={
                            "source": source,
                            "work_package_id": wp_id,
                            "code": matched_code,
                            "start_date": payload.start_date.isoformat() if payload.start_date else None,
                            "finish_date": payload.finish_date.isoformat() if payload.finish_date else None,
                        },
                    )

    return {
        "status": "Updated",
        "source": source,
        "project_id": str(parsed_id),
        "work_package_id": wp_id,
        "code": matched_code,
        "start_date": payload.start_date.isoformat() if payload.start_date else None,
        "finish_date": payload.finish_date.isoformat() if payload.finish_date else None,
        "item": updated_row,
        "openproject_item": openproject_item,
        "pushed_at": pushed_at,
    }


# ── Word(.docx) 보고서 내보내기 ─────────────────────────────────────────────
@app.get("/api/projects/{project_id}/wbs-docx")
async def export_wbs_docx(project_id: str, request: Request) -> StreamingResponse:
    """WBS 항목을 Word(.docx) 파일로 내보내기 (python-docx 없이 XML 직접 생성)"""
    # DB에서 WBS 행 조회
    pool = request.app.state.pool
    rows: list[dict] = []
    try:
        async with pool.acquire() as conn:
            parsed_id = safe_uuid(project_id)
            if not parsed_id:
                raise HTTPException(status_code=400, detail="Invalid project id")
            await fetch_tenant_project(conn, request, parsed_id)
            db_rows = await conn.fetch(
                """
                SELECT code, name, item_type, owner, start_date, finish_date, weight, metadata, tenant_id
                FROM wbs_project_wbs_items
                WHERE project_id = $1
                ORDER BY sort_order, code
                """,
                parsed_id,
            )
            for r in db_rows:
                meta = normalize_metadata(r["metadata"])
                rows.append({
                    "code":        r["code"] or "",
                    "name":        r["name"] or "",
                    "item_type":   r["item_type"] or "작업",
                    "owner":       r["owner"] or "",
                    "status":      meta.get("status") or meta.get("op_status") or "대기",
                    "progress":    meta.get("progress") or meta.get("op_progress") or 0,
                    "start_date":  str(r["start_date"]) if r["start_date"] else "",
                    "finish_date": str(r["finish_date"]) if r["finish_date"] else "",
                    "weight":      r["weight"] or 0,
                    "reviewer":    meta.get("reviewer", ""),
                    "approver":    meta.get("approver", ""),
                })
    except HTTPException:
        raise
    except Exception:
        rows = []

    # 간단한 OOXML 생성 (python-docx 불필요)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    def xml_esc(s: str) -> str:
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def make_row(cells: list[str], header: bool = False) -> str:
        bold_open  = "<w:b/>" if header else ""
        shading    = '<w:shd w:val="clear" w:color="auto" w:fill="E8F0FE"/>' if header else ""
        tcs = ""
        for cell in cells:
            tcs += f"""<w:tc><w:tcPr>{shading}</w:tcPr><w:p><w:pPr><w:jc w:val="left"/></w:pPr>
<w:r><w:rPr>{bold_open}<w:sz w:val="18"/></w:rPr><w:t xml:space="preserve">{xml_esc(cell)}</w:t></w:r></w:p></w:tc>"""
        return f"<w:tr>{tcs}</w:tr>"

    headers = ["WBS코드", "작업명", "유형", "담당자(R)", "검토자(A)", "승인자(C)", "상태", "진행률", "시작일", "종료일", "가중치"]
    table_rows = make_row(headers, header=True)
    for r in rows:
        table_rows += make_row([
            r["code"], r["name"], r["item_type"], r["owner"],
            r["reviewer"], r["approver"], r["status"],
            f"{r['progress']}%", r["start_date"], r["finish_date"], f"{r['weight']}%"
        ])

    col_widths = "".join(f'<w:gridCol w:w="{w}"/>' for w in [900,2400,700,900,900,900,800,700,900,900,700])

    doc_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"
  xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<w:body>
  <w:p><w:pPr><w:jc w:val="center"/></w:pPr>
    <w:r><w:rPr><w:b/><w:sz w:val="32"/></w:rPr><w:t>WBS 보고서</w:t></w:r>
  </w:p>
  <w:p><w:pPr><w:jc w:val="center"/></w:pPr>
    <w:r><w:rPr><w:color w:val="666666"/><w:sz w:val="18"/></w:rPr>
      <w:t>프로젝트 ID: {xml_esc(project_id)} · 생성: {now_str}</w:t>
    </w:r>
  </w:p>
  <w:p><w:r><w:t> </w:t></w:r></w:p>
  <w:tbl>
    <w:tblPr><w:tblStyle w:val="TableGrid"/><w:tblW w:w="10200" w:type="dxa"/>
      <w:tblBorders>
        <w:top w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
        <w:left w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
        <w:bottom w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
        <w:right w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
        <w:insideH w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
        <w:insideV w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>
      </w:tblBorders>
    </w:tblPr>
    <w:tblGrid>{col_widths}</w:tblGrid>
    {table_rows}
  </w:tbl>
  <w:p><w:r><w:t> </w:t></w:r></w:p>
  <w:sectPr><w:pgSz w:w="12240" w:h="15840"/><w:pgMar w:top="720" w:right="720" w:bottom="720" w:left="720"/></w:sectPr>
</w:body>
</w:document>"""

    rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""

    doc_rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
</Relationships>"""

    content_types_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""

    import zipfile
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", rels_xml)
        zf.writestr("word/document.xml", doc_xml)
        zf.writestr("word/_rels/document.xml.rels", doc_rels_xml)
    buf.seek(0)

    fname = f"WBS_Report_{project_id}_{datetime.now().strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── P2-02: 리스크 트래킹 ──────────────────────────────────────────────────────

@app.get("/api/risks")
async def list_risks(request: Request, project_id: str | None = None, status: str | None = None) -> list[dict[str, Any]]:
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        params: list[Any] = [tid]
        conditions = ["r.tenant_id = $1"]
        if project_id:
            try:
                pid = UUID(project_id)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid project_id")
            params.append(pid)
            conditions.append(f"r.project_id = ${len(params)}")
        if status:
            params.append(status)
            conditions.append(f"r.status = ${len(params)}")
        where = " AND ".join(conditions)
        rows = await connection.fetch(
            f"""
            SELECT r.*, p.name AS project_name
            FROM wbs_risks r
            JOIN wbs_projects p
              ON p.id = r.project_id
             AND p.tenant_id = r.tenant_id
            WHERE {where}
            ORDER BY
              CASE r.severity WHEN '높음' THEN 0 WHEN '보통' THEN 1 ELSE 2 END,
              r.created_at DESC
            """,
            *params,
        )
        return [normalize_record(row) for row in rows]


@app.post("/api/projects/{project_id}/risks", status_code=201)
async def create_risk(project_id: str, payload: RiskCreate, request: Request) -> dict[str, Any]:
    try:
        pid = UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid project_id")
    due = parse_date(payload.due_date) if payload.due_date else None
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, pid)
        ensure_project_mutate_role(project)
        user = getattr(request.state, "user", None)
        row = await connection.fetchrow(
            """
            INSERT INTO wbs_risks
              (project_id, title, description, severity, likelihood, owner, mitigation, wbs_code, due_date, created_by, tenant_id)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)
            RETURNING *
            """,
            pid, payload.title, payload.description, payload.severity, payload.likelihood,
            payload.owner, payload.mitigation, payload.wbs_code, due,
            UUID(str(user["id"])) if user else None, tid,
        )
        await insert_audit_event(
            connection, request=request,
            event_type="risk.created",
            entity_type="risk", entity_id=str(row["id"]),
            summary=f"리스크 등록: {payload.title} (프로젝트: {project['name']})",
            metadata={"severity": payload.severity, "project_id": project_id},
        )
        return normalize_record(row)


@app.patch("/api/risks/{risk_id}")
async def update_risk(risk_id: str, payload: RiskUpdate, request: Request) -> dict[str, Any]:
    try:
        rid = UUID(risk_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid risk_id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        risk = await connection.fetchrow("SELECT * FROM wbs_risks WHERE id = $1 AND tenant_id = $2", rid, tid)
        if not risk:
            raise HTTPException(status_code=404, detail="Risk not found")
        ensure_project_mutate_role(await fetch_tenant_project(connection, request, risk["project_id"]))
        updates: dict[str, Any] = {}
        for field in ("title", "description", "severity", "likelihood", "status", "owner", "mitigation", "wbs_code"):
            val = getattr(payload, field)
            if val is not None:
                updates[field] = val
        if payload.due_date is not None:
            updates["due_date"] = parse_date(payload.due_date) if payload.due_date else None
        if not updates:
            return normalize_record(risk)
        updates["updated_at"] = datetime.now(timezone.utc)
        set_clause = ", ".join(f"{k} = ${i+3}" for i, k in enumerate(updates))
        row = await connection.fetchrow(
            f"UPDATE wbs_risks SET {set_clause} WHERE id = $1 AND tenant_id = $2 RETURNING *",
            rid, tid, *updates.values(),
        )
        await insert_audit_event(
            connection, request=request,
            event_type="risk.updated",
            entity_type="risk", entity_id=risk_id,
            summary=f"리스크 수정: {row['title']}",
            metadata={"changes": list(updates.keys())},
        )
        return normalize_record(row)


# ── P2-02: 이슈 트래킹 ──────────────────────────────────────────────────────

@app.get("/api/issues")
async def list_issues(request: Request, project_id: str | None = None, status: str | None = None) -> list[dict[str, Any]]:
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        params: list[Any] = [tid]
        conditions = ["i.tenant_id = $1"]
        if project_id:
            try:
                pid = UUID(project_id)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid project_id")
            params.append(pid)
            conditions.append(f"i.project_id = ${len(params)}")
        if status:
            params.append(status)
            conditions.append(f"i.status = ${len(params)}")
        where = " AND ".join(conditions)
        rows = await connection.fetch(
            f"""
            SELECT i.*, p.name AS project_name
            FROM wbs_issues i
            JOIN wbs_projects p
              ON p.id = i.project_id
             AND p.tenant_id = i.tenant_id
            WHERE {where}
            ORDER BY
              CASE i.priority WHEN '높음' THEN 0 WHEN '보통' THEN 1 ELSE 2 END,
              i.created_at DESC
            """,
            *params,
        )
        return [normalize_record(row) for row in rows]


@app.post("/api/projects/{project_id}/issues", status_code=201)
async def create_issue(project_id: str, payload: IssueCreate, request: Request) -> dict[str, Any]:
    try:
        pid = UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid project_id")
    due = parse_date(payload.due_date) if payload.due_date else None
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, pid)
        ensure_project_mutate_role(project)
        user = getattr(request.state, "user", None)
        row = await connection.fetchrow(
            """
            INSERT INTO wbs_issues
              (project_id, title, description, priority, assignee, wbs_code, due_date, created_by, tenant_id)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
            RETURNING *
            """,
            pid, payload.title, payload.description, payload.priority,
            payload.assignee, payload.wbs_code, due,
            UUID(str(user["id"])) if user else None, tid,
        )
        await insert_audit_event(
            connection, request=request,
            event_type="issue.created",
            entity_type="issue", entity_id=str(row["id"]),
            summary=f"이슈 등록: {payload.title} (프로젝트: {project['name']})",
            metadata={"priority": payload.priority, "project_id": project_id},
        )
        return normalize_record(row)


@app.patch("/api/issues/{issue_id}")
async def update_issue(issue_id: str, payload: IssueUpdate, request: Request) -> dict[str, Any]:
    try:
        iid = UUID(issue_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid issue_id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        issue = await connection.fetchrow("SELECT * FROM wbs_issues WHERE id = $1 AND tenant_id = $2", iid, tid)
        if not issue:
            raise HTTPException(status_code=404, detail="Issue not found")
        ensure_project_mutate_role(await fetch_tenant_project(connection, request, issue["project_id"]))
        updates: dict[str, Any] = {}
        for field in ("title", "description", "priority", "status", "assignee", "wbs_code"):
            val = getattr(payload, field)
            if val is not None:
                updates[field] = val
        if payload.due_date is not None:
            updates["due_date"] = parse_date(payload.due_date) if payload.due_date else None
        if payload.status in ("Resolved", "Closed") and issue["status"] not in ("Resolved", "Closed"):
            updates["resolved_at"] = datetime.now(timezone.utc)
        if not updates:
            return normalize_record(issue)
        updates["updated_at"] = datetime.now(timezone.utc)
        set_clause = ", ".join(f"{k} = ${i+3}" for i, k in enumerate(updates))
        row = await connection.fetchrow(
            f"UPDATE wbs_issues SET {set_clause} WHERE id = $1 AND tenant_id = $2 RETURNING *",
            iid, tid, *updates.values(),
        )
        await insert_audit_event(
            connection, request=request,
            event_type="issue.updated",
            entity_type="issue", entity_id=issue_id,
            summary=f"이슈 수정: {row['title']}",
            metadata={"changes": list(updates.keys())},
        )
        return normalize_record(row)


# ── P2-04: 알림 시스템 ───────────────────────────────────────────────────────

@app.get("/api/notifications")
async def list_notifications(request: Request, limit: int = 30) -> list[dict[str, Any]]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    async with get_pool(request).acquire() as connection:
        rows = await connection.fetch(
            """
            SELECT * FROM wbs_notifications
            WHERE user_id = $1
            ORDER BY created_at DESC
            LIMIT $2
            """,
            UUID(str(user["id"])), limit,
        )
        return [normalize_record(row) for row in rows]


@app.post("/api/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, request: Request) -> dict[str, Any]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        nid = UUID(notif_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid notification id")
    async with get_pool(request).acquire() as connection:
        row = await connection.fetchrow(
            "UPDATE wbs_notifications SET is_read = true WHERE id = $1 AND user_id = $2 RETURNING *",
            nid, UUID(str(user["id"])),
        )
        if not row:
            raise HTTPException(status_code=404, detail="Notification not found")
        return normalize_record(row)


@app.post("/api/notifications/read-all")
async def mark_all_notifications_read(request: Request) -> dict[str, Any]:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    async with get_pool(request).acquire() as connection:
        count = await connection.fetchval(
            "UPDATE wbs_notifications SET is_read = true WHERE user_id = $1 AND is_read = false RETURNING count(*)",
            UUID(str(user["id"])),
        ) or 0
        return {"updated": count}


# ── P2-05: 멀티 테넌시 ───────────────────────────────────────────────────────

class TenantCreate(BaseModel):
    id: str = Field(..., min_length=2, max_length=60, pattern=r"^[a-z0-9\-]+$")
    name: str = Field(..., min_length=1, max_length=100)


@app.get("/api/tenants")
async def list_tenants(request: Request) -> list[dict[str, Any]]:
    require_admin_role(request)
    async with get_pool(request).acquire() as connection:
        rows = await connection.fetch(
            """
            SELECT t.*,
                   count(DISTINCT p.id)::integer AS project_count,
                   (count(DISTINCT (i.project_id, i.code)) FILTER (WHERE i.project_id IS NOT NULL))::integer AS wbs_item_count,
                   count(DISTINCT g.id)::integer AS group_count,
                   count(DISTINCT u.id)::integer AS user_count
            FROM wbs_tenants t
            LEFT JOIN wbs_projects p ON p.tenant_id = t.id
            LEFT JOIN wbs_project_wbs_items i ON i.project_id = p.id
            LEFT JOIN wbs_user_groups g ON g.tenant_id = t.id
            LEFT JOIN wbs_users u ON u.tenant_id = t.id
            GROUP BY t.id
            ORDER BY t.created_at
            """
        )
        return [normalize_record(r) for r in rows]


@app.post("/api/tenants", status_code=201)
async def create_tenant(payload: TenantCreate, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    async with get_pool(request).acquire() as connection:
        existing = await connection.fetchrow("SELECT id FROM wbs_tenants WHERE id = $1", payload.id)
        if existing:
            raise HTTPException(status_code=409, detail="Tenant ID already exists")
        row = await connection.fetchrow(
            "INSERT INTO wbs_tenants (id, name) VALUES ($1, $2) RETURNING *",
            payload.id, payload.name,
        )
        await ensure_default_user_group(connection, payload.id)
        await connection.execute(
            "INSERT INTO wbs_project_operation_policies (tenant_id) VALUES ($1) ON CONFLICT (tenant_id) DO NOTHING",
            payload.id,
        )
        await insert_audit_event(connection, request=request,
            event_type="tenant.created", entity_type="tenant", entity_id=payload.id,
            summary=f"테넌트 생성: {payload.name}")
        return normalize_record(row)


@app.patch("/api/tenants/{tenant_id}")
async def update_tenant_status(tenant_id: str, request: Request, status: str = "Active") -> dict[str, Any]:
    require_admin_role(request)
    if status not in ("Active", "Suspended"):
        raise HTTPException(status_code=400, detail="Invalid status")
    async with get_pool(request).acquire() as connection:
        row = await connection.fetchrow(
            "UPDATE wbs_tenants SET status=$2, updated_at=now() WHERE id=$1 RETURNING *",
            tenant_id, status,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Tenant not found")
        return normalize_record(row)


# ── P2-06: SSO / LDAP 설정 ───────────────────────────────────────────────────

@app.get("/api/settings/auth")
async def get_auth_settings(request: Request) -> dict[str, Any]:
    require_admin_role(request)
    return {
        "auth_backend": AUTH_BACKEND,
        "ldap_server": LDAP_SERVER,
        "ldap_port": LDAP_PORT,
        "ldap_use_ssl": LDAP_USE_SSL,
        "ldap_bind_dn": LDAP_BIND_DN,
        "ldap_base_dn": LDAP_BASE_DN,
        "ldap_user_filter": LDAP_USER_FILTER,
        "ldap_attr_email": LDAP_ATTR_EMAIL,
        "ldap_attr_name": LDAP_ATTR_NAME,
    }


class LdapTestRequest(BaseModel):
    email: str
    password: str


class LdapDiagnosticsRequest(BaseModel):
    email: str | None = None
    password: str | None = None


@app.post("/api/settings/auth/test-ldap")
async def test_ldap_connection(payload: LdapTestRequest, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    diagnostics = await asyncio.get_event_loop().run_in_executor(
        None, _run_ldap_diagnostics, payload.email, payload.password
    )
    ok = bool(diagnostics.get("success"))
    return {
        "success": ok,
        "message": "LDAP 인증 성공" if ok else "LDAP 인증 실패 (서버 연결 또는 자격증명 확인)",
        "diagnostics": diagnostics,
    }


@app.post("/api/settings/auth/ldap-diagnostics")
async def run_ldap_diagnostics(payload: LdapDiagnosticsRequest, request: Request) -> dict[str, Any]:
    require_admin_role(request)
    diagnostics = await asyncio.get_event_loop().run_in_executor(
        None,
        _run_ldap_diagnostics,
        payload.email,
        payload.password,
    )
    return diagnostics


# ── P2-03: PMO 리포팅 ────────────────────────────────────────────────────────

@app.get("/api/reports/weekly-excel")
async def report_weekly_excel(request: Request, from_date: str | None = None, to_date: str | None = None) -> StreamingResponse:
    """PMO 주간 보고서 Excel — 프로젝트 현황·리스크·이슈·승인 이력 시트."""
    require_roles(request, {"admin", "pmo"})
    tid = get_tenant_id(request)
    today = date.today()
    dt_from = parse_date(from_date) if from_date else today - timedelta(days=7)
    dt_to   = parse_date(to_date)   if to_date   else today

    async with get_pool(request).acquire() as connection:
        projects = await connection.fetch(
            "SELECT * FROM wbs_projects WHERE tenant_id = $1 ORDER BY created_at DESC",
            tid,
        )
        approvals = await connection.fetch(
            """
            SELECT a.*, p.name AS project_name
            FROM wbs_approval_requests a
            JOIN wbs_projects p
              ON p.id = a.project_id
             AND p.tenant_id = a.tenant_id
            WHERE a.tenant_id = $1
              AND a.created_at::date BETWEEN $2 AND $3
            ORDER BY a.created_at DESC
            """,
            tid, dt_from, dt_to,
        )
        risks = await connection.fetch(
            """
            SELECT r.*, p.name AS project_name
            FROM wbs_risks r
            JOIN wbs_projects p
              ON p.id = r.project_id
             AND p.tenant_id = r.tenant_id
            WHERE r.tenant_id = $1
            ORDER BY r.created_at DESC
            """,
            tid,
        ) if await _table_exists(connection, "wbs_risks") else []
        issues = await connection.fetch(
            """
            SELECT i.*, p.name AS project_name
            FROM wbs_issues i
            JOIN wbs_projects p
              ON p.id = i.project_id
             AND p.tenant_id = i.tenant_id
            WHERE i.tenant_id = $1
            ORDER BY i.created_at DESC
            """,
            tid,
        ) if await _table_exists(connection, "wbs_issues") else []

    wb = Workbook()

    # ── 시트 1: 프로젝트 현황
    ws = wb.active
    ws.title = "프로젝트 현황"
    hdr_fill = PatternFill("solid", fgColor="0071E3")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["프로젝트명", "담당자", "유형", "상태", "시작일", "OpenProject ID"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, p in enumerate(projects, 2):
        meta = normalize_metadata(p.get("metadata") or {})
        ws.cell(row=r_idx, column=1, value=p["name"])
        ws.cell(row=r_idx, column=2, value=p["owner"])
        ws.cell(row=r_idx, column=3, value=p["template_key"])
        ws.cell(row=r_idx, column=4, value=p["status"])
        ws.cell(row=r_idx, column=5, value=str(p["start_date"]) if p.get("start_date") else "")
        ws.cell(row=r_idx, column=6, value=p.get("openproject_project_id") or "")
    for col in range(1, 7):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    # ── 시트 2: 승인 이력
    ws2 = wb.create_sheet("승인 이력")
    hdrs2 = ["프로젝트", "제목", "요청 유형", "상태", "요청자", "검토자", "요청일", "결정일"]
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, a in enumerate(approvals, 2):
        ws2.cell(row=r_idx, column=1, value=a.get("project_name") or "")
        ws2.cell(row=r_idx, column=2, value=a["title"])
        ws2.cell(row=r_idx, column=3, value=a["request_type"])
        ws2.cell(row=r_idx, column=4, value=a["status"])
        ws2.cell(row=r_idx, column=5, value=a["requester"])
        ws2.cell(row=r_idx, column=6, value=a.get("reviewer") or "")
        ws2.cell(row=r_idx, column=7, value=str(a["created_at"].date()) if a.get("created_at") else "")
        ws2.cell(row=r_idx, column=8, value=str(a["decided_at"].date()) if a.get("decided_at") else "")
    for col in range(1, 9):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 20

    # ── 시트 3: 리스크
    ws3 = wb.create_sheet("리스크")
    hdrs3 = ["프로젝트", "제목", "심각도", "발생가능성", "담당자", "상태", "목표일", "대응전략"]
    for col, h in enumerate(hdrs3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, r in enumerate(risks, 2):
        ws3.cell(row=r_idx, column=1, value=r.get("project_name") or "")
        ws3.cell(row=r_idx, column=2, value=r["title"])
        ws3.cell(row=r_idx, column=3, value=r["severity"])
        ws3.cell(row=r_idx, column=4, value=r["likelihood"])
        ws3.cell(row=r_idx, column=5, value=r["owner"])
        ws3.cell(row=r_idx, column=6, value=r["status"])
        ws3.cell(row=r_idx, column=7, value=str(r["due_date"]) if r.get("due_date") else "")
        ws3.cell(row=r_idx, column=8, value=r.get("mitigation") or "")
    for col in range(1, 9):
        ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = 20

    # ── 시트 4: 이슈
    ws4 = wb.create_sheet("이슈")
    hdrs4 = ["프로젝트", "제목", "우선순위", "담당자", "상태", "목표일", "설명"]
    for col, h in enumerate(hdrs4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, i in enumerate(issues, 2):
        ws4.cell(row=r_idx, column=1, value=i.get("project_name") or "")
        ws4.cell(row=r_idx, column=2, value=i["title"])
        ws4.cell(row=r_idx, column=3, value=i["priority"])
        ws4.cell(row=r_idx, column=4, value=i["assignee"])
        ws4.cell(row=r_idx, column=5, value=i["status"])
        ws4.cell(row=r_idx, column=6, value=str(i["due_date"]) if i.get("due_date") else "")
        ws4.cell(row=r_idx, column=7, value=i.get("description") or "")
    for col in range(1, 8):
        ws4.column_dimensions[ws4.cell(row=1, column=col).column_letter].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"PMO_Weekly_Report_{today.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type=EXCEL_MEDIA_TYPE,
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


async def _table_exists(connection: asyncpg.Connection, table_name: str) -> bool:
    result = await connection.fetchval(
        "SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_name=$1)", table_name
    )
    return bool(result)


@app.get("/api/reports/executive-summary")
async def report_executive_summary(request: Request) -> dict[str, Any]:
    """경영진 요약 — HTML 렌더링용 JSON (포털에서 PDF 인쇄)."""
    require_roles(request, {"admin", "pmo"})
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project_count = await connection.fetchval("SELECT count(*) FROM wbs_projects WHERE tenant_id = $1", tid)
        status_dist   = await connection.fetch(
            "SELECT status, count(*) AS cnt FROM wbs_projects WHERE tenant_id = $1 GROUP BY status ORDER BY cnt DESC",
            tid,
        )
        pending_count = await connection.fetchval(
            "SELECT count(*) FROM wbs_approval_requests WHERE tenant_id = $1 AND status = 'Pending'",
            tid,
        )
        risk_count = issue_count = 0
        if await _table_exists(connection, "wbs_risks"):
            risk_count  = await connection.fetchval(
                "SELECT count(*) FROM wbs_risks WHERE tenant_id = $1 AND status != 'Closed'",
                tid,
            ) or 0
            issue_count = await connection.fetchval(
                "SELECT count(*) FROM wbs_issues WHERE tenant_id = $1 AND status != 'Closed'",
                tid,
            ) or 0
        top_risks = []
        if await _table_exists(connection, "wbs_risks"):
            rows = await connection.fetch(
                """
                SELECT r.title, r.severity, r.status, p.name AS project_name
                FROM wbs_risks r
                JOIN wbs_projects p
                  ON p.id = r.project_id
                 AND p.tenant_id = r.tenant_id
                WHERE r.tenant_id = $1
                  AND r.status != 'Closed'
                  AND r.severity = '높음'
                ORDER BY r.created_at DESC
                LIMIT 5
                """,
                tid,
            )
            top_risks = [normalize_record(r) for r in rows]

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_projects": project_count,
            "pending_approvals": pending_count,
            "open_risks": risk_count,
            "open_issues": issue_count,
        },
        "status_distribution": [{"status": r["status"], "count": r["cnt"]} for r in status_dist],
        "top_risks": top_risks,
    }


# ── P3-01: 프로젝트 상세 ─────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}")
async def get_project_detail(project_id: str, request: Request) -> dict[str, Any]:
    """프로젝트 상세 — WBS 통계, 리스크/이슈 요약, 최근 승인, 최근 감사 포함."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)

        # WBS 통계 (status는 metadata->>'status'에 저장)
        wbs_stats = await connection.fetchrow(
            """
            SELECT
                count(*) AS total_items,
                count(*) FILTER (WHERE metadata->>'status' = '완료') AS done_items,
                COALESCE(SUM(COALESCE((metadata->>'weight')::numeric, weight, 0)), 0) AS total_weight,
                COALESCE(SUM(
                    COALESCE((metadata->>'weight')::numeric, weight, 0)
                    * COALESCE((metadata->>'progress')::numeric, 0) / 100.0
                ), 0) AS earned_weight
            FROM wbs_project_wbs_items
            WHERE project_id = $1 AND item_type NOT IN ('프로젝트')
            """,
            parsed_id,
        )

        # 리스크 요약
        risks = await connection.fetch(
            """
            SELECT id, title, severity, status, owner, due_date, created_at
            FROM wbs_risks WHERE project_id = $1 AND tenant_id = $2
            ORDER BY CASE severity WHEN '높음' THEN 0 WHEN '보통' THEN 1 ELSE 2 END, created_at DESC
            LIMIT 10
            """,
            parsed_id, tid,
        )

        # 이슈 요약
        issues = await connection.fetch(
            """
            SELECT id, title, priority, status, assignee, due_date, created_at
            FROM wbs_issues WHERE project_id = $1 AND tenant_id = $2
            ORDER BY CASE priority WHEN '높음' THEN 0 WHEN '보통' THEN 1 ELSE 2 END, created_at DESC
            LIMIT 10
            """,
            parsed_id, tid,
        )

        # 최근 승인 이력
        approvals = await connection.fetch(
            f"""
            SELECT {APPROVAL_SELECT}
            FROM wbs_approval_requests a
            JOIN wbs_projects p ON p.id = a.project_id
            WHERE a.project_id = $1 AND a.tenant_id = $2
            ORDER BY a.created_at DESC LIMIT 5
            """,
            parsed_id, tid,
        )

        # 베이스라인
        baseline = await fetch_latest_project_baseline(connection, parsed_id)

        # 진행중인 동기화 현황
        last_sync = await connection.fetchrow(
            """
            SELECT id, status, started_at, completed_at, metadata
            FROM wbs_sync_runs WHERE project_id = $1
            ORDER BY started_at DESC LIMIT 1
            """,
            parsed_id,
        )

    tw = float(wbs_stats["total_weight"] or 0)
    ew = float(wbs_stats["earned_weight"] or 0)
    progress_pct = round(ew / tw * 100, 1) if tw else 0
    spi = round(ew / tw, 3) if tw else None

    return {
        "project": normalize_record(project),
        "project_role": project["_project_role"],
        "wbs_summary": {
            "total_items": wbs_stats["total_items"],
            "done_items": wbs_stats["done_items"],
            "total_weight": round(tw, 2),
            "earned_weight": round(ew, 2),
            "progress_pct": progress_pct,
            "spi": spi,
        },
        "risks": [normalize_record(r) for r in risks],
        "issues": [normalize_record(i) for i in issues],
        "approvals": [normalize_record(a) for a in approvals],
        "baseline": baseline,
        "last_sync": normalize_record(last_sync) if last_sync else None,
    }


# ── P3-02: 변경 관리 고도화 ──────────────────────────────────────────────────

class ChangeRequestCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    description: str = ""
    impact_scope: str = ""          # 범위: 일정/비용/품질/인력
    impact_schedule_days: int | None = None
    impact_cost: float | None = None
    priority: str = "보통"          # 높음/보통/낮음
    wbs_code: str | None = None
    requested_by: str = ""


class ChangeRequestUpdate(BaseModel):
    title: str | None = None
    description: str | None = None
    impact_scope: str | None = None
    impact_schedule_days: int | None = None
    impact_cost: float | None = None
    priority: str | None = None
    status: str | None = None       # Open/Approved/Rejected/Withdrawn
    resolution: str | None = None


class AnnouncementCreate(BaseModel):
    title: str = Field(..., min_length=1, max_length=200)
    body: str = Field("", max_length=4000)
    project_id: str | None = None   # None = 전사 공지
    pinned: bool = False


class AnnouncementUpdate(BaseModel):
    title: str | None = Field(None, min_length=1, max_length=200)
    body: str | None = Field(None, max_length=4000)
    pinned: bool | None = None


@app.get("/api/change-requests")
async def list_all_change_requests(request: Request) -> dict[str, Any]:
    """전사(테넌트) 전체 변경요청 — 프로젝트 단위 CR + WBS 항목별 CR을 합쳐 반환."""
    tid = get_tenant_id(request)
    user = getattr(request.state, "user", None) or {}
    is_admin = user.get("role") == "admin"
    user_id = safe_uuid(user.get("id"))

    member_join = "" if is_admin else "JOIN wbs_project_members m ON m.project_id = p.id AND m.tenant_id = p.tenant_id"
    member_where = "" if is_admin else "AND m.user_id = $2"
    params: list[Any] = [tid] if is_admin else [tid, user_id]

    async with get_pool(request).acquire() as connection:
        project_cr_records = await connection.fetch(
            f"""
            SELECT cr.*, p.name AS project_name
            FROM wbs_change_requests cr
            JOIN wbs_projects p ON p.id = cr.project_id
            {member_join}
            WHERE cr.tenant_id = $1 {member_where}
            ORDER BY cr.created_at DESC
            """,
            *params,
        )
        item_records = await connection.fetch(
            f"""
            SELECT i.project_id, i.code AS wbs_code, i.name AS item_name, i.metadata, p.name AS project_name
            FROM wbs_project_wbs_items i
            JOIN wbs_projects p ON p.id = i.project_id
            {member_join}
            WHERE i.tenant_id = $1 {member_where}
              AND jsonb_array_length(COALESCE(i.metadata->'cr_list', '[]'::jsonb)) > 0
            """,
            *params,
        )

    project_crs = []
    for record in project_cr_records:
        row = normalize_record(record)
        project_crs.append({
            "source": "project",
            "id": str(row.get("id")),
            "project_id": str(row.get("project_id")),
            "project_name": row.get("project_name"),
            "wbs_code": row.get("wbs_code"),
            "version": row.get("version"),
            "title": row.get("title"),
            "status": row.get("status"),
            "priority": row.get("priority"),
            "requested_by": row.get("requested_by"),
            "created_at": row.get("created_at"),
        })

    item_crs = []
    for record in item_records:
        row = normalize_record(record)
        meta = normalize_metadata(row.get("metadata"))
        for cr in (meta.get("cr_list") or []):
            if not isinstance(cr, dict):
                continue
            item_crs.append({
                "source": "item",
                "id": cr.get("id"),
                "project_id": str(row.get("project_id")),
                "project_name": row.get("project_name"),
                "wbs_code": row.get("wbs_code"),
                "wbs_name": row.get("item_name"),
                "version": None,
                "title": cr.get("title"),
                "status": cr.get("status") or "등록",
                "priority": cr.get("impact"),
                "requested_by": cr.get("requestedBy"),
                "created_at": cr.get("date"),
            })

    return {
        "project_change_requests": project_crs,
        "item_change_requests": item_crs,
        "total": len(project_crs) + len(item_crs),
    }


# ── 공지사항(게시판형) ───────────────────────────────────────────────────
# SMTP 메일 연동과 무관하게, 등록 즉시 인앱 알림(wbs_notifications)으로 대상자에게 전달된다.
# project_id가 NULL이면 전사(테넌트 전체) 공지, 값이 있으면 해당 프로젝트 멤버 대상 공지.

@app.get("/api/announcements")
async def list_announcements(request: Request, project_id: str | None = None) -> list[dict[str, Any]]:
    tid = get_tenant_id(request)
    user = getattr(request.state, "user", None) or {}
    # admin/pmo는 테넌트 내 모든 공지(프로젝트 멤버십 무관)를 조회·관리할 수 있다.
    is_tenant_manager = user.get("role") in MUTATING_ROLES
    user_id = safe_uuid(user.get("id"))

    async with get_pool(request).acquire() as connection:
        if project_id:
            parsed_pid = safe_uuid(project_id)
            if not parsed_pid:
                raise HTTPException(status_code=400, detail="Invalid project id")
            if is_tenant_manager:
                project = await fetch_project(connection, parsed_pid, tid)
                if not project:
                    raise HTTPException(status_code=404, detail="Project not found")
            else:
                await fetch_tenant_project(connection, request, parsed_pid)
            records = await connection.fetch(
                """
                SELECT a.*, p.name AS project_name
                FROM wbs_announcements a
                LEFT JOIN wbs_projects p ON p.id = a.project_id
                WHERE a.tenant_id = $1 AND a.project_id = $2
                ORDER BY a.pinned DESC, a.created_at DESC
                """,
                tid, parsed_pid,
            )
        elif is_tenant_manager:
            records = await connection.fetch(
                """
                SELECT a.*, p.name AS project_name
                FROM wbs_announcements a
                LEFT JOIN wbs_projects p ON p.id = a.project_id
                WHERE a.tenant_id = $1
                ORDER BY a.pinned DESC, a.created_at DESC
                """,
                tid,
            )
        else:
            records = await connection.fetch(
                """
                SELECT a.*, p.name AS project_name
                FROM wbs_announcements a
                LEFT JOIN wbs_projects p ON p.id = a.project_id
                WHERE a.tenant_id = $1
                  AND (
                    a.project_id IS NULL
                    OR a.project_id IN (
                      SELECT project_id FROM wbs_project_members
                      WHERE tenant_id = $1 AND user_id = $2
                    )
                  )
                ORDER BY a.pinned DESC, a.created_at DESC
                """,
                tid, user_id,
            )

    return [normalize_record(record) for record in records]


@app.post("/api/announcements", status_code=201)
async def create_announcement(payload: AnnouncementCreate, request: Request) -> dict[str, Any]:
    user = require_mutating_role(request)
    tid = get_tenant_id(request)
    author_id = safe_uuid(user.get("id"))

    async with get_pool(request).acquire() as connection:
        project: dict[str, Any] | None = None
        parsed_pid: UUID | None = None
        if payload.project_id:
            parsed_pid = safe_uuid(payload.project_id)
            if not parsed_pid:
                raise HTTPException(status_code=400, detail="Invalid project id")
            # require_mutating_role()로 이미 테넌트 admin/pmo임이 확인됐으므로,
            # 프로젝트 멤버십 여부와 무관하게 테넌트 내 프로젝트 공지 작성을 허용한다.
            project = await fetch_project(connection, parsed_pid, tid)
            if not project:
                raise HTTPException(status_code=404, detail="Project not found")

        record = await connection.fetchrow(
            """
            INSERT INTO wbs_announcements (tenant_id, project_id, title, body, pinned, author_id, author_name)
            VALUES ($1,$2,$3,$4,$5,$6,$7)
            RETURNING *
            """,
            tid, parsed_pid, payload.title.strip(), payload.body, payload.pinned,
            author_id, user.get("display_name") or user.get("email") or "",
        )
        announcement = normalize_record(record)
        announcement["project_name"] = project.get("name") if project else None

        # 작성자 본인에게도 등록 알림을 보내 인앱 알림으로 등록 결과를 확인할 수 있게 한다.
        if parsed_pid:
            rows = await connection.fetch(
                """
                SELECT u.id
                FROM wbs_project_members m
                JOIN wbs_users u ON u.id = m.user_id
                WHERE m.tenant_id = $1 AND m.project_id = $2 AND u.status = 'Active'
                """,
                tid, parsed_pid,
            )
        else:
            rows = await connection.fetch(
                """
                SELECT id FROM wbs_users
                WHERE tenant_id = $1 AND status = 'Active'
                """,
                tid,
            )
        recipient_ids = {row["id"] for row in rows}
        if author_id:
            recipient_ids.add(author_id)

        for recipient_id in recipient_ids:
            await send_notification(
                connection,
                user_id=recipient_id,
                event_type="announcement.new",
                title=f"[공지] {announcement['title']}",
                body=announcement.get("body") or "",
                entity_type="announcement",
                entity_id=str(announcement["id"]),
                metadata={
                    "project_id": str(parsed_pid) if parsed_pid else None,
                    "project_name": announcement.get("project_name"),
                },
                email_to=None,
            )

        await insert_audit_event(
            connection,
            request=request,
            event_type="announcement.created",
            summary=f"공지 등록: {announcement['title']}",
            entity_type="announcement",
            entity_id=announcement["id"],
            metadata={"project_id": str(parsed_pid) if parsed_pid else None},
        )

    return announcement


@app.patch("/api/announcements/{announcement_id}")
async def update_announcement(announcement_id: str, payload: AnnouncementUpdate, request: Request) -> dict[str, Any]:
    user = getattr(request.state, "user", None) or {}
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    parsed_id = safe_uuid(announcement_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid announcement id")
    tid = get_tenant_id(request)

    async with get_pool(request).acquire() as connection:
        existing = await connection.fetchrow(
            "SELECT * FROM wbs_announcements WHERE id = $1 AND tenant_id = $2",
            parsed_id, tid,
        )
        if not existing:
            raise HTTPException(status_code=404, detail="Announcement not found")
        existing = normalize_record(existing)
        if user.get("role") != "admin" and str(existing.get("author_id")) != str(user.get("id")):
            raise HTTPException(status_code=403, detail="Insufficient role")

        fields_set = payload.model_fields_set
        updates: list[str] = []
        params: list[Any] = []
        if "title" in fields_set and payload.title is not None:
            params.append(payload.title.strip())
            updates.append(f"title = ${len(params)}")
        if "body" in fields_set and payload.body is not None:
            params.append(payload.body)
            updates.append(f"body = ${len(params)}")
        if "pinned" in fields_set and payload.pinned is not None:
            params.append(payload.pinned)
            updates.append(f"pinned = ${len(params)}")

        if updates:
            updates.append("updated_at = now()")
            params.append(parsed_id)
            params.append(tid)
            record = await connection.fetchrow(
                f"UPDATE wbs_announcements SET {', '.join(updates)} "
                f"WHERE id = ${len(params) - 1} AND tenant_id = ${len(params)} RETURNING *",
                *params,
            )
            existing = normalize_record(record)

        project_name = None
        if existing.get("project_id"):
            project_name = await connection.fetchval(
                "SELECT name FROM wbs_projects WHERE id = $1", safe_uuid(existing["project_id"]),
            )
        existing["project_name"] = project_name

    return existing


@app.delete("/api/announcements/{announcement_id}", status_code=204)
async def delete_announcement(announcement_id: str, request: Request) -> None:
    user = getattr(request.state, "user", None) or {}
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    parsed_id = safe_uuid(announcement_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid announcement id")
    tid = get_tenant_id(request)

    async with get_pool(request).acquire() as connection:
        existing = await connection.fetchrow(
            "SELECT author_id FROM wbs_announcements WHERE id = $1 AND tenant_id = $2",
            parsed_id, tid,
        )
        if not existing:
            raise HTTPException(status_code=404, detail="Announcement not found")
        if user.get("role") != "admin" and str(existing["author_id"]) != str(user.get("id")):
            raise HTTPException(status_code=403, detail="Insufficient role")
        await connection.execute(
            "DELETE FROM wbs_announcements WHERE id = $1 AND tenant_id = $2", parsed_id, tid,
        )
        await connection.execute(
            "DELETE FROM wbs_notifications WHERE entity_type = 'announcement' AND entity_id = $1",
            str(parsed_id),
        )


@app.get("/api/projects/{project_id}/change-requests")
async def list_change_requests(
    project_id: str, request: Request,
    status: str | None = None,
) -> list[dict[str, Any]]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        where = "project_id = $1 AND tenant_id = $2"
        params: list[Any] = [parsed_id, get_tenant_id(request)]
        if status:
            params.append(status)
            where += f" AND status = ${len(params)}"
        rows = await connection.fetch(
            f"""
            SELECT * FROM wbs_change_requests
            WHERE {where}
            ORDER BY CASE priority WHEN '높음' THEN 0 WHEN '보통' THEN 1 ELSE 2 END,
                     created_at DESC
            """,
            *params,
        )
        return [normalize_record(r) for r in rows]


@app.post("/api/projects/{project_id}/change-requests", status_code=201)
async def create_change_request(
    project_id: str, payload: ChangeRequestCreate, request: Request
) -> dict[str, Any]:
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    tid = get_tenant_id(request)
    user = getattr(request.state, "user", None)
    async with get_pool(request).acquire() as connection:
        project = await fetch_tenant_project(connection, request, parsed_id)
        ensure_project_mutate_role(project)
        # 버전 자동 증가
        last_ver = await connection.fetchval(
            "SELECT count(*) FROM wbs_change_requests WHERE project_id = $1", parsed_id
        ) or 0
        version = f"CR-{int(last_ver) + 1:03d}"
        row = await connection.fetchrow(
            """
            INSERT INTO wbs_change_requests
              (project_id, version, title, description, impact_scope,
               impact_schedule_days, impact_cost, priority, wbs_code,
               requested_by, created_by, tenant_id)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
            RETURNING *
            """,
            parsed_id, version, payload.title, payload.description,
            payload.impact_scope, payload.impact_schedule_days,
            payload.impact_cost, payload.priority, payload.wbs_code,
            payload.requested_by,
            UUID(str(user["id"])) if user else None,
            tid,
        )
        await insert_audit_event(
            connection, request=request,
            event_type="cr.created", entity_type="change_request", entity_id=str(row["id"]),
            summary=f"변경요청 생성: {version} {payload.title}",
            metadata={"project_id": project_id, "priority": payload.priority},
        )
        return normalize_record(row)


@app.patch("/api/change-requests/{cr_id}")
async def update_change_request(
    cr_id: str, payload: ChangeRequestUpdate, request: Request
) -> dict[str, Any]:
    parsed = safe_uuid(cr_id)
    if not parsed:
        raise HTTPException(status_code=400, detail="Invalid id")
    updates: dict[str, Any] = {}
    for f in ("title", "description", "impact_scope", "impact_schedule_days",
              "impact_cost", "priority", "resolution"):
        v = getattr(payload, f)
        if v is not None:
            updates[f] = v
    if payload.status is not None:
        if payload.status not in ("Open", "Approved", "Rejected", "Withdrawn"):
            raise HTTPException(status_code=400, detail="Invalid status")
        updates["status"] = payload.status
        if payload.status == "Approved":
            updates["approved_at"] = datetime.now(timezone.utc)
        elif payload.status == "Rejected":
            updates["rejected_at"] = datetime.now(timezone.utc)
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.now(timezone.utc)
    set_clause = ", ".join(f"{k} = ${i+2}" for i, k in enumerate(updates))
    params: list[Any] = [parsed, *updates.values()]
    async with get_pool(request).acquire() as connection:
        current = await connection.fetchrow(
            "SELECT project_id FROM wbs_change_requests WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
            parsed, get_tenant_id(request),
        )
        if not current:
            raise HTTPException(status_code=404, detail="Change request not found")
        ensure_project_mutate_role(await fetch_tenant_project(connection, request, current["project_id"]))
        row = await connection.fetchrow(
            f"UPDATE wbs_change_requests SET {set_clause} WHERE id = $1 AND tenant_id = ${len(params) + 1} RETURNING *",
            *params,
            get_tenant_id(request),
        )
        await insert_audit_event(
            connection, request=request,
            event_type="cr.updated", entity_type="change_request", entity_id=cr_id,
            summary=f"변경요청 수정: {row['version']} → {updates.get('status', row['status'])}",
        )
        return normalize_record(row)


# ── P3-03: WBS 버전 비교(Diff) ───────────────────────────────────────────────

@app.get("/api/projects/{project_id}/wbs-diff")
async def wbs_diff(project_id: str, request: Request, baseline_id: str | None = None) -> dict[str, Any]:
    """베이스라인 vs 현재 WBS Diff. baseline_id 미지정 시 최신 베이스라인 사용."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        # 현재 WBS
        current_rows = await connection.fetch(
            "SELECT * FROM wbs_project_wbs_items WHERE project_id = $1 ORDER BY sort_order, code",
            parsed_id,
        )
        current = {r["code"]: normalize_record(r) for r in current_rows}

        # 베이스라인 WBS
        if baseline_id:
            bl_parsed = safe_uuid(baseline_id)
            bl_row = await connection.fetchrow(
                "SELECT snapshot FROM wbs_project_baselines WHERE id = $1 AND project_id = $2",
                bl_parsed, parsed_id,
            ) if bl_parsed else None
        else:
            bl_row = await connection.fetchrow(
                "SELECT snapshot FROM wbs_project_baselines WHERE project_id = $1 ORDER BY created_at DESC LIMIT 1",
                parsed_id,
            )

        if not bl_row:
            return {"added": list(current.values()), "removed": [], "changed": [], "unchanged": []}

        baseline_snapshot = bl_row["snapshot"] if isinstance(bl_row["snapshot"], list) else json.loads(bl_row["snapshot"])
        baseline = {item["code"]: item for item in baseline_snapshot}

        added, removed, changed, unchanged = [], [], [], []
        all_codes = set(current) | set(baseline)
        for code in sorted(all_codes):
            if code in current and code not in baseline:
                added.append(current[code])
            elif code not in current and code in baseline:
                removed.append(baseline[code])
            else:
                cur = current[code]
                bas = baseline[code]
                diff_fields: dict[str, Any] = {}
                for field in ("name", "status", "owner", "weight", "sort_order", "item_type", "parent_code"):
                    cv, bv = cur.get(field), bas.get(field)
                    if str(cv) != str(bv):
                        diff_fields[field] = {"before": bv, "after": cv}
                # 진행률 비교
                cp = (cur.get("metadata") or {}).get("progress")
                bp = (bas.get("metadata") or {}).get("progress")
                if str(cp) != str(bp):
                    diff_fields["progress"] = {"before": bp, "after": cp}
                if diff_fields:
                    changed.append({"code": code, "name": cur.get("name"), "changes": diff_fields})
                else:
                    unchanged.append(code)

    return {
        "project_id": project_id,
        "baseline_id": baseline_id,
        "summary": {
            "added": len(added),
            "removed": len(removed),
            "changed": len(changed),
            "unchanged": len(unchanged),
        },
        "added": added,
        "removed": removed,
        "changed": changed,
    }


# ── P3-03b: 담당자 검증 / 일괄 매핑 ─────────────────────────────────────────

@app.get("/api/projects/{project_id}/owner-map")
async def get_owner_map(project_id: str, request: Request) -> dict[str, Any]:
    """WBS 항목의 owner/reviewer/approver를 집계하고 계정 매칭 결과를 반환합니다."""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        tenant_id = get_tenant_id(request)

        rows = await connection.fetch(
            "SELECT code, owner, metadata FROM wbs_project_wbs_items WHERE project_id = $1",
            parsed_id,
        )
        db_users = await connection.fetch(
            "SELECT id::text, email, display_name, role FROM wbs_users WHERE tenant_id = $1 AND status = 'Active'",
            tenant_id,
        )
        user_list = [dict(u) for u in db_users]
        name_idx = {resource_identity_key(u["display_name"]): u for u in user_list}
        email_idx = {resource_identity_key(u["email"] or ""): u for u in user_list if u.get("email")}

        def find_user(name: str | None) -> dict[str, Any] | None:
            if not name:
                return None
            k = resource_identity_key(name)
            return name_idx.get(k) or email_idx.get(k)

        raw_names: dict[str, dict[str, Any]] = {}

        def record(name: str | None, code: str, field: str) -> None:
            if not name or not name.strip():
                return
            n = name.strip()
            if n not in raw_names:
                u = find_user(n)
                raw_names[n] = {
                    "raw_name": n,
                    "usages": [],
                    "matched_user": {"id": u["id"], "display_name": u["display_name"], "email": u["email"]} if u else None,
                }
            raw_names[n]["usages"].append({"code": code, "field": field})

        for row in rows:
            meta = row["metadata"] or {}
            record(row["owner"], row["code"], "owner")
            if isinstance(meta, dict):
                record(meta.get("reviewer"), row["code"], "reviewer")
                record(meta.get("approver"), row["code"], "approver")

        assignees = sorted(raw_names.values(), key=lambda x: (x["matched_user"] is not None, x["raw_name"]))
        return {
            "assignees": assignees,
            "users": user_list,
        }


@app.post("/api/projects/{project_id}/owner-remap")
async def remap_owners(project_id: str, request: Request) -> dict[str, Any]:
    """담당자 이름을 일괄 교체합니다. body: { mappings: { 'old': 'new' } }"""
    parsed_id = safe_uuid(project_id)
    if not parsed_id:
        raise HTTPException(status_code=400, detail="Invalid project id")
    body = await request.json()
    mappings: dict[str, str | None] = body.get("mappings", {})
    if not mappings:
        raise HTTPException(status_code=400, detail="mappings required")

    async with get_pool(request).acquire() as connection:
        await fetch_tenant_project(connection, request, parsed_id)
        rows = await connection.fetch(
            "SELECT id, code, owner, metadata FROM wbs_project_wbs_items WHERE project_id = $1",
            parsed_id,
        )
        updated = 0
        async with connection.transaction():
            for row in rows:
                new_owner = mappings.get(row["owner"], row["owner"]) if row["owner"] else row["owner"]
                meta = row["metadata"] or {}
                new_meta: dict[str, Any] = dict(meta) if isinstance(meta, dict) else {}
                changed = new_owner != row["owner"]

                if isinstance(meta, dict):
                    old_reviewer = meta.get("reviewer")
                    old_approver = meta.get("approver")
                    if old_reviewer and old_reviewer in mappings:
                        new_meta["reviewer"] = mappings[old_reviewer]
                        changed = True
                    if old_approver and old_approver in mappings:
                        new_meta["approver"] = mappings[old_approver]
                        changed = True

                if changed:
                    await connection.execute(
                        "UPDATE wbs_project_wbs_items SET owner = $1, metadata = $2::jsonb, updated_at = now() WHERE id = $3",
                        new_owner, json.dumps(new_meta, ensure_ascii=False), row["id"],
                    )
                    updated += 1

        return {"updated": updated}


# ── P3-04: 포트폴리오 번다운 ─────────────────────────────────────────────────

@app.get("/api/reports/burndown")
async def portfolio_burndown(
    request: Request,
    from_date: str | None = None,
    to_date: str | None = None,
    project_id: str | None = None,
) -> dict[str, Any]:
    """일별 earned weight 스냅샷 기반 번다운. 스냅샷 없으면 현재값 단일 포인트 반환."""
    tid = get_tenant_id(request)
    today = date.today()
    d_from = parse_date(from_date) or (today - timedelta(days=30))
    d_to = parse_date(to_date) or today
    async with get_pool(request).acquire() as connection:
        pid_filter = ""
        params: list[Any] = [tid, d_from, d_to]
        if project_id:
            p = safe_uuid(project_id)
            if p:
                params.append(p)
                pid_filter = f" AND s.project_id = ${len(params)}"

        snapshot_rows = await connection.fetch(
            f"""
            SELECT s.snapshot_date, SUM(s.earned_weight) AS earned, SUM(s.total_weight) AS total
            FROM wbs_progress_snapshots s
            JOIN wbs_projects p ON p.id = s.project_id
            WHERE p.tenant_id = $1 AND s.snapshot_date BETWEEN $2 AND $3{pid_filter}
            GROUP BY s.snapshot_date ORDER BY s.snapshot_date
            """,
            *params,
        ) if await _table_exists(connection, "wbs_progress_snapshots") else []

        # 스냅샷 없으면 현재 값으로 단일 포인트 생성
        if not snapshot_rows:
            cur = await connection.fetchrow(
                """
                SELECT COALESCE(SUM(
                    CASE WHEN i.item_type NOT IN ('프로젝트') THEN
                        COALESCE((i.metadata->>'weight')::numeric, i.weight, 0) ELSE 0 END
                ), 0) AS total_weight,
                COALESCE(SUM(
                    CASE WHEN i.item_type NOT IN ('프로젝트') THEN
                        COALESCE((i.metadata->>'weight')::numeric, i.weight, 0)
                        * COALESCE((i.metadata->>'progress')::numeric, 0) / 100.0
                    ELSE 0 END
                ), 0) AS earned_weight
                FROM wbs_projects p
                LEFT JOIN wbs_project_wbs_items i ON i.project_id = p.id
                WHERE p.tenant_id = $1 AND p.status NOT IN ('Closed')
                """,
                tid,
            )
            snapshot_rows = [{"snapshot_date": today, "earned": cur["earned_weight"], "total": cur["total_weight"]}]

    series = [
        {
            "date": str(r["snapshot_date"]),
            "earned": round(float(r["earned"] or 0), 2),
            "total": round(float(r["total"] or 0), 2),
            "progress_pct": round(float(r["earned"] or 0) / float(r["total"] or 1) * 100, 1),
        }
        for r in snapshot_rows
    ]
    return {"from_date": str(d_from), "to_date": str(d_to), "series": series}


# ── P3-05: 자원 배분 ─────────────────────────────────────────────────────────

RESOURCE_DEFAULT_TASK_HOURS = 8.0


def resource_identity_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s._-]+", "", text)


def parse_effort_hours(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return max(float(value), 0.0)
    if isinstance(value, (int, float)):
        return max(float(value), 0.0)
    text = str(value).strip().lower().replace(",", ".")
    if not text:
        return 0.0
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return max(float(match.group(1)), 0.0) if match else 0.0


def count_weekdays(start: date, end: date) -> int:
    if end < start:
        start, end = end, start
    current = start
    total = 0
    while current <= end:
        if current.weekday() < 5:
            total += 1
        current += timedelta(days=1)
    return max(total, 1)


def utilization_state(value: float) -> str:
    if value >= 100:
        return "over"
    if value >= 80:
        return "high"
    if value >= 50:
        return "normal"
    return "low"


def empty_resource_bucket(name: str) -> dict[str, Any]:
    return {
        "assignee": name,
        "total_tasks": 0,
        "done_tasks": 0,
        "open_tasks": 0,
        "total_weight": 0.0,
        "total_effort_hours": 0.0,
        "estimated_effort_hours": 0.0,
        "inferred_effort_tasks": 0,
        "avg_progress": 0.0,
        "projects": [],
        "tasks": [],
        "_project_index": {},
        "_progress_sum": 0.0,
    }


def add_task_to_bucket(bucket: dict[str, Any], task: dict[str, Any]) -> None:
    bucket["total_tasks"] += 1
    bucket["done_tasks"] += 1 if task["done"] else 0
    bucket["open_tasks"] += 0 if task["done"] else 1
    bucket["total_weight"] += task["weight"]
    bucket["total_effort_hours"] += task["effort_hours"]
    bucket["estimated_effort_hours"] += task["estimated_effort_hours"]
    bucket["inferred_effort_tasks"] += 1 if task["effort_basis"] == "default" else 0
    bucket["_progress_sum"] += task["progress"]
    bucket["tasks"].append({
        "project_id": task["project_id"],
        "project_name": task["project_name"],
        "code": task["code"],
        "name": task["name"],
        "item_type": task["item_type"],
        "status": task["status"],
        "progress": task["progress"],
        "start_date": task["start_date"],
        "finish_date": task["finish_date"],
        "weight": task["weight"],
        "effort_hours": task["effort_hours"],
        "estimated_effort_hours": task["estimated_effort_hours"],
        "effort_basis": task["effort_basis"],
    })
    project = bucket["_project_index"].setdefault(task["project_id"], {
        "project_id": task["project_id"],
        "project_name": task["project_name"],
        "task_count": 0,
        "done_count": 0,
        "open_count": 0,
        "avg_progress": 0.0,
        "total_weight": 0.0,
        "total_effort_hours": 0.0,
        "estimated_effort_hours": 0.0,
        "_progress_sum": 0.0,
    })
    project["task_count"] += 1
    project["done_count"] += 1 if task["done"] else 0
    project["open_count"] += 0 if task["done"] else 1
    project["total_weight"] += task["weight"]
    project["total_effort_hours"] += task["effort_hours"]
    project["estimated_effort_hours"] += task["estimated_effort_hours"]
    project["_progress_sum"] += task["progress"]


def finalize_resource_bucket(bucket: dict[str, Any]) -> dict[str, Any]:
    total = max(int(bucket["total_tasks"]), 1)
    bucket["avg_progress"] = round(bucket["_progress_sum"] / total, 1)
    projects = []
    for project in bucket["_project_index"].values():
        project["avg_progress"] = round(project["_progress_sum"] / max(project["task_count"], 1), 1)
        project["total_weight"] = round(project["total_weight"], 2)
        project["total_effort_hours"] = round(project["total_effort_hours"], 1)
        project["estimated_effort_hours"] = round(project["estimated_effort_hours"], 1)
        project.pop("_progress_sum", None)
        projects.append(project)
    bucket["projects"] = sorted(projects, key=lambda item: (-item["task_count"], item["project_name"]))
    bucket["tasks"] = sorted(
        bucket["tasks"],
        key=lambda item: (item["finish_date"] or "9999-12-31", item["project_name"], item["code"] or ""),
    )
    bucket["total_weight"] = round(bucket["total_weight"], 2)
    bucket["total_effort_hours"] = round(bucket["total_effort_hours"], 1)
    bucket["estimated_effort_hours"] = round(bucket["estimated_effort_hours"], 1)
    bucket.pop("_project_index", None)
    bucket.pop("_progress_sum", None)
    return bucket

@app.get("/api/reports/resource-load")
async def resource_load(
    request: Request,
    from_date: str | None = None,
    to_date: str | None = None,
) -> dict[str, Any]:
    """담당자별 작업 부하, 사용자 계정 매핑, PMO 가동율을 함께 제공한다."""
    require_roles(request, {"admin", "pmo"})
    tid = get_tenant_id(request)
    today = date.today()
    d_from = parse_date(from_date) or date(today.year, today.month, 1)
    d_to = parse_date(to_date) or today
    if d_to < d_from:
        d_from, d_to = d_to, d_from
    workdays = count_weekdays(d_from, d_to)
    async with get_pool(request).acquire() as connection:
        users = await connection.fetch(
            """
            SELECT id, email, display_name, role, status, last_login_at
            FROM wbs_users
            WHERE tenant_id = $1
            ORDER BY CASE role WHEN 'admin' THEN 0 WHEN 'pmo' THEN 1 ELSE 2 END, email
            """,
            tid,
        )
        rows = await connection.fetch(
            """
            SELECT
                i.owner AS assignee,
                p.id AS project_id,
                p.name AS project_name,
                i.code,
                i.name AS task_name,
                i.item_type,
                i.start_date,
                i.finish_date,
                i.weight,
                i.metadata
            FROM wbs_project_wbs_items i
            JOIN wbs_projects p ON p.id = i.project_id
            WHERE p.tenant_id = $1
              AND i.item_type IN ('작업', '산출물', '마일스톤')
              AND i.owner IS NOT NULL AND i.owner <> ''
              AND (i.start_date IS NULL OR i.start_date <= $3)
              AND (i.finish_date IS NULL OR i.finish_date >= $2)
            ORDER BY i.owner, p.name, i.sort_order, i.code
            """,
            tid, d_from, d_to,
        )

    user_list = [normalize_record(user) for user in users]
    active_users = [user for user in user_list if user.get("status") == "Active"]
    pmo_users = [user for user in active_users if user.get("role") in {"admin", "pmo"}]
    fallback_pmo_user = next((user for user in active_users if user.get("role") == "pmo"), None)
    fallback_pmo_user = fallback_pmo_user or next((user for user in pmo_users), None)
    user_keys: dict[str, dict[str, Any]] = {}
    for user in active_users:
        candidates = [
            user.get("display_name"),
            user.get("email"),
            str(user.get("email") or "").partition("@")[0],
        ]
        for candidate in candidates:
            key = resource_identity_key(candidate)
            if key:
                user_keys.setdefault(key, user)

    def matched_user_for_owner(owner: str) -> dict[str, Any] | None:
        key = resource_identity_key(owner)
        if key in user_keys:
            return user_keys[key]
        if key == "pmo" and fallback_pmo_user:
            return fallback_pmo_user
        return None

    by_assignee: dict[str, dict[str, Any]] = {}
    account_buckets: dict[str, dict[str, Any]] = {}
    for user in active_users:
        uid = str(user["id"])
        account_buckets[uid] = {
            "user_id": uid,
            "email": user.get("email"),
            "display_name": user.get("display_name"),
            "role": user.get("role"),
            "status": user.get("status"),
            "last_login_at": user.get("last_login_at"),
            "matched_owners": [],
            **empty_resource_bucket(str(user.get("display_name") or user.get("email") or "")),
        }

    for r in rows:
        metadata = normalize_metadata(r["metadata"])
        status = str(metadata.get("status") or "")
        progress = parse_effort_hours(metadata.get("progress"))
        done = status == "완료" or progress >= 100
        effort_hours = parse_effort_hours(metadata.get("effort"))
        estimated_effort_hours = effort_hours if effort_hours > 0 else RESOURCE_DEFAULT_TASK_HOURS
        task = {
            "assignee": r["assignee"],
            "project_id": str(r["project_id"]),
            "project_name": r["project_name"],
            "code": r["code"],
            "name": r["task_name"],
            "item_type": r["item_type"],
            "status": status or "대기",
            "progress": min(progress, 100.0),
            "start_date": r["start_date"].isoformat() if r["start_date"] else None,
            "finish_date": r["finish_date"].isoformat() if r["finish_date"] else None,
            "weight": float(r["weight"] or 0),
            "effort_hours": effort_hours,
            "estimated_effort_hours": estimated_effort_hours,
            "effort_basis": "metadata" if effort_hours > 0 else "default",
            "done": done,
        }

        owner = task["assignee"]
        by_assignee.setdefault(owner, empty_resource_bucket(owner))
        matched_user = matched_user_for_owner(owner)
        if matched_user:
            by_assignee[owner]["matched_user"] = {
                "id": str(matched_user["id"]),
                "email": matched_user.get("email"),
                "display_name": matched_user.get("display_name"),
                "role": matched_user.get("role"),
            }
        add_task_to_bucket(by_assignee[owner], task)

        if matched_user:
            user_bucket = account_buckets[str(matched_user["id"])]
            if owner not in user_bucket["matched_owners"]:
                user_bucket["matched_owners"].append(owner)
            add_task_to_bucket(user_bucket, task)

    assignees = sorted(
        (finalize_resource_bucket(bucket) for bucket in by_assignee.values()),
        key=lambda item: (-item["total_tasks"], item["assignee"]),
    )
    account_tasks = sorted(
        (finalize_resource_bucket(bucket) for bucket in account_buckets.values()),
        key=lambda item: (-item["total_tasks"], item["role"], item["email"] or ""),
    )
    unmapped_assignments = [
        {
            "assignee": item["assignee"],
            "total_tasks": item["total_tasks"],
            "open_tasks": item["open_tasks"],
            "total_weight": item["total_weight"],
            "estimated_effort_hours": item["estimated_effort_hours"],
        }
        for item in assignees
        if not item.get("matched_user")
    ]
    account_by_user_id = {item["user_id"]: item for item in account_tasks}
    pmo_capacity = []
    capacity_hours = workdays * RESOURCE_DEFAULT_TASK_HOURS
    for user in pmo_users:
        uid = str(user["id"])
        account = account_by_user_id.get(uid) or {}
        planned_hours = float(account.get("estimated_effort_hours") or 0)
        utilization = round((planned_hours / capacity_hours) * 100, 1) if capacity_hours else 0.0
        pmo_capacity.append({
            "user_id": uid,
            "email": user.get("email"),
            "display_name": user.get("display_name"),
            "role": user.get("role"),
            "capacity_hours": round(capacity_hours, 1),
            "planned_hours": round(planned_hours, 1),
            "actual_effort_hours": round(float(account.get("total_effort_hours") or 0), 1),
            "inferred_effort_tasks": int(account.get("inferred_effort_tasks") or 0),
            "utilization_pct": utilization,
            "state": utilization_state(utilization),
            "task_count": int(account.get("total_tasks") or 0),
            "open_tasks": int(account.get("open_tasks") or 0),
        })
    summary = {
        "period_days": (d_to - d_from).days + 1,
        "workdays": workdays,
        "daily_capacity_hours": RESOURCE_DEFAULT_TASK_HOURS,
        "total_assignees": len(assignees),
        "total_tasks": sum(item["total_tasks"] for item in assignees),
        "done_tasks": sum(item["done_tasks"] for item in assignees),
        "open_tasks": sum(item["open_tasks"] for item in assignees),
        "total_weight": round(sum(float(item["total_weight"]) for item in assignees), 2),
        "total_effort_hours": round(sum(float(item["total_effort_hours"]) for item in assignees), 1),
        "estimated_effort_hours": round(sum(float(item["estimated_effort_hours"]) for item in assignees), 1),
        "mapped_accounts": sum(1 for item in account_tasks if item["total_tasks"] > 0),
        "unmapped_assignees": len(unmapped_assignments),
        "pmo_members": len(pmo_capacity),
        "pmo_capacity_hours": round(sum(item["capacity_hours"] for item in pmo_capacity), 1),
        "pmo_planned_hours": round(sum(item["planned_hours"] for item in pmo_capacity), 1),
    }
    return {
        "from_date": str(d_from),
        "to_date": str(d_to),
        "data_source": "internal_wbs",
        "summary": summary,
        "assignees": assignees,
        "account_tasks": account_tasks,
        "unmapped_assignments": unmapped_assignments,
        "pmo_capacity": sorted(pmo_capacity, key=lambda item: (-item["utilization_pct"], item["display_name"] or "")),
        "capacity_basis": {
            "daily_hours": RESOURCE_DEFAULT_TASK_HOURS,
            "workdays": workdays,
            "fallback": "공수 미입력 작업은 8h/건으로 추정",
        },
    }


# ── P3-06: 감사 로그 고도화 ──────────────────────────────────────────────────

@app.get("/api/audit-events/export.csv")
async def export_audit_csv(
    request: Request,
    event_type: str | None = None,
    entity_type: str | None = None,
    actor: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    limit: int = 5000,
) -> StreamingResponse:
    require_admin_role(request)
    tid = get_tenant_id(request)
    async with get_pool(request).acquire() as connection:
        conditions = ["tenant_id = $1"]
        params: list[Any] = [tid]
        if event_type:
            params.append(event_type); conditions.append(f"event_type = ${len(params)}")
        if entity_type:
            params.append(entity_type); conditions.append(f"entity_type = ${len(params)}")
        if actor:
            params.append(f"%{actor}%"); conditions.append(f"actor_email ILIKE ${len(params)}")
        if from_date:
            d = parse_date(from_date)
            if d:
                params.append(d); conditions.append(f"created_at >= ${len(params)}")
        if to_date:
            d = parse_date(to_date)
            if d:
                params.append(d + timedelta(days=1)); conditions.append(f"created_at < ${len(params)}")
        params.append(min(limit, 10000))
        where = " AND ".join(conditions)
        rows = await connection.fetch(
            f"""
            SELECT id, event_type, entity_type, entity_id, summary,
                   actor_email, actor_role, created_at
            FROM wbs_audit_events
            WHERE {where}
            ORDER BY created_at DESC
            LIMIT ${len(params)}
            """,
            *params,
        )
    import csv, io
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "event_type", "entity_type", "entity_id", "summary",
                     "actor_email", "actor_role", "created_at"])
    for r in rows:
        writer.writerow([
            str(r["id"]), r["event_type"], r["entity_type"] or "", r["entity_id"] or "",
            r["summary"] or "", r["actor_email"] or "", r["actor_role"] or "",
            r["created_at"].isoformat() if r["created_at"] else "",
        ])
    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=audit-{date.today()}.csv"},
    )


# ── P3-07: 백업 / 복원 ───────────────────────────────────────────────────────

def _backup_dir() -> Path:
    d = BACKUP_DIR
    try:
        d.mkdir(parents=True, exist_ok=True)
        # 쓰기 테스트
        test = d / ".write_test"
        test.touch(); test.unlink()
    except OSError:
        # 읽기 전용이면 /tmp/wbs-backups 사용
        d = Path("/tmp/wbs-backups")
        d.mkdir(parents=True, exist_ok=True)
    return d


@app.get("/api/operations/backups")
async def list_backups(request: Request) -> list[dict[str, Any]]:
    """백업 파일 목록 반환 (최신순 20개)."""
    require_admin_role(request)
    bd = _backup_dir()
    files = sorted(bd.glob("*.sql.gz"), key=lambda f: f.stat().st_mtime, reverse=True)[:20]
    result = []
    for f in files:
        st = f.stat()
        result.append({
            "filename": f.name,
            "size_bytes": st.st_size,
            "size_human": _human_size(st.st_size),
            "created_at": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
        })
    return result


def _human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


@app.post("/api/operations/backup", status_code=202)
async def trigger_backup(request: Request) -> dict[str, Any]:
    """pg_dump을 비동기 실행해 .sql.gz 파일 생성."""
    require_admin_role(request)
    bd = _backup_dir()
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    filename = f"wbs-backup-{ts}.sql.gz"
    filepath = bd / filename

    # DATABASE_URL 파싱으로 pg_dump 인자 구성
    import urllib.parse
    parsed = urllib.parse.urlparse(DATABASE_URL)
    env_vars = {
        **dict(__import__("os").environ),
        "PGPASSWORD": parsed.password or "",
    }
    cmd = [
        "pg_dump",
        "-h", parsed.hostname or "localhost",
        "-p", str(parsed.port or 5432),
        "-U", parsed.username or "wbs",
        "-d", parsed.path.lstrip("/") or "wbs_platform",
        "--no-password",
        "--format=plain",
        "--encoding=UTF8",
    ]

    async def _run_backup():
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env=env_vars,
        )
        stdout, stderr = await proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(stderr.decode()[:500])
        import gzip
        with gzip.open(filepath, "wb") as gz:
            gz.write(stdout)
        return filepath.stat().st_size

    try:
        size = await asyncio.wait_for(_run_backup(), timeout=120)
        async with get_pool(request).acquire() as conn:
            await insert_audit_event(
                conn, request=request,
                event_type="system.backup_created", entity_type="backup", entity_id=filename,
                summary=f"DB 백업 생성: {filename} ({_human_size(size)})",
            )
        return {"status": "created", "filename": filename, "size_bytes": size, "size_human": _human_size(size)}
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Backup timed out (>120s)")
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {e}")


@app.get("/api/operations/backups/{filename}")
async def download_backup(filename: str, request: Request) -> StreamingResponse:
    """백업 파일 다운로드."""
    require_admin_role(request)
    # 경로 순회 방지
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = _backup_dir() / filename
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(status_code=404, detail="Backup file not found")

    def _iter_file():
        with open(filepath, "rb") as f:
            while chunk := f.read(64 * 1024):
                yield chunk

    return StreamingResponse(
        _iter_file(),
        media_type="application/gzip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/api/operations/backups/{filename}", status_code=204)
async def delete_backup(filename: str, request: Request) -> None:
    """백업 파일 삭제."""
    require_admin_role(request)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = _backup_dir() / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")
    filepath.unlink()
    async with get_pool(request).acquire() as conn:
        await insert_audit_event(
            conn, request=request,
            event_type="system.backup_deleted", entity_type="backup", entity_id=filename,
            summary=f"DB 백업 삭제: {filename}",
        )
